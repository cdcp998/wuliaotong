"""基础资料接口测试（P1，L2 门禁）。需要本地 MySQL 已初始化。

幂等性：所有编码/名称带随机后缀，可重复执行。
"""
import io
import uuid

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.core.security import hash_password
from app.db import SessionLocal
from app.main import app
from app.models.sys import SysRole, SysRolePermission, SysUser

client = TestClient(app)
_TAG = uuid.uuid4().hex[:6]  # 本次运行的唯一后缀


def _login_admin() -> None:
    r = client.post("/api/v1/auth/login", json={"username": "admin", "password": "admin123"})
    assert r.status_code == 200 and r.json()["code"] == 0


def _make_xlsx(headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================ 分类 ============================

def test_category_tree_crud():
    _login_admin()
    tag = "类" + _TAG
    r = client.post("/api/v1/categories", json={"name": tag})
    assert r.json()["code"] == 0, r.text
    root_id = r.json()["data"]["id"]
    r = client.post("/api/v1/categories", json={"parent_id": root_id, "name": tag + "子"})
    child_id = r.json()["data"]["id"]
    assert r.json()["code"] == 0

    tree = client.get("/api/v1/categories").json()["data"]
    root = [n for n in tree if n["id"] == root_id][0]
    assert root["children"][0]["id"] == child_id

    r = client.put(f"/api/v1/categories/{root_id}", json={"name": tag + "改", "parent_id": 0})
    assert r.json()["code"] == 0

    # 有子分类不能删（HTTP 409 冲突）
    rd = client.delete(f"/api/v1/categories/{root_id}")
    assert rd.status_code == 409 and rd.json()["code"] == 4006
    # 删子分类成功
    assert client.delete(f"/api/v1/categories/{child_id}").json()["code"] == 0
    assert client.delete(f"/api/v1/categories/{root_id}").json()["code"] == 0


def test_category_three_level_and_delete_404():
    """三级体系：三级分类可创建并可挂材料；四级创建被拒；删除不存在分类返回 404/4003。"""
    _login_admin()
    tag = "三" + _TAG
    r1 = client.post("/api/v1/categories", json={"name": tag})
    assert r1.json()["code"] == 0, r1.text
    l1 = r1.json()["data"]["id"]
    r2 = client.post("/api/v1/categories", json={"parent_id": l1, "name": tag + "二"})
    l2 = r2.json()["data"]["id"]
    r3 = client.post("/api/v1/categories", json={"parent_id": l2, "name": tag + "三"})
    assert r3.json()["code"] == 0, r3.text
    l3 = r3.json()["data"]["id"]

    # 四级创建被拒（4006）
    assert client.post("/api/v1/categories", json={"parent_id": l3, "name": tag + "四"}).json()["code"] == 4006

    # 材料可挂二级/三级，挂顶级被拒（4006）
    unit = client.post("/api/v1/units", json={"name": "个" + _TAG + "x"}).json()["data"]["id"]
    p1 = client.post("/api/v1/products", json={"name": tag + "料A", "unit_id": unit, "category_id": l2}).json()["data"]
    p2 = client.post("/api/v1/products", json={"name": tag + "料B", "unit_id": unit, "category_id": l3}).json()["data"]
    assert p1["category_id"] == l2 and p2["category_id"] == l3
    assert client.post("/api/v1/products", json={"name": tag + "料C", "unit_id": unit, "category_id": l1}).json()["code"] == 4006

    # 二级分类已挂材料 → 不能再建子分类（409/4006）
    r = client.post("/api/v1/categories", json={"parent_id": l2, "name": tag + "二新"})
    assert r.status_code == 409 and r.json()["code"] == 4006, r.text

    # 分类维度商品查询附带挂载材料明细（含数量列）
    rows = client.get(f"/api/v1/products?category_id={l3}").json()["data"]
    assert rows["total"] == 1 and rows["list"][0]["id"] == p2["id"]
    assert "stock_qty" in rows["list"][0]

    # 单独改挂：取消挂载（0）+ 改挂三级；挂顶级被拒
    assert client.put(f"/api/v1/products/{p1['id']}/category", json={"category_id": 0}).json()["code"] == 0
    assert client.get(f"/api/v1/products/{p1['id']}").json()["data"]["category_id"] == 0
    assert client.put(f"/api/v1/products/{p1['id']}/category", json={"category_id": l3}).json()["code"] == 0
    assert client.put(f"/api/v1/products/{p1['id']}/category", json={"category_id": l1}).json()["code"] == 4006

    # 取消挂载后二级分类可再建子分类（规则仅约束"已挂材料时"）
    r = client.post("/api/v1/categories", json={"parent_id": l2, "name": tag + "二新"})
    assert r.json()["code"] == 0, r.text
    assert client.delete(f"/api/v1/categories/{r.json()['data']['id']}").json()["code"] == 0

    # 删除不存在的分类：HTTP 404 + code=4003
    r = client.delete("/api/v1/categories/999999999")
    assert r.status_code == 404 and r.json()["code"] == 4003 and r.json()["message"] == "分类不存在"

    # 清理（先取消挂载，再自底向上删）
    for pid in (p1["id"], p2["id"]):
        client.put(f"/api/v1/products/{pid}/category", json={"category_id": 0})
        client.delete(f"/api/v1/products/{pid}")
    assert client.delete(f"/api/v1/categories/{l3}").json()["code"] == 0
    assert client.delete(f"/api/v1/categories/{l2}").json()["code"] == 0
    assert client.delete(f"/api/v1/categories/{l1}").json()["code"] == 0


# ============================ 单位 ============================

def test_unit_crud():
    _login_admin()
    tag = "件" + _TAG + "x"  # 合法名（避免 6 位 hex 尾被垃圾名校验拦截）
    r = client.post("/api/v1/units", json={"name": tag})
    assert r.json()["code"] == 0, r.text
    unit_id = r.json()["data"]["id"]
    # 重名拦截
    assert client.post("/api/v1/units", json={"name": tag}).json()["code"] == 4006
    r = client.put(f"/api/v1/units/{unit_id}", json={"name": tag + "改"})
    assert r.json()["code"] == 0
    assert client.delete(f"/api/v1/units/{unit_id}").json()["code"] == 0


def test_department_auto_code():
    """单位编码自动生成（隐藏）：不传 code 创建 → 返回纯数字编码（yyyyMMdd + 当日序号），编辑不涉及编码。"""
    _login_admin()
    tag = "自动码" + _TAG
    r = client.post("/api/v1/departments", json={"name": tag})
    assert r.json()["code"] == 0, r.text
    code1 = r.json()["data"]["code"]
    assert code1.isdigit() and len(code1) == 12  # 8 位日期 + 4 位序号
    assert code1.startswith(__import__("datetime").date.today().strftime("%Y%m%d"))
    dept_id = r.json()["data"]["id"]

    # 再次创建：序号递增且唯一
    r2 = client.post("/api/v1/departments", json={"name": tag + "2"})
    assert r2.json()["code"] == 0, r2.text
    code2 = r2.json()["data"]["code"]
    assert code2 != code1 and code2.isdigit()

    # 列表返回的编码全为数字（自动生成，前端隐藏展示）
    lst = client.get("/api/v1/departments").json()["data"]
    assert any(x["id"] == dept_id and x["code"].isdigit() for x in lst)

    # 编辑单位不涉及编码（名称/备注可改）
    assert client.put(f"/api/v1/departments/{dept_id}", json={"name": tag + "改", "remark": "备注"}).json()["code"] == 0
    updated = [x for x in client.get("/api/v1/departments").json()["data"] if x["id"] == dept_id][0]
    assert updated["name"] == tag + "改" and updated["code"] == code1


# ============================ 供应商 ============================

def test_supplier_crud_and_import():
    _login_admin()
    code = "SUP" + _TAG
    r = client.post("/api/v1/suppliers", json={"code": code, "name": "华东五金"})
    assert r.json()["code"] == 0, r.text
    sup_id = r.json()["data"]["id"]
    assert client.post("/api/v1/suppliers", json={"code": code, "name": "重复"}).json()["code"] == 4006
    r = client.put(f"/api/v1/suppliers/{sup_id}", json={"code": code, "name": "华东五金批发"})
    assert r.json()["code"] == 0
    # 软删除（status=0）
    assert client.delete(f"/api/v1/suppliers/{sup_id}").json()["code"] == 0

    xlsx = _make_xlsx(["编码", "名称", "联系人", "电话", "地址"], [
        ["SUP2" + _TAG, "南方轴承", "王五", "13800000000", "苏州"],
        ["", "无名供应商", "", "", ""],
    ])
    r = client.post(
        "/api/v1/suppliers/import",
        files={"file": ("suppliers.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.json()["code"] == 0
    data = r.json()["data"]
    assert data["success_count"] == 2 and data["fail_rows"] == []


# ============================ 仓库/货架/库位 ============================

def test_warehouse_shelf_location_crud():
    _login_admin()
    wh_code = "CK" + _TAG
    r = client.post("/api/v1/warehouses", json={"code": wh_code, "name": "一号仓库"})
    assert r.json()["code"] == 0, r.text
    wh_id = r.json()["data"]["id"]
    assert client.post("/api/v1/warehouses", json={"code": wh_code, "name": "重复"}).json()["code"] == 4006

    r = client.post("/api/v1/shelves", json={"warehouse_id": wh_id, "code": "J01", "name": "货架一"})
    assert r.json()["code"] == 0, r.text
    shelf_id = r.json()["data"]["id"]

    r = client.post("/api/v1/locations", json={"warehouse_id": wh_id, "shelf_id": shelf_id, "layer_no": 1})
    assert r.json()["code"] == 0, r.text
    loc = r.json()["data"]
    assert loc["code"] == f"{wh_code}-J01-01"  # 自动生成库位编码

    locs = client.get(f"/api/v1/locations?warehouse_id={wh_id}").json()["data"]
    assert len(locs) == 1
    # 有货架不能删仓库
    assert client.delete(f"/api/v1/warehouses/{wh_id}").json()["code"] == 4006
    assert client.delete(f"/api/v1/locations/{loc['id']}").json()["code"] == 0
    assert client.delete(f"/api/v1/shelves/{shelf_id}").json()["code"] == 0
    assert client.delete(f"/api/v1/warehouses/{wh_id}").json()["code"] == 0


# ============================ 商品 ============================

def test_product_crud():
    _login_admin()
    unit_id = client.get("/api/v1/units").json()["data"][0]["id"]  # 复用种子单位

    code = "9" + str(int(_TAG, 16) % 10**9)
    barcode = "69" + str(int(_TAG, 16) % 10**12)  # 条码全局唯一（条码唯一性校验）
    payload = {
        "code": code, "barcode": barcode, "sku": "SKU-1", "name": "轴承6204",
        "spec": "20x12", "unit_id": unit_id, "purchase_price": "8.50",
        "min_stock": "10", "max_stock": "500",
        "units": [{"unit_id": unit_id, "rate": "1", "is_default": 1}],
    }
    r = client.post("/api/v1/products", json=payload)
    assert r.json()["code"] == 0, r.text
    pid = r.json()["data"]["id"]
    assert r.json()["data"]["purchase_price"] == "8.50"

    # 编码重复
    assert client.post("/api/v1/products", json={**payload, "code": code}).json()["code"] == 4006

    # 搜索
    r = client.get(f"/api/v1/products?keyword={code}").json()
    assert r["data"]["total"] == 1
    r = client.get(f"/api/v1/products?barcode={barcode}").json()
    assert r["data"]["total"] == 1

    # 修改
    r = client.put(f"/api/v1/products/{pid}", json={**payload, "name": "轴承6204-2RS"})
    assert r.json()["code"] == 0, r.text
    assert client.get(f"/api/v1/products/{pid}").json()["data"]["name"] == "轴承6204-2RS"

    # 软删除
    assert client.delete(f"/api/v1/products/{pid}").json()["code"] == 0
    r = client.get(f"/api/v1/products?keyword={code}").json()
    assert r["data"]["total"] == 0


def test_product_import_export():
    _login_admin()
    unit_name = "盒"  # 复用种子单位（导入命中已有单位，不新建）
    cat_name = "标准件" + _TAG
    xlsx = _make_xlsx(
        ["编码", "条码", "SKU", "名称", "分类", "规格", "单位", "进价", "下限", "上限"],
        [
            ["9" + str(int(_TAG, 16) % 10**9 + 1), "6900000000002", "", "螺丝M6", cat_name, "30mm", unit_name, "2.50", "100", "5000"],
            ["9" + str(int(_TAG, 16) % 10**9 + 2), "", "", "", cat_name, "", unit_name, "1.00", "0", "0"],  # 名称为空 → 失败行
        ],
    )
    r = client.post(
        "/api/v1/products/import",
        files={"file": ("products.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.json()["code"] == 0, r.text
    data = r.json()["data"]
    assert data["success_count"] == 1 and data["fail_rows"] and data["fail_rows"][0]["row"] == 3

    # 导入的分类/单位自动创建
    cats = client.get("/api/v1/categories").json()["data"]
    assert any(c["name"] == cat_name for c in cats)

    # 模板
    r = client.get("/api/v1/products/import-template")
    assert r.status_code == 200 and "spreadsheet" in r.headers["content-type"]

    # 导出
    r = client.get("/api/v1/products/export")
    assert r.status_code == 200 and "spreadsheet" in r.headers["content-type"]
    wb = __import__("openpyxl").load_workbook(io.BytesIO(r.content))
    rows = list(wb.active.iter_rows(values_only=True))
    assert len(rows) >= 2  # 表头 + 至少 1 行


# ============================ 权限校验 ============================

def test_permission_denied():
    # 新建一个"使用者"角色账号（无 base:product 写权限）
    db = SessionLocal()
    try:
        if not db.query(SysUser).filter(SysUser.username == "tester_user").first():
            db.add(SysUser(username="tester_user", password_hash=hash_password("123456"), real_name="测试使用者", role_id=4))
            db.commit()
    finally:
        db.close()

    c = TestClient(app)
    r = c.post("/api/v1/auth/login", json={"username": "tester_user", "password": "123456"})
    assert r.json()["code"] == 0
    # 无 base:product 写权限 → 403；查询对所有登录用户开放（领用选商品需要）
    assert c.post("/api/v1/products", json={"code": "X1", "name": "x", "unit_id": 1}).status_code == 403
    assert c.get("/api/v1/products").status_code == 200
    # 未登录 → 401
    assert TestClient(app).get("/api/v1/products").status_code == 401


def test_category_write_with_ai_suggestion_perm():
    """AI 建议处理员（仅 ai:suggestion，无 base:category）可在确认新增材料时维护分类：
    可新增/编辑分类，删除仍要求 base:category（403）。"""
    tag = "AI" + _TAG
    db = SessionLocal()
    try:
        role = SysRole(code=f"ai_only_{tag}", name="AI处理员(测试)", description="test", is_builtin=0)
        db.add(role)
        db.flush()
        db.add(SysRolePermission(role_id=role.id, permission_id=26))  # ai:suggestion
        db.add(SysUser(username=f"ai_worker_{tag}", password_hash=hash_password("123456"),
                       real_name="AI处理员(测试)", role_id=role.id))
        db.commit()
    finally:
        db.close()

    c = TestClient(app)
    r = c.post("/api/v1/auth/login", json={"username": f"ai_worker_{tag}", "password": "123456"})
    assert r.status_code == 200 and r.json()["code"] == 0

    name = f"AI内联分类{tag}"
    r = c.post("/api/v1/categories", json={"parent_id": 0, "name": name, "sort": 0})
    assert r.status_code == 200 and r.json()["code"] == 0, r.text
    cid = r.json()["data"]["id"]
    assert c.put(f"/api/v1/categories/{cid}", json={"parent_id": 0, "name": name + "改", "sort": 1}).json()["code"] == 0
    # 删除仍要求 base:category → 403
    assert c.delete(f"/api/v1/categories/{cid}").status_code == 403

    # 清理：admin 删除测试分类
    _login_admin()
    assert client.delete(f"/api/v1/categories/{cid}").status_code == 200
