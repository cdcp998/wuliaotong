"""报表与看板接口（《后端API设计.md》§6 /stock/summary、§8 报表）。

报表全部基于 stk_stock_log 聚合（唯一事实来源）：
- 期初 = 期间起点前的累计净变动；入库 = change_qty>0 之和；出库 = |change_qty<0| 之和；结存 = 期初+入库-出库
- 预警判定与 scheduler.scan_stock_alerts 一致：min_stock/max_stock 非 0 才生效
"""
from __future__ import annotations

import io
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from urllib.parse import quote

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins
from sqlalchemy import and_, case, func, or_, select, tuple_
from sqlalchemy.orm import Session

from app.core.cache import cache_aside_json
from app.core.deps import get_current_user, require_permission
from app.core.excel_guard import safe_excel_value
from app.core.response import BizError, E_PARAM, ok
from app.db import get_db
from app.models.advanced import StkCheck, StkTransfer
from app.models.base import BaseCategory, BaseLocation, BaseProduct, BaseUnit, BaseWarehouse
from app.models.requisition import REQ_STATUS_PENDING, OutRequisition
from app.models.stock import StkStock, StkStockLog
from app.schemas.stock import PageData

router = APIRouter(tags=["报表"], dependencies=[Depends(get_current_user)])

_DEC2 = Decimal("0.01")
_DEC3 = Decimal("0.001")
_DEC0 = Decimal(0)
_MAX_EXPORT_ROWS = 20000  # 流水/库存导出行数上限（防全表加载导致 OOM）
_MAX_INVENTORY_AGG_ROWS = 50000  # 进销存汇总聚合行数上限


def _fmt_qty(v: Decimal | int | str | None) -> str:
    """数量统一格式化：保留至 3 位小数并去尾零（30.000 → 30，5.500 → 5.5）。"""
    d = Decimal(v or 0)
    return format(d.quantize(_DEC3), "f").rstrip("0").rstrip(".") or "0"
_EPOCH = datetime(2000, 1, 1)


def _parse_range(start: str, end: str) -> tuple[datetime, datetime]:
    """日期区间解析：缺省为 本月1号 00:00:00 ~ 今天 23:59:59。"""
    today = date.today()
    try:
        s = datetime.combine(date.fromisoformat(start), time.min) if start else datetime.combine(today.replace(day=1), time.min)
        e = datetime.combine(date.fromisoformat(end), time.max) if end else datetime.combine(today, time.max)
    except ValueError:
        raise BizError(E_PARAM, "日期格式错误，应为 YYYY-MM-DD")
    if s > e:
        raise BizError(E_PARAM, "开始日期不能晚于结束日期")
    return s, e


def _in_out_sum(db: Session, start: datetime | None, end: datetime | None, warehouse_id: int = 0) -> tuple[Decimal, Decimal]:
    """区间内入库件数（change_qty>0）与出库件数（|change_qty<0|）。"""
    stmt = select(
        func.coalesce(func.sum(case((StkStockLog.change_qty > 0, StkStockLog.change_qty), else_=0)), 0),
        func.coalesce(func.sum(case((StkStockLog.change_qty < 0, -StkStockLog.change_qty), else_=0)), 0),
    ).select_from(StkStockLog)
    if warehouse_id:
        stmt = stmt.where(StkStockLog.warehouse_id == warehouse_id)
    if start:
        stmt = stmt.where(StkStockLog.created_at >= start)
    if end:
        stmt = stmt.where(StkStockLog.created_at <= end)
    in_qty, out_qty = db.execute(stmt).one()
    return Decimal(in_qty or 0), Decimal(out_qty or 0)


def _alert_count(db: Session) -> int:
    """低于下限/高于上限的库存行数（口径同 scheduler）。"""
    return (
        db.scalar(
            select(func.count())
            .select_from(StkStock)
            .join(BaseProduct, BaseProduct.id == StkStock.product_id)
            .where(StkStock.qty != 0)
            .where(
                or_(
                    and_(BaseProduct.min_stock > 0, StkStock.qty < BaseProduct.min_stock),
                    and_(BaseProduct.max_stock > 0, StkStock.qty > BaseProduct.max_stock),
                )
            )
        )
        or 0
    )


def _trend_7d(db: Session) -> list[dict]:
    """近 7 日（含今天）每日入库/出库件数。"""
    start = datetime.combine(date.today() - timedelta(days=6), time.min)
    rows = db.execute(
        select(
            func.date(StkStockLog.created_at).label("d"),
            func.coalesce(func.sum(case((StkStockLog.change_qty > 0, StkStockLog.change_qty), else_=0)), 0),
            func.coalesce(func.sum(case((StkStockLog.change_qty < 0, -StkStockLog.change_qty), else_=0)), 0),
        )
        .where(StkStockLog.created_at >= start)
        .group_by(func.date(StkStockLog.created_at))
    ).all()
    m = {str(d): (Decimal(i or 0), Decimal(o or 0)) for d, i, o in rows}
    out = []
    for i in range(6, -1, -1):
        d = (date.today() - timedelta(days=i)).isoformat()
        in_qty, out_qty = m.get(d, (_DEC0, _DEC0))
        out.append({"date": d, "in_qty": _fmt_qty(in_qty), "out_qty": _fmt_qty(out_qty)})
    return out


def _pending_todos(db: Session) -> dict:
    """待办：领用待审计、调拨待审核、盘点进行中。"""
    return {
        "pending_requisitions": db.scalar(select(func.count()).select_from(OutRequisition).where(OutRequisition.status == REQ_STATUS_PENDING)) or 0,
        "pending_transfers": db.scalar(select(func.count()).select_from(StkTransfer).where(StkTransfer.status == 0)) or 0,
        "pending_checks": db.scalar(select(func.count()).select_from(StkCheck).where(StkCheck.status == 1)) or 0,
    }


@router.get("/stock/summary")
def stock_summary(db: Session = Depends(get_db)) -> dict:
    """看板：SKU 数、总件数、今日入库/出库件数、预警数、待审计单数、近 7 日趋势。

    60 秒 TTL 缓存；库存变动（post_stock_change）即时失效。
    """
    def _load() -> dict:
        today_start = datetime.combine(date.today(), time.min)
        today_in, today_out = _in_out_sum(db, today_start, None)
        return {
            "sku_count": db.scalar(select(func.count(func.distinct(StkStock.product_id))).where(StkStock.qty != 0)) or 0,
            "total_qty": _fmt_qty(db.scalar(select(func.coalesce(func.sum(StkStock.qty), 0))) or 0),
            "today_in_qty": _fmt_qty(today_in),
            "today_out_qty": _fmt_qty(today_out),
            "alert_count": _alert_count(db),
            "pending_requisition_count": db.scalar(
                select(func.count()).select_from(OutRequisition).where(OutRequisition.status == REQ_STATUS_PENDING)
            )
            or 0,
            "trend_7d": _trend_7d(db),
        }

    return ok(cache_aside_json("dash:stock_summary", 60, _load))


@router.get("/reports/dashboard", dependencies=[Depends(require_permission("report:view"))])
def dashboard(db: Session = Depends(get_db)) -> dict:
    """经营看板：今日/本周/本月出入库、预警、待办、近 7 日趋势。60 秒 TTL 缓存。"""

    def _load() -> dict:
        now = datetime.now()
        today_start = datetime.combine(now.date(), time.min)
        week_start = today_start - timedelta(days=now.weekday())
        month_start = today_start.replace(day=1)
        today_in, today_out = _in_out_sum(db, today_start, None)
        week_in, week_out = _in_out_sum(db, week_start, None)
        month_in, month_out = _in_out_sum(db, month_start, None)
        return {
            "today": {"in_qty": _fmt_qty(today_in), "out_qty": _fmt_qty(today_out)},
            "week": {"in_qty": _fmt_qty(week_in), "out_qty": _fmt_qty(week_out)},
            "month": {"in_qty": _fmt_qty(month_in), "out_qty": _fmt_qty(month_out)},
            "sku_count": db.scalar(select(func.count(func.distinct(StkStock.product_id))).where(StkStock.qty != 0)) or 0,
            "total_qty": _fmt_qty(db.scalar(select(func.coalesce(func.sum(StkStock.qty), 0))) or 0),
            "alert_count": _alert_count(db),
            "todos": _pending_todos(db),
            "trend_7d": _trend_7d(db),
        }

    return ok(cache_aside_json("dash:dashboard", 60, _load))


def _inventory_rows(db: Session, start: datetime, end: datetime, warehouse_id: int = 0, product_id: int = 0) -> list[dict]:
    """期间进销存汇总（全量，供分页与导出复用）。按商品聚合：期初+入库-出库=结存。

    金额口径（与《后端API设计.md》一致）：入库/出库金额 = 流水变动数量 × 该笔流水成本价；
    期初金额 = 结存金额 - 入库金额 + 出库金额（保持 期初+入库-出库=结存 恒等）。
    """
    in_period = and_(StkStockLog.created_at >= start, StkStockLog.created_at <= end)
    stmt = select(
        StkStockLog.product_id,
        StkStockLog.location_id,
        func.coalesce(func.sum(case((StkStockLog.created_at < start, StkStockLog.change_qty), else_=0)), 0),
        func.coalesce(func.sum(case((and_(in_period, StkStockLog.change_qty > 0), StkStockLog.change_qty), else_=0)), 0),
        func.coalesce(func.sum(case((and_(in_period, StkStockLog.change_qty < 0), -StkStockLog.change_qty), else_=0)), 0),
        func.coalesce(func.sum(case((and_(in_period, StkStockLog.change_qty > 0), StkStockLog.change_qty * StkStockLog.cost_price), else_=0)), 0),
        func.coalesce(func.sum(case((and_(in_period, StkStockLog.change_qty < 0), -StkStockLog.change_qty * StkStockLog.cost_price), else_=0)), 0),
    ).select_from(StkStockLog)
    if warehouse_id:
        stmt = stmt.where(StkStockLog.warehouse_id == warehouse_id)
    if product_id:
        stmt = stmt.where(StkStockLog.product_id == product_id)
    rows = db.execute(stmt.group_by(StkStockLog.product_id, StkStockLog.location_id)).all()

    # 结存金额 = 各库位结存 × 该库位当前成本价（移动加权），避免跨库位成本混淆。
    # 只查结果集涉及的 (product_id, location_id)，不再整表加载 stk_stock。
    pairs = sorted({(pid, loc_id) for pid, loc_id, *_ in rows})
    if pairs:
        cost_map = {
            (s.product_id, s.location_id): Decimal(s.cost_price or 0)
            for s in db.scalars(
                select(StkStock).where(tuple_(StkStock.product_id, StkStock.location_id).in_(pairs))
            ).all()
        }
    else:
        cost_map = {}
    agg: dict[int, dict] = {}
    for pid, loc_id, opening, in_qty, out_qty, in_amt, out_amt in rows:
        opening = Decimal(opening or 0)
        in_qty = Decimal(in_qty or 0)
        out_qty = Decimal(out_qty or 0)
        in_amt = Decimal(in_amt or 0)
        out_amt = Decimal(out_amt or 0)
        closing = opening + in_qty - out_qty
        a = agg.setdefault(pid, {"opening": _DEC0, "in_qty": _DEC0, "out_qty": _DEC0, "closing": _DEC0, "amount": _DEC0, "in_amount": _DEC0, "out_amount": _DEC0})
        a["opening"] += opening
        a["in_qty"] += in_qty
        a["out_qty"] += out_qty
        a["closing"] += closing
        a["amount"] += closing * cost_map.get((pid, loc_id), _DEC0)
        a["in_amount"] += in_amt
        a["out_amount"] += out_amt
    return [
        {
            "product_id": pid,
            "opening_qty": _fmt_qty(a["opening"]),
            "in_qty": _fmt_qty(a["in_qty"]),
            "out_qty": _fmt_qty(a["out_qty"]),
            "closing_qty": _fmt_qty(a["closing"]),
            "closing_amount": str(a["amount"].quantize(_DEC2)),
            "opening_amount": str((a["amount"] - a["in_amount"] + a["out_amount"]).quantize(_DEC2)),
            "in_amount": str(a["in_amount"].quantize(_DEC2)),
            "out_amount": str(a["out_amount"].quantize(_DEC2)),
            "unit_price": str((a["amount"] / a["closing"]).quantize(_DEC2)) if a["closing"] else "0.00",
        }
        for pid, a in sorted(agg.items())
    ]


@router.get("/reports/inventory-summary", dependencies=[Depends(require_permission("report:view"))])
def inventory_summary(
    warehouse_id: int = Query(0),
    product_id: int = Query(0),
    start: str = Query(""),
    end: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
) -> dict:
    s, e = _parse_range(start, end)
    rows = _inventory_rows(db, s, e, warehouse_id, product_id)
    if len(rows) > _MAX_INVENTORY_AGG_ROWS:
        raise BizError(E_PARAM, f"汇总结果超过 {_MAX_INVENTORY_AGG_ROWS} 行，请缩小查询范围或导出")
    total = len(rows)
    page_rows = rows[(page - 1) * page_size : page * page_size]
    # 批量预取商品/单位，避免每行 2 次回表（N+1）
    pids = [r["product_id"] for r in page_rows]
    prod_map = {p.id: p for p in db.scalars(select(BaseProduct).where(BaseProduct.id.in_(pids))).all()} if pids else {}
    uids = {p.unit_id for p in prod_map.values() if p.unit_id}
    unit_map = {u.id: u for u in db.scalars(select(BaseUnit).where(BaseUnit.id.in_(uids))).all()} if uids else {}
    out = []
    for r in page_rows:
        p = prod_map.get(r["product_id"])
        unit = unit_map.get(p.unit_id) if p and p.unit_id else None
        out.append(
            {
                **r,
                "code": p.code if p else "",
                "name": p.name if p else "",
                "spec": p.spec if p else "",
                "unit_name": unit.name if unit else "",
            }
        )
    return ok(PageData(list=out, total=total, page=page, page_size=page_size).model_dump())


def _stock_rows(
    db: Session,
    warehouse_id: int,
    sort: str,
    limit: int | None = None,
    offset: int = 0,
) -> list[dict]:
    """库存余额行：qty/amount/turnover 排序，含 30 天出库与呆滞天数。

    limit/offset 在 SQL 层生效（分页不再全表加载）；缺省（导出）返回全量。
    """
    out30 = (
        select(StkStockLog.product_id, func.sum(-StkStockLog.change_qty).label("out30"))
        .where(StkStockLog.change_qty < 0, StkStockLog.created_at >= datetime.now() - timedelta(days=30))
        .group_by(StkStockLog.product_id)
        .subquery()
    )
    last = (
        select(StkStockLog.product_id, func.max(StkStockLog.created_at).label("last"))
        .group_by(StkStockLog.product_id)
        .subquery()
    )
    stmt = (
        select(StkStock, func.coalesce(out30.c.out30, 0).label("out30"), func.coalesce(last.c.last, _EPOCH).label("last"))
        .join(out30, out30.c.product_id == StkStock.product_id, isouter=True)
        .join(last, last.c.product_id == StkStock.product_id, isouter=True)
        .where(StkStock.qty > 0)
    )
    if warehouse_id:
        stmt = stmt.where(StkStock.warehouse_id == warehouse_id)
    if sort == "qty":
        stmt = stmt.order_by(StkStock.qty.desc())
    elif sort == "amount":
        stmt = stmt.order_by((StkStock.qty * StkStock.cost_price).desc())
    else:  # turnover：30 天出库多优先，久未变动靠前
        stmt = stmt.order_by(func.coalesce(out30.c.out30, 0).desc(), func.coalesce(last.c.last, _EPOCH).asc())
    if limit is not None:
        stmt = stmt.limit(limit).offset(offset)
    now = datetime.now()
    rows_all = db.execute(stmt).all()
    # 批量预取商品/仓库，避免每行 2 次回表（N+1）
    pids = {stock.product_id for stock, _, _ in rows_all}
    wids = {stock.warehouse_id for stock, _, _ in rows_all}
    prod_map = {p.id: p for p in db.scalars(select(BaseProduct).where(BaseProduct.id.in_(pids))).all()} if pids else {}
    wh_map = {w.id: w for w in db.scalars(select(BaseWarehouse).where(BaseWarehouse.id.in_(wids))).all()} if wids else {}
    out = []
    for stock, out30_qty, last_moved in rows_all:
        p = prod_map.get(stock.product_id)
        wh = wh_map.get(stock.warehouse_id)
        # 兼容驱动返回 str / datetime / None（coalesce 兜底 _EPOCH 可能改变列类型）
        if isinstance(last_moved, datetime):
            last_dt: datetime | None = last_moved
        elif last_moved:
            last_dt = datetime.fromisoformat(str(last_moved).replace(" ", "T"))
        else:
            last_dt = None
        days = (now - last_dt).days if last_dt else 9999
        out.append(
            {
                "product_id": stock.product_id,
                "code": p.code if p else "",
                "name": p.name if p else "",
                "spec": p.spec if p else "",
                "warehouse_name": wh.name if wh else "",
                "qty": _fmt_qty(stock.qty),
                "cost_price": str(stock.cost_price),
                "amount": str((stock.qty * stock.cost_price).quantize(_DEC2)),
                "out_qty_30d": _fmt_qty(out30_qty),
                "last_moved_at": last_dt.strftime("%Y-%m-%d %H:%M:%S") if last_dt else "",
                "dormant_days": days,
            }
        )
    return out


@router.get("/reports/stock", dependencies=[Depends(require_permission("report:view"))])
def stock_report(
    warehouse_id: int = Query(0),
    sort: str = Query("qty"),  # qty | amount | turnover
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
) -> dict:
    """库存余额/周转/呆滞。"""
    if sort not in ("qty", "amount", "turnover"):
        raise BizError(E_PARAM, "sort 仅支持 qty|amount|turnover")
    total_stmt = select(func.count()).select_from(StkStock).where(StkStock.qty > 0)
    if warehouse_id:
        total_stmt = total_stmt.where(StkStock.warehouse_id == warehouse_id)
    total = db.scalar(total_stmt) or 0
    rows = _stock_rows(db, warehouse_id, sort, limit=page_size, offset=(page - 1) * page_size)
    return ok(PageData(list=rows, total=total, page=page, page_size=page_size).model_dump())


# ============================ Excel 导出（样式对齐 testdata/匹配导出表格/库存金额收发存（2026.06）.xlsx） ============================

_FONT_TITLE = Font(name="宋体", size=14, bold=True)
_FONT_HEADER = Font(name="宋体", size=14, bold=True)
_FONT_HEADER_RED = Font(name="宋体", size=14, bold=True, color="FFFF0000")  # 模板 S/T 列（已使用数量/金额）红色表头
_FONT_DATA = Font(name="宋体", size=11)
_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
# 模板表头仅左/右/上边框，表头与数据的分隔线由数据行上边框提供（避免双线）
_BORDER_LRT = Border(left=_THIN, right=_THIN, top=_THIN)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_DATA = Alignment(vertical="center")

# 库存金额收发存表列（模板 21 列，顺序固定）
_INV_EXPORT_HEADERS = [
    "年月", "仓库名称", "物料分类编码", "物料分类名称", "物料编码", "物料名称", "规格型号", "计量单位",
    "月度期初数量", "月度期初金额", "月度入库数量", "月度入库金额", "月度出库数量", "月度出库金额",
    "月度结存数量", "月度结存金额", "领料员", "备注（板块）", "已使用数量", "金额（元）", "单价",
]
# 模板列宽（未列出的列保持默认宽；I~N 隐藏）
_INV_COL_WIDTHS = {"B": 12, "C": 7.375, "D": 9.625, "E": 16, "F": 25.25, "G": 22.5, "H": 4.75, "Q": 11.5, "R": 17.5, "S": 14.75, "T": 13.5, "U": 10.5}
_INV_TITLE_MERGE = "A1:P1"  # 模板标题仅跨 A~P（16 列）


def _export_xlsx(headers: list[str], data: list[list], filename: str, sheet: str = "报表", title: str | None = None) -> StreamingResponse:
    """通用导出：可选标题行 + 加粗居中表头 + 宋体数据行，风格与收发存模板一致。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    row = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        c = ws.cell(1, 1, title)
        c.font, c.alignment, c.border = _FONT_TITLE, _ALIGN_CENTER, _BORDER_ALL
        ws.row_dimensions[1].height = 47.45
        row = 2
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.font, c.alignment, c.border = _FONT_HEADER, _ALIGN_CENTER, _BORDER_LRT
    ws.row_dimensions[row].height = 52.15
    for r in data:
        row += 1
        ws.row_dimensions[row].height = 25.15
        for col, v in enumerate(r, 1):
            c = ws.cell(row, col, safe_excel_value(v))
            c.font, c.alignment, c.border = _FONT_DATA, _ALIGN_DATA, _BORDER_ALL
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{quote(filename)}"'},
    )


def _export_inventory_xlsx(rows: list[list], title: str, filename: str) -> StreamingResponse:
    """库存金额收发存表导出：1:1 复刻 testdata/匹配导出表格/库存金额收发存（2026.06）.xlsx 的
    格式/布局/样式（标题合并 A1:P1、21 列表头（S/T 红色）、I~N 隐藏、列宽/行高、A4 纵向）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    for col, w in _INV_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    for col in "IJKLMN":
        ws.column_dimensions[col].hidden = True
        ws.column_dimensions[col].width = 0  # 与模板一致：隐藏列宽 0
    # 标题行
    ws.merge_cells(_INV_TITLE_MERGE)
    c = ws["A1"]
    c.value = title
    c.font, c.alignment, c.border = _FONT_TITLE, _ALIGN_CENTER, _BORDER_ALL
    ws.row_dimensions[1].height = 47.45
    # 表头行（S/T 列红色加粗）
    for col, h in enumerate(_INV_EXPORT_HEADERS, 1):
        c = ws.cell(2, col, h)
        c.font = _FONT_HEADER_RED if col in (19, 20) else _FONT_HEADER
        c.alignment, c.border = _ALIGN_CENTER, _BORDER_LRT
    ws.row_dimensions[2].height = 52.15
    # 数据行（U 列单价 0.00 两位小数格式）
    for i, r in enumerate(rows, start=3):
        ws.row_dimensions[i].height = 25.15
        for col, v in enumerate(r, 1):
            c = ws.cell(i, col, safe_excel_value(v))
            c.font, c.alignment, c.border = _FONT_DATA, _ALIGN_DATA, _BORDER_ALL
            if col == 21:
                c.number_format = "0.00_ "
    # 页面设置：A4 纵向（与模板一致）
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75, header=0.3, footer=0.3)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # 中文文件名：RFC 5987 filename*（ASCII 兜底）
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="inventory-summary.xlsx"; filename*=UTF-8\'\'{quote(filename)}'},
    )


@router.get("/reports/export", dependencies=[Depends(require_permission("report:export"))])
def export_report(
    type: str = Query("inventory-summary"),  # inventory-summary | stock | flow
    warehouse_id: int = Query(0),
    start: str = Query(""),
    end: str = Query(""),
    sort: str = Query("qty"),
    product_id: int = Query(0),
    bill_no: str = Query("", max_length=30),
    change_type: str = Query("", max_length=20),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    today = date.today().isoformat().replace("-", "")
    if type == "inventory-summary":
        s, e = _parse_range(start, end)
        same_month = (s.year, s.month) == (e.year, e.month)
        period = f"{s.year}年{s.month}月" if same_month else f"{s:%Y-%m}~{e:%Y-%m}"
        ym = f"{s:%Y-%m}" if same_month else f"{s:%Y-%m}~{e:%Y-%m}"
        wh_name = ""
        if warehouse_id:
            wh = db.get(BaseWarehouse, warehouse_id)
            wh_name = wh.name if wh else ""
        data = []
        inv_rows = _inventory_rows(db, s, e, warehouse_id, product_id)
        if len(inv_rows) > _MAX_INVENTORY_AGG_ROWS:
            raise BizError(E_PARAM, f"导出汇总超过 {_MAX_INVENTORY_AGG_ROWS} 行，请缩小范围后分批导出")
        # 批量预取商品/单位/分类，避免每行 3 次回表（N+1）
        pids = [r["product_id"] for r in inv_rows]
        prod_map = {p.id: p for p in db.scalars(select(BaseProduct).where(BaseProduct.id.in_(pids))).all()} if pids else {}
        uids = {p.unit_id for p in prod_map.values() if p.unit_id}
        unit_map = {u.id: u for u in db.scalars(select(BaseUnit).where(BaseUnit.id.in_(uids))).all()} if uids else {}
        cids = {p.category_id for p in prod_map.values() if p.category_id}
        cat_map = {c.id: c for c in db.scalars(select(BaseCategory).where(BaseCategory.id.in_(cids))).all()} if cids else {}
        for r in inv_rows:
            p = prod_map.get(r["product_id"])
            unit = unit_map.get(p.unit_id) if p and p.unit_id else None
            cat = cat_map.get(p.category_id) if p and p.category_id else None
            data.append(
                [
                    ym,
                    wh_name,
                    str(cat.id) if cat else "",  # 物料分类编码（系统无独立编码字段，用分类 ID）
                    cat.name if cat else "",
                    p.code if p else "",
                    p.name if p else "",
                    p.spec if p else "",
                    unit.name if unit else "",
                    r["opening_qty"],
                    r["opening_amount"],
                    r["in_qty"],
                    r["in_amount"],
                    r["out_qty"],
                    r["out_amount"],
                    r["closing_qty"],
                    r["closing_amount"],
                    "", "", "", 0,  # 领料员 / 备注（板块） / 已使用数量（模板列，库存口径无数据）
                    r["unit_price"],
                ]
            )
        now = datetime.now().strftime("%H%M")
        filename = f"库存金额收发存{s:%Y.%m}_{now}.xlsx" if same_month else f"库存金额收发存{s:%Y.%m}-{e:%Y.%m}_{now}.xlsx"
        return _export_inventory_xlsx(data, f"{period}库存金额收发存表", filename)
    if type == "stock":
        if sort not in ("qty", "amount", "turnover"):
            raise BizError(E_PARAM, "sort 仅支持 qty|amount|turnover")
        stock_rows = _stock_rows(db, warehouse_id, sort)
        if len(stock_rows) > _MAX_EXPORT_ROWS:
            raise BizError(E_PARAM, f"导出库存超过 {_MAX_EXPORT_ROWS} 行，请按仓库分批导出")
        data = [
            [
                r["code"], r["name"], r["spec"], r["warehouse_name"], r["qty"],
                r["cost_price"], r["amount"], r["out_qty_30d"], r["last_moved_at"], r["dormant_days"],
            ]
            for r in stock_rows
        ]
        return _export_xlsx(
            ["商品编码", "商品名称", "规格", "仓库", "数量", "成本价", "金额", "30天出库", "最近变动", "呆滞天数"],
            data,
            f"stock_{today}.xlsx",
            title="库存报表",
        )
    if type == "flow":
        stmt = select(StkStockLog)
        if product_id:
            stmt = stmt.where(StkStockLog.product_id == product_id)
        if bill_no:
            stmt = stmt.where(StkStockLog.bill_no.like(f"%{bill_no}%"))
        if change_type:
            stmt = stmt.where(StkStockLog.change_type == change_type)
        if start:
            try:
                date.fromisoformat(start)
            except ValueError:
                raise BizError(E_PARAM, "start 格式错误，应为 YYYY-MM-DD")
            stmt = stmt.where(StkStockLog.created_at >= f"{start} 00:00:00")
        if end:
            try:
                date.fromisoformat(end)
            except ValueError:
                raise BizError(E_PARAM, "end 格式错误，应为 YYYY-MM-DD")
            stmt = stmt.where(StkStockLog.created_at <= f"{end} 23:59:59")
        rows = db.scalars(stmt.order_by(StkStockLog.id.desc()).limit(_MAX_EXPORT_ROWS + 1)).all()
        if len(rows) > _MAX_EXPORT_ROWS:
            raise BizError(E_PARAM, f"导出流水超过 {_MAX_EXPORT_ROWS} 行，请缩小日期范围后分批导出")
        # 批量预取商品/仓库/库位，避免每行 3 次回表（N+1）
        pids = {log.product_id for log in rows}
        wids = {log.warehouse_id for log in rows}
        lids = {log.location_id for log in rows}
        prod_map = {p.id: p for p in db.scalars(select(BaseProduct).where(BaseProduct.id.in_(pids))).all()} if pids else {}
        wh_map = {w.id: w for w in db.scalars(select(BaseWarehouse).where(BaseWarehouse.id.in_(wids))).all()} if wids else {}
        loc_map = {l.id: l for l in db.scalars(select(BaseLocation).where(BaseLocation.id.in_(lids))).all()} if lids else {}
        data = []
        for log in rows:
            p = prod_map.get(log.product_id)
            wh = wh_map.get(log.warehouse_id)
            loc = loc_map.get(log.location_id)
            data.append(
                [
                    log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    p.code if p else "",
                    p.name if p else "",
                    wh.name if wh else "",
                    loc.code if loc else "",
                    log.change_type,
                    log.bill_no,
                    str(log.before_qty),
                    str(log.change_qty),
                    str(log.after_qty),
                    str(log.cost_price),
                    log.remark,
                ]
            )
        return _export_xlsx(
            ["时间", "商品编码", "商品名称", "仓库", "库位", "类型", "单据号", "变动前", "变动数量", "变动后", "成本价", "备注"],
            data,
            f"flow_{today}.xlsx",
            title="库存流水",
        )
    raise BizError(E_PARAM, "type 仅支持 inventory-summary|stock|flow")


@router.post("/reports/ai-summary", dependencies=[Depends(require_permission("report:view"))])
def report_ai_summary(body: dict, db: Session = Depends(get_db)) -> dict:
    """AI 月报摘要（P9-P1⑦）：服务端聚合经营数据 → DeepSeek 生成 200-300 字摘要（未配置降级规则版）。"""
    try:
        start = date.fromisoformat(str(body.get("start") or ""))
        end = date.fromisoformat(str(body.get("end") or ""))
    except ValueError:
        raise BizError(E_PARAM, "start/end 必须为 YYYY-MM-DD 日期")
    if end < start:
        raise BizError(E_PARAM, "end 不能早于 start")
    from app.services.ai.report_summary import report_summary

    return ok(report_summary(db, start, end))
