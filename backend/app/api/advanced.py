"""库存进阶接口：调拨/盘点/其他出入库（《后端API设计.md》§5）。

库存一致性：全部走 services/stock.py 的 post_stock_change()，与单据状态同事务提交。
盘点为物品级别：按商品聚合账面数量，差异在审核时按库位分摊入账。
"""
from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Annotated
from urllib.parse import quote

from fastapi import APIRouter, Body, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from sqlalchemy import and_, case, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.deps import get_current_user, require_permission
from app.core.excel_guard import safe_excel_value
from app.core.response import BizError, E_BILL_STATUS, E_NOT_FOUND, E_PARAM, E_STOCK_NOT_ENOUGH, ok
from app.db import get_db
from app.models.advanced import (
    StkCheck,
    StkCheckItem,
    StkOtherIo,
    StkOtherIoItem,
    StkTransfer,
    StkTransferItem,
)
from app.models.base import BaseCategory, BaseLocation, BaseProduct, BaseUnit, BaseWarehouse
from app.models.stock import StkStock, StkStockLog
from app.models.sys import SysUser
from app.schemas.advanced import (
    CheckItemOut,
    CheckItemsReq,
    CheckOut,
    CheckReq,
    IN_TYPES,
    OtherIoItemReq,
    OtherIoOut,
    OtherIoReq,
    PageData,
    TransferItemReq,
    TransferOut,
    TransferReq,
)
from app.services.stock import bill_no_conflict, generate_bill_no, post_stock_change

router = APIRouter(tags=["库存进阶"], dependencies=[Depends(get_current_user)])

_DECIMAL_RE = re.compile(r"^\d+(\.\d+)?$")


def _parse_qty(v: str) -> Decimal:
    if not _DECIMAL_RE.match(v) or Decimal(v) <= 0:
        raise BizError(E_PARAM, "数量必须为正数")
    return Decimal(v)


def _user_name(db: Session, uid: int) -> str:
    u = db.get(SysUser, uid)
    return u.real_name if u else ""


def _loc_code(db: Session, loc_id: int) -> str:
    loc = db.get(BaseLocation, loc_id)
    return loc.code if loc else ""


def _item_out(db: Session, product_id: int, location_id: int, qty: Decimal, item_id: int) -> dict:
    return {
        "id": item_id,
        "product_id": product_id,
        "product_name": (p.name if (p := db.get(BaseProduct, product_id)) else ""),
        "code": (p.code if (p := db.get(BaseProduct, product_id)) else ""),
        "location_id": location_id,
        "location_code": _loc_code(db, location_id),
        "qty": format(qty, "f"),
    }


# ============================ 调拨 ============================


@router.post("/transfers", dependencies=[Depends(require_permission("stk:transfer"))])
def create_transfer(
    req: Annotated[TransferReq, Body()],
    db: Session = Depends(get_db),
) -> dict:
    if req.from_warehouse_id == req.to_warehouse_id:
        raise BizError(E_PARAM, "调出仓库与调入仓库不能相同")
    if db.get(BaseWarehouse, req.from_warehouse_id) is None or db.get(BaseWarehouse, req.to_warehouse_id) is None:
        raise BizError(E_PARAM, "仓库不存在")
    for attempt in range(5):  # 单号并发冲突重试
        bill = StkTransfer(
            bill_no=generate_bill_no(db, "DB", StkTransfer),
            from_warehouse_id=req.from_warehouse_id,
            to_warehouse_id=req.to_warehouse_id,
            status=0,
            remark=req.remark,
        )
        db.add(bill)
        db.flush()
        try:
            for idx, item in enumerate(req.items):
                if db.get(BaseProduct, item.product_id) is None:
                    raise BizError(E_NOT_FOUND, f"商品 id={item.product_id} 不存在")
                if db.get(BaseLocation, item.from_location_id) is None or db.get(BaseLocation, item.to_location_id) is None:
                    raise BizError(E_NOT_FOUND, "库位不存在")
                db.add(StkTransferItem(
                    transfer_id=bill.id, product_id=item.product_id, qty=_parse_qty(item.qty),
                    from_location_id=item.from_location_id, to_location_id=item.to_location_id,
                ))
            db.commit()
            return ok({"id": bill.id, "bill_no": bill.bill_no, "status": 0})
        except IntegrityError as exc:
            db.rollback()
            if not bill_no_conflict(exc):
                raise BizError(E_PARAM, f"调拨单保存失败：{exc.orig}") from exc
    raise BizError(E_PARAM, "单据编号生成失败，请重试")


def _transfer_out(db: Session, b: StkTransfer) -> dict:
    items = db.scalars(select(StkTransferItem).where(StkTransferItem.transfer_id == b.id)).all()
    return TransferOut(
        id=b.id, bill_no=b.bill_no,
        from_warehouse_id=b.from_warehouse_id,
        from_warehouse_name=(w.name if (w := db.get(BaseWarehouse, b.from_warehouse_id)) else ""),
        to_warehouse_id=b.to_warehouse_id,
        to_warehouse_name=(w.name if (w := db.get(BaseWarehouse, b.to_warehouse_id)) else ""),
        status=b.status, audit_name=_user_name(db, b.audit_by), audit_time=b.audit_time, remark=b.remark,
        items=[_item_out(db, it.product_id, it.from_location_id, it.qty, it.id) for it in items],
    ).model_dump()


@router.get("/transfers")
def list_transfers(
    status: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    stmt = select(StkTransfer)
    if status is not None:
        stmt = stmt.where(StkTransfer.status == status)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(StkTransfer.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    return ok(PageData(list=[_transfer_out(db, b) for b in rows], total=total, page=page, page_size=page_size).model_dump())


@router.get("/transfers/{bill_id}")
def get_transfer(bill_id: int, db: Session = Depends(get_db)) -> dict:
    b = db.get(StkTransfer, bill_id)
    if b is None:
        raise BizError(E_NOT_FOUND, "调拨单不存在")
    return ok(_transfer_out(db, b))


@router.post("/transfers/{bill_id}/audit", dependencies=[Depends(require_permission("stk:transfer"))])
def audit_transfer(
    bill_id: int,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    b = db.get(StkTransfer, bill_id)
    if b is None:
        raise BizError(E_NOT_FOUND, "调拨单不存在")
    if b.status != 0:
        raise BizError(E_BILL_STATUS, "仅草稿可审核")
    items = db.scalars(select(StkTransferItem).where(StkTransferItem.transfer_id == b.id)).all()
    if not items:
        raise BizError(E_PARAM, "调拨明细为空")
    # 同事务：调出仓扣（负）、调入仓加（正），任一库存不足整单回滚
    for it in items:
        post_stock_change(
            db,
            product_id=it.product_id, warehouse_id=b.from_warehouse_id, location_id=it.from_location_id,
            change_type="调拨出", bill_type="stk_transfer", bill_no=b.bill_no,
            bill_item_id=it.id, qty_delta=-it.qty, operator_id=user.id,
        )
        post_stock_change(
            db,
            product_id=it.product_id, warehouse_id=b.to_warehouse_id, location_id=it.to_location_id,
            change_type="调拨入", bill_type="stk_transfer", bill_no=b.bill_no,
            bill_item_id=it.id, qty_delta=it.qty, cost_price=Decimal(0), operator_id=user.id,
        )
    b.status = 1
    b.audit_by = user.id
    b.audit_time = datetime.now()
    db.commit()
    return ok()


@router.post("/transfers/{bill_id}/reject", dependencies=[Depends(require_permission("stk:transfer"))])
def reject_transfer(
    bill_id: int,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """调拨审核驳回：仅草稿可驳回（status=0 → -2 已驳回），不产生库存变动。"""
    b = db.get(StkTransfer, bill_id)
    if b is None:
        raise BizError(E_NOT_FOUND, "调拨单不存在")
    if b.status != 0:
        raise BizError(E_BILL_STATUS, "仅草稿可驳回")
    b.status = -2  # 已驳回
    b.audit_by = user.id
    b.audit_time = datetime.now()
    db.commit()
    return ok()


@router.post("/transfers/{bill_id}/void", dependencies=[Depends(require_permission("stk:transfer"))])
def void_transfer(
    bill_id: int,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    b = db.get(StkTransfer, bill_id)
    if b is None:
        raise BizError(E_NOT_FOUND, "调拨单不存在")
    if b.status == -1:
        raise BizError(E_BILL_STATUS, "单据已作废")
    if b.status == 1:
        # 已审核：反向冲销（调入仓扣回、调出仓加回），校验调入仓库存充足
        items = db.scalars(select(StkTransferItem).where(StkTransferItem.transfer_id == b.id)).all()
        for it in items:
            post_stock_change(
                db,
                product_id=it.product_id, warehouse_id=b.to_warehouse_id, location_id=it.to_location_id,
                change_type="调拨作废", bill_type="stk_transfer", bill_no=b.bill_no,
                bill_item_id=it.id, qty_delta=-it.qty, operator_id=user.id, remark="作废冲销",
            )
            post_stock_change(
                db,
                product_id=it.product_id, warehouse_id=b.from_warehouse_id, location_id=it.from_location_id,
                change_type="调拨作废", bill_type="stk_transfer", bill_no=b.bill_no,
                bill_item_id=it.id, qty_delta=it.qty, cost_price=Decimal(0), operator_id=user.id, remark="作废冲销",
            )
    b.status = -1
    db.commit()
    return ok()


# ============================ 盘点 ============================


@router.post("/checks", dependencies=[Depends(require_permission("stk:check"))])
def create_check(
    req: CheckReq,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    if db.get(BaseWarehouse, req.warehouse_id) is None:
        raise BizError(E_PARAM, "仓库不存在")
    for attempt in range(5):  # 单号并发冲突重试
        bill = StkCheck(
            bill_no=generate_bill_no(db, "PD", StkCheck),
            warehouse_id=req.warehouse_id,
            status=0,
            checker_id=user.id,
            check_date=datetime.now(),
            remark=req.remark,
        )
        db.add(bill)
        db.flush()
        # 物品级别盘点：按商品聚合该仓库账面库存（跨库位汇总，每物品一行）
        rows = db.execute(
            select(StkStock.product_id, func.sum(StkStock.qty))
            .where(StkStock.warehouse_id == req.warehouse_id, StkStock.qty != 0)
            .group_by(StkStock.product_id)
        ).all()
        for product_id, qty in rows:
            if not qty:
                continue
            # location_id=0 表示物品级汇总行，审核时按库位分摊入账
            db.add(StkCheckItem(check_id=bill.id, product_id=product_id, location_id=0, book_qty=qty, diff_qty=0))
        try:
            db.commit()
            return ok({"id": bill.id, "bill_no": bill.bill_no, "status": 0, "item_count": len(rows)})
        except IntegrityError as exc:
            db.rollback()
            if not bill_no_conflict(exc):
                raise BizError(E_PARAM, f"盘点单保存失败：{exc.orig}") from exc
    raise BizError(E_PARAM, "单据编号生成失败，请重试")


def _check_out(db: Session, b: StkCheck) -> dict:
    items = db.scalars(select(StkCheckItem).where(StkCheckItem.check_id == b.id).order_by(StkCheckItem.id)).all()
    return CheckOut(
        id=b.id, bill_no=b.bill_no, warehouse_id=b.warehouse_id,
        warehouse_name=(w.name if (w := db.get(BaseWarehouse, b.warehouse_id)) else ""),
        status=b.status, checker_name=_user_name(db, b.checker_id), check_date=b.check_date, remark=b.remark,
        items=[
            CheckItemOut(
                id=it.id, product_id=it.product_id,
                product_name=(p.name if (p := db.get(BaseProduct, it.product_id)) else ""),
                code=(p.code if (p := db.get(BaseProduct, it.product_id)) else ""),
                material_code=(p.material_code if (p := db.get(BaseProduct, it.product_id)) else ""),
                spec=(p.spec if (p := db.get(BaseProduct, it.product_id)) else ""),
                unit_name=(u.name if (p := db.get(BaseProduct, it.product_id)) and (u := db.get(BaseUnit, p.unit_id)) else ""),
                category_name=(
                    (c.name if (p := db.get(BaseProduct, it.product_id)) and p.category_id and (c := db.get(BaseCategory, p.category_id)) else "")
                ),
                location_id=it.location_id, location_code=_loc_code(db, it.location_id),
                book_qty=it.book_qty, real_qty=it.real_qty, diff_qty=it.diff_qty,
                photo_file_id=it.photo_file_id,
            )
            for it in items
        ],
    ).model_dump()


@router.get("/checks")
def list_checks(
    status: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    stmt = select(StkCheck)
    if status is not None:
        stmt = stmt.where(StkCheck.status == status)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(StkCheck.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    return ok(PageData(list=[_check_out(db, b) for b in rows], total=total, page=page, page_size=page_size).model_dump())


@router.get("/checks/{bill_id}")
def get_check(bill_id: int, db: Session = Depends(get_db)) -> dict:
    b = db.get(StkCheck, bill_id)
    if b is None:
        raise BizError(E_NOT_FOUND, "盘点单不存在")
    return ok(_check_out(db, b))


@router.put("/checks/{bill_id}/items", dependencies=[Depends(require_permission("stk:check"))])
def update_check_items(bill_id: int, req: CheckItemsReq, db: Session = Depends(get_db)) -> dict:
    b = db.get(StkCheck, bill_id)
    if b is None:
        raise BizError(E_NOT_FOUND, "盘点单不存在")
    if b.status not in (0, 1):
        raise BizError(E_BILL_STATUS, "已审核的盘点单不可修改")
    for item in req.items:
        if item.check_item_id == 0:
            # 当场新增账外物料：账面 0，实盘为录入值，审核时按盘盈入账
            if item.product_id <= 0 or db.get(BaseProduct, item.product_id) is None:
                raise BizError(E_PARAM, "新增物料无效")
            dup = db.scalar(
                select(func.count()).select_from(StkCheckItem).where(
                    StkCheckItem.check_id == b.id, StkCheckItem.product_id == item.product_id
                )
            )
            if dup:
                raise BizError(E_PARAM, f"物料 id={item.product_id} 已在本盘点单中")
            real = _parse_qty(item.real_qty)
            db.add(StkCheckItem(
                check_id=b.id, product_id=item.product_id, location_id=0,
                book_qty=Decimal(0), real_qty=real, diff_qty=real,
            ))
            continue
        ci = db.get(StkCheckItem, item.check_item_id)
        if ci is None or ci.check_id != b.id:
            raise BizError(E_NOT_FOUND, f"盘点明细 id={item.check_item_id} 不存在")
        ci.real_qty = _parse_qty(item.real_qty)
        ci.diff_qty = (ci.real_qty - ci.book_qty).quantize(Decimal("0.001"))
        ci.photo_file_id = item.photo_file_id  # 盘点拍照记录（可选）
    if b.status == 0:
        b.status = 1  # 录入实盘后进入"盘点中"
    db.commit()
    return ok()


@router.post("/checks/{bill_id}/audit", dependencies=[Depends(require_permission("stk:check"))])
def audit_check(
    bill_id: int,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    b = db.get(StkCheck, bill_id)
    if b is None:
        raise BizError(E_NOT_FOUND, "盘点单不存在")
    if b.status != 1:
        raise BizError(E_BILL_STATUS, "仅盘点中（已录实盘）的单据可审核")
    items = db.scalars(select(StkCheckItem).where(StkCheckItem.check_id == b.id)).all()
    if not items:
        raise BizError(E_PARAM, "盘点明细为空")
    for it in items:
        if it.real_qty is None:
            raise BizError(E_PARAM, f"盘点明细 id={it.id} 未录入实盘数量")
    # 一次性取该仓库全部商品库位库存（数量降序），内存按商品分组分摊，避免每明细一次查询（N+1）
    pids = [it.product_id for it in items if it.diff_qty != 0]
    loc_by_product: dict[int, list[tuple[int, Decimal, Decimal]]] = {}
    if pids:
        loc_rows = db.execute(
            select(StkStock.product_id, StkStock.location_id, StkStock.qty, StkStock.cost_price)
            .where(
                StkStock.warehouse_id == b.warehouse_id,
                StkStock.product_id.in_(pids),
                StkStock.qty != 0,
            )
            .order_by(StkStock.product_id, StkStock.qty.desc())
        ).all()
        for pid, lid, q, c in loc_rows:
            loc_by_product.setdefault(pid, []).append((lid, q, c))
    first_loc_cache: dict[int, int | None] = {}
    for it in items:
        if it.diff_qty == 0:
            continue
        change_type = "盘盈" if it.diff_qty > 0 else "盘亏"
        remark = f"账面 {format(it.book_qty, 'f')} → 实盘 {format(it.real_qty, 'f')}"
        # 物品级差异按库位分摊：盘亏从库存多的库位依次扣，盘盈记入库存最多的库位（无库存则仓库首个库位）
        loc_rows = loc_by_product.get(it.product_id, [])
        if it.diff_qty > 0:
            if loc_rows:
                loc_id = loc_rows[0][0]
                # 按当前账面成本入账，避免摊薄移动加权成本
                cost_price = loc_rows[0][2] or Decimal(0)
            else:
                if it.product_id not in first_loc_cache:
                    first_loc_cache[it.product_id] = db.scalar(
                        select(BaseLocation.id).where(BaseLocation.warehouse_id == b.warehouse_id).order_by(BaseLocation.id).limit(1)
                    )
                loc_id = first_loc_cache[it.product_id]
                cost_price = Decimal(0)
            if loc_id is None:
                raise BizError(E_PARAM, "仓库无库位，无法入账盘盈")
            post_stock_change(
                db, product_id=it.product_id, warehouse_id=b.warehouse_id, location_id=loc_id,
                change_type=change_type, bill_type="stk_check", bill_no=b.bill_no,
                bill_item_id=it.id, qty_delta=it.diff_qty, cost_price=cost_price, operator_id=user.id, remark=remark,
            )
        else:
            remain = -it.diff_qty
            for loc_id, qty, _cost in loc_rows:
                if remain <= 0:
                    break
                take = min(remain, qty)
                post_stock_change(
                    db, product_id=it.product_id, warehouse_id=b.warehouse_id, location_id=loc_id,
                    change_type=change_type, bill_type="stk_check", bill_no=b.bill_no,
                    bill_item_id=it.id, qty_delta=-take, cost_price=Decimal(0), operator_id=user.id, remark=remark,
                )
                remain -= take
            if remain > 0:
                raise BizError(E_STOCK_NOT_ENOUGH, f"商品 id={it.product_id} 库存不足，无法盘亏 {format(remain, 'f')}")
    b.status = 2
    db.commit()
    return ok()


# ============================ 盘点导出（收发存模板格式） ============================

_CHECK_HEADERS = [
    "年月", "仓库名称", "物料分类编码", "物料分类名称", "物料编码", "物料名称", "规格型号", "计量单位",
    "月度期初数量", "月度期初金额", "月度入库数量", "月度入库金额", "月度出库数量", "月度出库金额",
    "月度结存数量", "月度结存金额", "账面数量", "实盘数量", "盘盈盘亏数量", "盘盈盘亏金额", "备注",
]


def _month_flow_rows(db: Session, warehouse_id: int, start: datetime, end: datetime) -> dict[int, dict]:
    """盘点月收发存（按商品聚合）：期初/入库/出库/结存的数量与金额（金额=变动量×流水成本）。"""
    rows = db.execute(
        select(
            StkStockLog.product_id,
            func.coalesce(func.sum(case((StkStockLog.created_at < start, StkStockLog.change_qty), else_=0)), 0),
            func.coalesce(
                func.sum(case((and_(StkStockLog.created_at >= start, StkStockLog.created_at <= end, StkStockLog.change_qty > 0), StkStockLog.change_qty), else_=0)), 0
            ),
            func.coalesce(
                func.sum(case((and_(StkStockLog.created_at >= start, StkStockLog.created_at <= end, StkStockLog.change_qty < 0), -StkStockLog.change_qty), else_=0)), 0
            ),
            func.coalesce(func.sum(case((StkStockLog.created_at < start, StkStockLog.change_qty * StkStockLog.cost_price), else_=0)), 0),
            func.coalesce(
                func.sum(case((and_(StkStockLog.created_at >= start, StkStockLog.created_at <= end, StkStockLog.change_qty > 0), StkStockLog.change_qty * StkStockLog.cost_price), else_=0)), 0
            ),
            func.coalesce(
                func.sum(case((and_(StkStockLog.created_at >= start, StkStockLog.created_at <= end, StkStockLog.change_qty < 0), -StkStockLog.change_qty * StkStockLog.cost_price), else_=0)), 0
            ),
        )
        .where(StkStockLog.warehouse_id == warehouse_id)
        .group_by(StkStockLog.product_id)
    ).all()
    out: dict[int, dict] = {}
    for pid, op_q, in_q, out_q, op_a, in_a, out_a in rows:
        op_q, in_q, out_q = Decimal(op_q or 0), Decimal(in_q or 0), Decimal(out_q or 0)
        op_a, in_a, out_a = Decimal(op_a or 0), Decimal(in_a or 0), Decimal(out_a or 0)
        out[pid] = {
            "opening_qty": op_q, "opening_amount": op_a,
            "in_qty": in_q, "in_amount": in_a,
            "out_qty": out_q, "out_amount": out_a,
            "closing_qty": op_q + in_q - out_q, "closing_amount": op_a + in_a - out_a,
        }
    return out


@router.get("/checks/{bill_id}/export", dependencies=[Depends(require_permission("stk:check"))])
def export_check(bill_id: int, db: Session = Depends(get_db)):
    """导出盘点结果 Excel：列结构与《库存金额收发存（2026.06）》模板一致，追加账面/实盘/盈亏列。"""
    b = db.get(StkCheck, bill_id)
    if b is None:
        raise BizError(E_NOT_FOUND, "盘点单不存在")
    wh = db.get(BaseWarehouse, b.warehouse_id)
    items = db.scalars(select(StkCheckItem).where(StkCheckItem.check_id == b.id).order_by(StkCheckItem.id)).all()
    month_start = b.check_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_end = datetime(month_start.year + 1, 1, 1) if month_start.month == 12 else month_start.replace(month=month_start.month + 1)
    flow = _month_flow_rows(db, b.warehouse_id, month_start, month_end)

    wb = Workbook()
    ws = wb.active
    ws.title = "盘点结果"
    ncols = len(_CHECK_HEADERS)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name="宋体", size=14, bold=True)
    head_font = Font(name="宋体", size=14, bold=True)
    body_font = Font(name="宋体", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 标题行（合并，与模板 A1:P1 一致）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title = f"{month_start.year}年{month_start.month}月库存金额收发存表（盘点结果 {b.bill_no}）"
    c = ws.cell(row=1, column=1, value=title)
    c.font = title_font
    c.alignment = center
    c.border = border
    for col in range(2, ncols + 1):
        ws.cell(row=1, column=col).border = border
    # 表头行
    for col, name in enumerate(_CHECK_HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=name)
        c.font = head_font
        c.alignment = center
        c.border = border
    # 数据行
    # 批量预取商品/分类/单位，避免每行 3 次回表（N+1）
    pids = {it.product_id for it in items}
    prod_map = {p.id: p for p in db.scalars(select(BaseProduct).where(BaseProduct.id.in_(pids))).all()} if pids else {}
    cids = {p.category_id for p in prod_map.values() if p.category_id}
    cat_map = {c.id: c for c in db.scalars(select(BaseCategory).where(BaseCategory.id.in_(cids))).all()} if cids else {}
    uids = {p.unit_id for p in prod_map.values() if p.unit_id}
    unit_map = {u.id: u for u in db.scalars(select(BaseUnit).where(BaseUnit.id.in_(uids))).all()} if uids else {}
    row = 3
    for it in items:
        p = prod_map.get(it.product_id)
        cat = cat_map.get(p.category_id) if p and p.category_id else None
        unit = unit_map.get(p.unit_id) if p and p.unit_id else None
        f = flow.get(it.product_id, {})
        closing_q = f.get("closing_qty", Decimal(0))
        closing_a = f.get("closing_amount", Decimal(0))
        cost = closing_a / closing_q if closing_q else Decimal(0)
        real = it.real_qty if it.real_qty is not None else ""
        values = [
            month_start.strftime("%Y-%m"),
            wh.name if wh else "",
            p.category_id if p and p.category_id else "",
            cat.name if cat else "",
            (p.material_code or p.code) if p else "",
            p.name if p else "",
            p.spec if p else "",
            unit.name if unit else "",
            f.get("opening_qty", 0), f.get("opening_amount", 0),
            f.get("in_qty", 0), f.get("in_amount", 0),
            f.get("out_qty", 0), f.get("out_amount", 0),
            closing_q, closing_a,
            it.book_qty, real, it.diff_qty, it.diff_qty * cost,
            "",
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=float(v) if isinstance(v, Decimal) else safe_excel_value(v))
            c.font = body_font
            c.border = border
            c.alignment = Alignment(vertical="center")
        row += 1
    # 列宽
    widths = [10, 18, 13, 14, 20, 26, 22, 10, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 14, 14, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w
    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"库存金额收发存表_{b.bill_no}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ============================ 其他出入库 ============================


@router.post("/other-io", dependencies=[Depends(require_permission("stk:other"))])
def create_other_io(
    req: OtherIoReq,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    if db.get(BaseWarehouse, req.warehouse_id) is None:
        raise BizError(E_PARAM, "仓库不存在")
    for attempt in range(5):  # 单号并发冲突重试
        bill = StkOtherIo(
            bill_no=generate_bill_no(db, "QT", StkOtherIo),
            warehouse_id=req.warehouse_id,
            io_type=req.io_type,
            status=1,  # 直接过账
            operator_id=user.id,
            remark=req.remark,
        )
        db.add(bill)
        db.flush()
        direction = 1 if req.io_type in IN_TYPES else -1
        try:
            for idx, item in enumerate(req.items):
                if db.get(BaseProduct, item.product_id) is None:
                    raise BizError(E_NOT_FOUND, f"商品 id={item.product_id} 不存在")
                if db.get(BaseLocation, item.location_id) is None:
                    raise BizError(E_NOT_FOUND, f"库位 id={item.location_id} 不存在")
                qty = _parse_qty(item.qty)
                item_row = StkOtherIoItem(
                    bill_id=bill.id, product_id=item.product_id, qty=qty,
                    location_id=item.location_id, photo_file_id=item.photo_file_id, sort=idx,
                )
                db.add(item_row)
                db.flush()
                post_stock_change(
                    db,
                    product_id=item.product_id, warehouse_id=req.warehouse_id, location_id=item.location_id,
                    change_type=req.io_type, bill_type="stk_other_io", bill_no=bill.bill_no,
                    bill_item_id=item_row.id, qty_delta=direction * qty, cost_price=Decimal(0),
                    photo_file_id=item.photo_file_id, operator_id=user.id,
                )
            db.commit()
            return ok({"id": bill.id, "bill_no": bill.bill_no, "status": 1})
        except IntegrityError as exc:
            db.rollback()
            if not bill_no_conflict(exc):
                raise BizError(E_PARAM, f"其他出入库单保存失败：{exc.orig}") from exc
    raise BizError(E_PARAM, "单据编号生成失败，请重试")


def _other_out(db: Session, b: StkOtherIo) -> dict:
    items = db.scalars(select(StkOtherIoItem).where(StkOtherIoItem.bill_id == b.id).order_by(StkOtherIoItem.sort)).all()
    return OtherIoOut(
        id=b.id, bill_no=b.bill_no, warehouse_id=b.warehouse_id,
        warehouse_name=(w.name if (w := db.get(BaseWarehouse, b.warehouse_id)) else ""),
        io_type=b.io_type, status=b.status, operator_name=_user_name(db, b.operator_id), remark=b.remark,
        items=[_item_out(db, it.product_id, it.location_id, it.qty, it.id) for it in items],
    ).model_dump()


@router.get("/other-io")
def list_other_io(
    io_type: str = Query("", max_length=20),
    status: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    stmt = select(StkOtherIo)
    if io_type:
        stmt = stmt.where(StkOtherIo.io_type == io_type)
    if status is not None:
        stmt = stmt.where(StkOtherIo.status == status)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(StkOtherIo.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    return ok(PageData(list=[_other_out(db, b) for b in rows], total=total, page=page, page_size=page_size).model_dump())


@router.get("/other-io/{bill_id}")
def get_other_io(bill_id: int, db: Session = Depends(get_db)) -> dict:
    b = db.get(StkOtherIo, bill_id)
    if b is None:
        raise BizError(E_NOT_FOUND, "其他出入库单不存在")
    return ok(_other_out(db, b))


@router.post("/other-io/{bill_id}/void", dependencies=[Depends(require_permission("stk:other"))])
def void_other_io(
    bill_id: int,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    b = db.get(StkOtherIo, bill_id)
    if b is None:
        raise BizError(E_NOT_FOUND, "其他出入库单不存在")
    if b.status != 1:
        raise BizError(E_BILL_STATUS, "仅已过账单据可作废")
    items = db.scalars(select(StkOtherIoItem).where(StkOtherIoItem.bill_id == b.id)).all()
    direction = -1 if b.io_type in IN_TYPES else 1  # 反向冲销
    for it in items:
        post_stock_change(
            db,
            product_id=it.product_id, warehouse_id=b.warehouse_id, location_id=it.location_id,
            change_type=f"{b.io_type}作废", bill_type="stk_other_io", bill_no=b.bill_no,
            bill_item_id=it.id, qty_delta=direction * it.qty, cost_price=Decimal(0),
            operator_id=user.id, remark="作废冲销",
        )
    b.status = -1
    db.commit()
    return ok()
