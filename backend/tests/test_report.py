"""报表与看板测试（P6，L2 门禁）：看板、进销存汇总、库存报表、Excel 导出、2D 货架图、权限。

测试库与其他测试共享，全局计数断言一律用增量（before/after）或按本测试创建的商品/仓库过滤。
"""
from __future__ import annotations

import io
import re
import uuid
from datetime import date

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from urllib.parse import unquote

from app.main import app

client = TestClient(app)


def _login_admin() -> None:
    r = client.post("/api/v1/auth/login", json={"username": "admin", "password": "admin123"})
    assert r.status_code == 200 and r.json()["code"] == 0


def _login_tester() -> None:
    r = client.post("/api/v1/auth/login", json={"username": "tester_user", "password": "123456"})
    assert r.status_code == 200 and r.json()["code"] == 0


def _setup(qty: str = "30", min_stock: str = "30") -> tuple[int, int, int]:
    """建仓/货架/库位/商品（下限 min_stock）+ 采购入库 qty（单价 2.00）。返回 (wh, loc, pid)。"""
    tag = uuid.uuid4().hex[:6]
    r = client.post("/api/v1/warehouses", json={"code": "RP" + tag, "name": "报表仓"})
    wh = r.json()["data"]["id"]
    r = client.post("/api/v1/shelves", json={"warehouse_id": wh, "code": "A1"})
    r = client.post("/api/v1/locations", json={"warehouse_id": wh, "shelf_id": r.json()["data"]["id"], "layer_no": 1})
    loc = r.json()["data"]["id"]

    unit_id = client.get("/api/v1/units").json()["data"][0]["id"]  # 复用种子单位
    r = client.post("/api/v1/products", json={
        "code": "9" + str(int(tag, 16) % 10**9), "name": "报表物料", "unit_id": unit_id,
        "purchase_price": "2.00", "min_stock": min_stock,
    })
    pid = r.json()["data"]["id"]

    r = client.post("/api/v1/purchase-in", json={
        "warehouse_id": wh, "items": [{"product_id": pid, "qty": qty, "price": "2.00", "location_id": loc}],
    })
    assert r.json()["code"] == 0, r.text
    return wh, loc, pid


def _out_5(wh: int, loc: int, pid: int) -> None:
    """其他出库（报废）5 件，制造出库流水。"""
    r = client.post("/api/v1/other-io", json={
        "io_type": "报废", "warehouse_id": wh,
        "items": [{"product_id": pid, "qty": "5", "location_id": loc}],
    })
    assert r.json()["code"] == 0, r.text


def _alert_before() -> int:
    return client.get("/api/v1/stock/summary").json()["data"]["alert_count"]


# ============================ 看板 ============================

def test_stock_summary() -> None:
    _login_admin()
    before = _alert_before()
    _setup(qty="30", min_stock="40")  # 30 < 40 → 低库存预警 +1
    d = client.get("/api/v1/stock/summary").json()["data"]
    assert d["sku_count"] >= 1
    assert float(d["total_qty"]) >= 30
    assert float(d["today_in_qty"]) >= 30
    assert float(d["today_out_qty"]) >= 0
    assert d["alert_count"] == before + 1
    assert len(d["trend_7d"]) == 7
    assert d["trend_7d"][-1]["date"] == date.today().isoformat()
    assert float(d["trend_7d"][-1]["in_qty"]) >= 30


def test_dashboard_today_and_todos() -> None:
    _login_admin()
    wh, loc, pid = _setup(qty="30", min_stock="20")
    _out_5(wh, loc, pid)
    # 待审计领用单（不审计，保持 pending）
    r = client.post("/api/v1/requisitions", json={
        "warehouse_id": wh, "use_location": "车间A", "use_reason": "报表测试",
        "items": [{"product_id": pid, "qty": "1", "location_id": loc}],
    })
    assert r.json()["code"] == 0, r.text
    before_req = client.get("/api/v1/reports/dashboard").json()["data"]["todos"]["pending_requisitions"]

    d = client.get("/api/v1/reports/dashboard").json()["data"]
    assert float(d["today"]["in_qty"]) >= 30
    assert float(d["today"]["out_qty"]) >= 5
    assert float(d["week"]["in_qty"]) >= 30
    assert float(d["month"]["in_qty"]) >= 30
    assert d["todos"]["pending_requisitions"] == before_req  # 本测试未新增待审计
    assert d["sku_count"] >= 1
    assert float(d["total_qty"]) >= 25


def test_dashboard_alert_high() -> None:
    _login_admin()
    before = _alert_before()
    _setup(qty="30", min_stock="20")  # 30 > 20 且 max_stock=0 → 不预警
    assert client.get("/api/v1/reports/dashboard").json()["data"]["alert_count"] == before
    _setup(qty="30", min_stock="40")  # 30 < 40 → 低库存预警 +1
    assert client.get("/api/v1/reports/dashboard").json()["data"]["alert_count"] == before + 1


# ============================ 进销存汇总 ============================

def test_inventory_summary_period() -> None:
    _login_admin()
    wh, loc, pid = _setup(qty="30", min_stock="20")
    _out_5(wh, loc, pid)
    today = date.today().isoformat()
    r = client.get(f"/api/v1/reports/inventory-summary?product_id={pid}&start={today}&end={today}")
    assert r.status_code == 200 and r.json()["code"] == 0, r.text
    d = r.json()["data"]
    assert d["total"] == 1
    row = d["list"][0]
    assert row["opening_qty"] == "0"
    assert row["in_qty"] == "30"
    assert row["out_qty"] == "5"
    assert row["closing_qty"] == "25"
    assert row["closing_amount"] == "50.00"  # 25 × 移动加权成本 2.00
    assert row["in_amount"] == "60.00"  # 30 × 2.00
    assert row["out_amount"] == "10.00"  # 5 × 2.00
    assert row["opening_amount"] == "0.00"  # 结存金额 - 入库金额 + 出库金额
    assert row["unit_price"] == "2.00"
    # warehouse 过滤：不存在的仓库 → 空
    r = client.get(f"/api/v1/reports/inventory-summary?warehouse_id={wh + 999999}&product_id={pid}&start={today}&end={today}")
    assert r.json()["data"]["total"] == 0


def test_inventory_summary_opening() -> None:
    """期初口径：start 之前的净变动计入期初。"""
    _login_admin()
    wh, loc, pid = _setup(qty="30", min_stock="20")
    _out_5(wh, loc, pid)
    future = "2099-01-01"  # 期间起点在未来 → 全部变动计入期初
    r = client.get(f"/api/v1/reports/inventory-summary?product_id={pid}&start={future}&end={future}")
    row = r.json()["data"]["list"][0]
    assert row["opening_qty"] == "25"
    assert row["in_qty"] == "0"
    assert row["closing_qty"] == "25"
    assert row["opening_amount"] == "50.00"  # 期初 25 × 成本 2.00


# ============================ 库存报表 ============================

def test_stock_report_sorts() -> None:
    _login_admin()
    wh, loc, pid = _setup(qty="30", min_stock="20")
    _out_5(wh, loc, pid)
    r = client.get(f"/api/v1/reports/stock?warehouse_id={wh}&sort=qty")
    assert r.status_code == 200 and r.json()["code"] == 0, r.text
    row = next(x for x in r.json()["data"]["list"] if x["product_id"] == pid)
    assert row["qty"] == "25"
    assert row["amount"] == "50.00"
    assert row["out_qty_30d"] == "5"
    assert row["dormant_days"] == 0
    for sort in ("amount", "turnover"):
        r = client.get(f"/api/v1/reports/stock?warehouse_id={wh}&sort={sort}")
        assert r.json()["code"] == 0
    r = client.get("/api/v1/reports/stock?sort=bad")
    assert r.json()["code"] == 4006


# ============================ Excel 导出 ============================

def test_export_inventory_summary() -> None:
    """导出布局/样式与 testdata/匹配导出表格/库存金额收发存（2026.06）.xlsx 一致：
    标题合并 A1:P1 + 21 列表头（S/T 红色）+ I~N 隐藏 + 数据行自第 3 行起。"""
    _login_admin()
    wh, loc, pid = _setup()
    _out_5(wh, loc, pid)
    today = date.today().isoformat()
    r = client.get(f"/api/v1/reports/export?type=inventory-summary&product_id={pid}&start={today}&end={today}")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws.title == "模板"
    assert ws.max_row == 3  # 标题 + 表头 + 1 行数据
    assert str(next(iter(ws.merged_cells.ranges))) == "A1:P1"
    assert ws.cell(1, 1).value == f"{today[:4]}年{int(today[5:7])}月库存金额收发存表"
    assert ws.cell(1, 1).font.bold
    assert ws.cell(2, 1).value == "年月"
    assert ws.cell(2, 5).value == "物料编码"
    assert ws.cell(2, 19).value == "已使用数量"
    assert ws.cell(2, 19).font.color.rgb == "FFFF0000"  # S/T 红色表头
    assert ws.column_dimensions["I"].hidden
    assert ws.row_dimensions[1].height == 47.45
    assert float(ws.cell(3, 15).value) == 25  # 月度结存数量（O 列）
    assert float(ws.cell(3, 21).value) == 2.0  # 单价 = 50 / 25
    assert ws.cell(3, 21).number_format == "0.00_ "
    # 文件名：无括号 + 末尾追加当前时间（精确到分钟）
    cd = r.headers["content-disposition"]
    m = re.search(r"filename\*=UTF-8''([^;]+)", cd)
    fname = unquote(m.group(1)) if m else ""
    assert "库存金额收发存" in fname
    assert "（" not in fname and "）" not in fname
    assert re.search(r"_\d{4}\.xlsx$", fname)


def test_export_stock_and_flow() -> None:
    _login_admin()
    wh, loc, pid = _setup()
    _out_5(wh, loc, pid)
    r = client.get(f"/api/v1/reports/export?type=stock&warehouse_id={wh}")
    assert r.status_code == 200 and "spreadsheetml" in r.headers["content-type"]
    assert load_workbook(io.BytesIO(r.content)).active.max_row >= 2
    r = client.get(f"/api/v1/reports/export?type=flow&product_id={pid}")
    assert r.status_code == 200 and "spreadsheetml" in r.headers["content-type"]
    assert load_workbook(io.BytesIO(r.content)).active.max_row >= 2
    r = client.get("/api/v1/reports/export?type=bad")
    assert r.json()["code"] == 4006


# ============================ 2D 货架图数据源 ============================

def test_location_summary() -> None:
    _login_admin()
    wh, _, pid = _setup(qty="30", min_stock="40")
    r = client.get(f"/api/v1/stock/location-summary?warehouse_id={wh}")
    assert r.status_code == 200 and r.json()["code"] == 0, r.text
    rows = r.json()["data"]
    assert len(rows) == 1
    assert rows[0]["location_code"]
    assert rows[0]["layer_no"] == 1
    items = rows[0]["items"]
    assert len(items) == 1 and items[0]["product_id"] == pid
    assert items[0]["qty"] == "30"
    assert items[0]["alert"] == "low"  # 30 < 下限 40 → 红
    # 空库位也返回，items 为空
    r = client.post("/api/v1/shelves", json={"warehouse_id": wh, "code": "B2"})
    r = client.post("/api/v1/locations", json={"warehouse_id": wh, "shelf_id": r.json()["data"]["id"], "layer_no": 2})
    r = client.get(f"/api/v1/stock/location-summary?warehouse_id={wh}")
    assert len(r.json()["data"]) == 2
    assert r.json()["data"][1]["items"] == []


# ============================ 权限 ============================

def test_report_permission() -> None:
    _login_tester()  # 使用者角色无 report:view
    assert client.get("/api/v1/reports/dashboard").json()["code"] == 4005
    assert client.get("/api/v1/reports/export?type=stock").json()["code"] == 4005
    # /stock/summary 登录即可（手机端库存查询）
    assert client.get("/api/v1/stock/summary").json()["code"] == 0
