"""统一导出格式服务（所有表格导出的唯一入口）。

三级格式合并（优先级从高到低）：
  1. 请求级 spec —— ExportFormatModal 随请求传入的 fmt（用户当次自定义）
  2. 模块级覆盖 —— sys_config['export.module.<key>']（管理员在设置页按模块配置）
  3. 全局默认   —— sys_config['export.global']（管理员统一调整）
  4. 内置默认   —— BUILTIN_FORMAT（代码内兜底，保证零配置可用）

模块接入方式（新增导出必须走这里，禁止各自拼 openpyxl）：
    from app.services.export_service import write_table_xlsx
    return write_table_xlsx(db, module="stock_query", headers=[...], rows=[...],
                            filename="stock.xlsx", title="库存报表")

配置存储：sys_config KV（JSON 字符串）
    export.global               全局默认（部分字段即可，深合并）
    export.module.<module_key>  模块覆盖（同上，深合并到全局之上）
"""
from __future__ import annotations

import io
import json
import re
from datetime import date, datetime
from typing import Any
from urllib.parse import quote

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.excel_guard import safe_excel_value
from app.models.sys import SysConfig

EXPORT_CONFIG_KEYS: dict[str, str] = {
    "global": "export.global",
    "stock_query": "export.module.stock_query",
    "operation_logs": "export.module.operation_logs",
    "check_export": "export.module.check_export",
    "flow": "export.module.flow",
}

# 内置默认（代码兜底；管理员不可见不可改此层）
BUILTIN_FORMAT: dict[str, Any] = {
    "header": {"bg": "F6F8FE", "font": "宋体", "size": 12, "bold": True, "align": "center"},
    "body": {"font": "宋体", "size": 11},
    "columnWidth": {"mode": "auto", "fixed": 20, "min": 8, "max": 55},
    "rowHeight": {"data": 25.15, "title": 47.45},
    "dataFormat": {"dateFormat": "yyyy-mm-dd", "timeFormat": "yyyy-mm-dd hh:mm:ss",
                    "thousands": False, "decimals": 2},
    "options": {"freezeHeader": True, "autoFilter": True, "gridlines": False,
                 "pageMargin": 0.7, "printTitleRows": True},
    "longNumberAsText": True,
    "wrapText": True,
}

_LONG_DIGITS = re.compile(r"^\d{15,}$")
_SENSITIVE_KEYS = ("password", "passwd", "secret", "token", "captcha")


# ============================ 配置读取与合并 ============================

def _deep_merge(base: dict, override: dict) -> dict:
    out = dict(base)
    for k, v in override.items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = _deep_merge(out[k], v)
        else:
            out[k] = v
    return out


def _read_config_json(db: Session, key: str) -> dict | None:
    from sqlalchemy import select

    row = db.scalar(select(SysConfig).where(SysConfig.config_key == key))
    if row is None or not row.config_value:
        return None
    try:
        obj = json.loads(row.config_value)
        return obj if isinstance(obj, dict) else None
    except ValueError:
        return None


def get_effective_format(db: Session, module_key: str | None = None) -> dict:
    """内置 ← 全局 ← 模块，逐层深合并。module_key 不在登记表中时忽略模块层。"""
    merged = BUILTIN_FORMAT
    glob = _read_config_json(db, EXPORT_CONFIG_KEYS["global"])
    if glob:
        merged = _deep_merge(merged, glob)
    if module_key and module_key in EXPORT_CONFIG_KEYS and module_key != "global":
        mod = _read_config_json(db, EXPORT_CONFIG_KEYS[module_key])
        if mod:
            merged = _deep_merge(merged, mod)
    return merged


def get_all_formats(db: Session) -> dict:
    """管理界面读取：全局已存配置 + 各模块已存覆盖 + 内置默认（供前端回显合并结果）。"""
    return {
        "builtin": BUILTIN_FORMAT,
        "global": _read_config_json(db, EXPORT_CONFIG_KEYS["global"]) or {},
        "modules": {k: (_read_config_json(db, EXPORT_CONFIG_KEYS[k]) or None)
                     for k in ("stock_query", "operation_logs", "check_export", "flow")},
        "effective": {k: get_effective_format(db, k if k != "global" else None)
                       for k in ("stock_query", "operation_logs", "check_export", "flow")},
    }


def save_global_format(db: Session, cfg: dict) -> None:
    _upsert_config(db, EXPORT_CONFIG_KEYS["global"], json.dumps(cfg, ensure_ascii=False))


def save_module_format(db: Session, module_key: str, cfg: dict | None) -> None:
    key = EXPORT_CONFIG_KEYS.get(module_key)
    if key is None or module_key == "global":
        raise ValueError("未知模块标识")
    _upsert_config(db, key, json.dumps(cfg, ensure_ascii=False) if cfg else "")


def _upsert_config(db: Session, key: str, value: str) -> None:
    from sqlalchemy import select

    row = db.scalar(select(SysConfig).where(SysConfig.config_key == key))
    if row is None:
        db.add(SysConfig(config_key=key, config_value=value))
    else:
        row.config_value = value
    db.commit()


# ============================ 渲染 ============================

def _num(v):
    """数值列：Decimal/int/float 原样（供 number_format 生效），其余原值。"""
    if isinstance(v, bool):
        return v
    return v


def _cell_value(v, fmt: dict, header_hint: str = ""):
    """按全局数据格式约定转换单元格值；返回 (value, number_format|None)。"""
    s = safe_excel_value(v)
    date_fmt = fmt["dataFormat"]["dateFormat"]
    time_fmt = fmt["dataFormat"]["timeFormat"]
    thousands = fmt["dataFormat"]["thousands"]
    decimals = int(fmt["dataFormat"].get("decimals", 2))

    if isinstance(v, (datetime,)):
        return v, time_fmt
    if isinstance(v, (date,)):
        return v, date_fmt
    if isinstance(s, str):
        # ≥15 位纯数字强制文本（防科学计数法/精度丢失）
        if fmt["longNumberAsText"] and _LONG_DIGITS.match(s.strip()):
            return s, "@"
        if looks_like_time := bool(re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", s)):
            return s, time_fmt
        if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
            return s, date_fmt
        return s, None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        nf = f"#{',##0.' + '0' * decimals}" if thousands else f"0.{('0' * decimals)}" if decimals else "0"
        return v, nf
    return s, None


def write_table_xlsx(
    db: Session,
    module: str,
    *,
    headers: list[str],
    rows: list[list],
    filename: str,
    sheet: str = "报表",
    title: str | None = None,
    column_widths: list[float] | None = None,
    request_spec: dict | None = None,
) -> StreamingResponse:
    """统一表格导出入口。

    - module：模块标识（决定模块级格式覆盖）
    - request_spec：请求级格式（ExportFormatModal fmt 参数），叠加在配置之上（键为列位置）
      形如 {"colWidths": {"1": 30}, "textCols": [0], "numberCols": {"3": 2}}
    """
    fmt = get_effective_format(db, module)
    req = request_spec or {}
    text_cols = set(req.get("textCols") or [])
    number_cols: dict[int, int] = {int(k): v for k, v in (req.get("numberCols") or {}).items()}
    req_widths: dict[int, float] = {int(k): float(v) for k, v in (req.get("colWidths") or {}).items()}

    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.sheet_view.showGridLines = bool(fmt["options"]["gridlines"])

    h = fmt["header"]
    head_font = Font(name=h["font"], size=h["size"], bold=h["bold"])
    head_fill = PatternFill("solid", fgColor=(h["bg"] or "#FFFFFF").lstrip("#"))
    head_align = Alignment(horizontal=h["align"], vertical="center", wrap_text=bool(fmt["wrapText"]))
    body_font = Font(name=fmt["body"]["font"], size=fmt["body"]["size"])
    body_align = Alignment(vertical="center", wrap_text=bool(fmt["wrapText"]), shrink_to_fit=False)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        c = ws.cell(1, 1, title)
        c.font = Font(name=fmt["header"]["font"], size=fmt["rowHeight"] and 14, bold=True)
        c.alignment, c.border = head_align, border
        ws.row_dimensions[1].height = fmt["rowHeight"]["title"]
        row = 2

    for ci, name in enumerate(headers, 1):
        c = ws.cell(row, ci, name)
        c.font, c.fill, c.alignment, c.border = head_font, head_fill, head_align, border
    ws.row_dimensions[row].height = 30

    for r in rows:
        row += 1
        ws.row_dimensions[row].height = fmt["rowHeight"]["data"]
        for ci, v in enumerate(r, 1):
            value, nf = _cell_value(v, fmt)
            if (ci - 1) in text_cols:
                value, nf = "" if v is None else str(v), "@"
            elif (ci - 1) in number_cols:
                d = number_cols[ci - 1]
                value = float(v) if v not in (None, "") else 0
                nf = f"0.{('0' * d)}" if d else "0"
            c = ws.cell(row, ci, value)
            c.font, c.alignment, c.border = body_font, body_align, border
            if nf:
                c.number_format = nf

    # 列宽：请求级手动 > 模块 fixed > 自适应(min~max)
    cw = fmt["columnWidth"]
    min_w, max_w = float(cw.get("min", 8)), float(cw.get("max", 55))
    for ci in range(1, len(headers) + 1):
        if (ci - 1) in req_widths:
            width = req_widths[ci - 1]
        elif column_widths and ci <= len(column_widths):
            width = column_widths[ci - 1]
        elif cw["mode"] == "fixed":
            width = float(cw.get("fixed", 20))
        else:
            samples = [str(headers[ci - 1])] + [str(r[ci - 1]) for r in rows[:300] if ci - 1 < len(r)]
            content_max = max((len(x) for x in samples), default=8)
            width = max(min_w, min(max_w, content_max + 4))
        ws.column_dimensions[get_column_letter(ci)].width = width

    if fmt["options"]["freezeHeader"]:
        ws.freeze_panes = f"A{row + 1}"
    if fmt["options"]["autoFilter"]:
        last = row + max(len(rows), 1)
        ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{last}"
    ws.page_setup.left_margin = ws.page_setup.right_margin = float(fmt["options"]["pageMargin"])
    if fmt["options"]["printTitleRows"]:
        ws.print_title_rows = f"{row}:{row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{quote(filename)}"'},
    )
