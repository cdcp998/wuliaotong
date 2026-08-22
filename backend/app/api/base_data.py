"""基础资料接口：分类/单位/供应商/商品/仓库/货架/库位 + Excel 导入导出（《后端API设计.md》§2）。

权限点：base:category / base:product / base:supplier / base:warehouse / base:stock-location。
删除语义：基础资料不做物理删除，DELETE = 停用（status=0）；有引用关系的禁止停用。
"""
from __future__ import annotations

import io
import re
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import func, or_, select, text, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.cache import cache_aside, cache_delete, cache_delete_pattern
from app.core.deps import (
    SUPER_ADMIN_ROLE_CODE,
    get_current_user,
    require_any_permission,
    require_manager_role,
    require_permission,
)
from app.core.excel_guard import safe_excel_value
from app.core.response import BizError, E_NOT_FOUND, E_PARAM, ok
from app.db import get_db
from app.models.base import (
    BaseCategory,
    BaseDepartment,
    BaseDepartmentShelf,
    BaseLocation,
    BaseProduct,
    BaseProductSupplier,
    BaseProductUnit,
    BaseShelf,
    BaseSupplier,
    BaseUnit,
    BaseWarehouse,
)
from app.models.stock import PchPurchaseIn, StkStock
from app.models.sys import SysDeleteReview, SysNotification, SysRole, SysUser
from app.schemas.admin import DeptOut, DeptReq, DeptShelvesReq, DeptUpdateReq
from app.schemas.base import (
    CategoryNode,
    CategoryReq,
    DeleteReviewOut,
    DeleteReviewRejectReq,
    DeleteReviewReq,
    LocationOut,
    LocationReq,
    PageData,
    ProductCategoryReq,
    ProductOut,
    ProductReq,
    ShelfOut,
    ShelfReq,
    SupplierOut,
    SupplierReq,
    UnitOut,
    UnitReq,
    WarehouseOut,
    WarehouseReq,
    WarehouseUpdateReq,
)

router = APIRouter(tags=["基础资料"], dependencies=[Depends(get_current_user)])
# 静态路径路由（/products/import-template、/products/export）必须优先于 /products/{product_id} 注册
static_router = APIRouter(tags=["基础资料"], dependencies=[Depends(get_current_user)])

_DECIMAL_RE = re.compile(r"^\d+(\.\d+)?$")

# 缓存 TTL（秒）：基础资料字典 10 分钟，商品详情/条码查询 10 分钟（写操作即时失效）
_DICT_TTL = 600
_PRODUCT_TTL = 600

# 商品导入列（与模板/导出一致，顺序固定）
PRODUCT_IMPORT_COLUMNS = ["编码", "条码", "SKU", "名称", "分类", "规格", "单位", "进价", "下限", "上限"]


def _parse_dec(v: str, field: str) -> Decimal:
    if not _DECIMAL_RE.match(v):
        raise BizError(E_PARAM, f"{field} 必须是数字")
    return Decimal(v)


def _category_path(db: Session, parent_id: int) -> str:
    if parent_id == 0:
        return "/"
    parent = db.get(BaseCategory, parent_id)
    if parent is None:
        raise BizError(E_PARAM, "父分类不存在")
    return f"{parent.path}{parent.id}/"


def _rebuild_path(db: Session, cat: BaseCategory) -> None:
    """重算分类 path（含子孙）。"""
    cat.path = _category_path(db, cat.parent_id)
    children = db.scalars(select(BaseCategory).where(BaseCategory.parent_id == cat.id)).all()
    for child in children:
        _rebuild_path(db, child)


def _subtree_category_ids(db: Session, cat: BaseCategory) -> list[int]:
    """分类子树全部 id（含自身）：path 只存祖先链（如 "/2/5/"），子孙 path LIKE "path/自身id/%"。"""
    prefix = f"{cat.path}{cat.id}/"
    ids = list(db.scalars(select(BaseCategory.id).where(BaseCategory.path.like(prefix + "%"))).all())
    return [cat.id] + ids


# ============================ 分类 ============================


@router.get("/categories")
def list_categories(db: Session = Depends(get_db)) -> dict:
    def _load() -> list[dict]:
        cats = db.scalars(
            select(BaseCategory).order_by(BaseCategory.sort, BaseCategory.id)
        ).all()
        nodes: dict[int, dict] = {c.id: {"id": c.id, "parent_id": c.parent_id, "name": c.name, "sort": c.sort, "children": []} for c in cats}
        tree: list[dict] = []
        for c in cats:
            node = nodes[c.id]
            if c.parent_id and c.parent_id in nodes:
                nodes[c.parent_id]["children"].append(node)
            else:
                tree.append(node)
        return tree

    tree: list[dict] = cache_aside("dict:categories", _DICT_TTL, _load)

    # 已挂材料数实时统计（不进树缓存，避免创建/改挂材料后数字滞后）
    counts = dict(
        db.execute(
            select(BaseProduct.category_id, func.count())
            .where(BaseProduct.category_id > 0, BaseProduct.status == 1)
            .group_by(BaseProduct.category_id)
        ).all()
    )

    def _attach(nodes: list[dict]) -> None:
        for n in nodes:
            n["product_count"] = counts.get(n["id"], 0)
            if n.get("children"):
                _attach(n["children"])

    _attach(tree)
    return ok(tree)


@router.post("/categories", dependencies=[Depends(require_any_permission("base:category", "ai:suggestion"))])
def create_category(req: CategoryReq, db: Session = Depends(get_db)) -> dict:
    if req.parent_id:
        parent = db.get(BaseCategory, req.parent_id)
        if parent is None:
            raise BizError(E_PARAM, "父分类不存在")
        if _cat_depth(parent) >= 3:
            raise BizError(E_PARAM, "分类最多三级，三级分类下不能再建子分类")
        # 规则：二级分类已挂材料时不能再建子分类（挂材料与建子分类二选一）
        if parent.parent_id != 0 and _product_cnt(db, parent.id) > 0:
            raise BizError(E_PARAM, "该分类已挂载材料，不能再创建子分类（请先取消挂载材料或另建分类）", http_status=409)
    cat = BaseCategory(
        parent_id=req.parent_id,
        name=req.name,
        sort=req.sort,
        path=_category_path(db, req.parent_id),
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)
    cache_delete("dict:categories")  # 分类树缓存失效
    return ok(CategoryNode(id=cat.id, parent_id=cat.parent_id, name=cat.name, sort=cat.sort).model_dump())


@router.put("/categories/{cat_id}", dependencies=[Depends(require_any_permission("base:category", "ai:suggestion"))])
def update_category(cat_id: int, req: CategoryReq, db: Session = Depends(get_db)) -> dict:
    cat = db.get(BaseCategory, cat_id)
    if cat is None:
        raise BizError(E_NOT_FOUND, "分类不存在")
    if req.parent_id == cat_id:
        raise BizError(E_PARAM, "父分类不能是自己")
    if req.parent_id != cat.parent_id:
        parent = db.get(BaseCategory, req.parent_id) if req.parent_id else None
        if req.parent_id and parent is None:
            raise BizError(E_PARAM, "父分类不存在")
        if parent and _cat_depth(parent) >= 3:
            raise BizError(E_PARAM, "分类最多三级，三级分类下不能再建子分类")
        # 规则：二级分类已挂材料时不能再建子分类（防止通过移动绕过创建校验）
        if parent and parent.parent_id != 0 and _product_cnt(db, parent.id) > 0:
            raise BizError(E_PARAM, "该分类已挂载材料，不能再创建子分类（请先取消挂载材料或另建分类）", http_status=409)
        # 父分类不能是自己的子孙（防环）；移动后最深子孙不得超过三级
        if parent and _is_descendant(db, parent.id, cat.id):
            raise BizError(E_PARAM, "父分类不能是自己的子分类")
        new_depth = (_cat_depth(parent) if parent else 0) + 1
        if new_depth + _subtree_height(db, cat) - 1 > 3:
            raise BizError(E_PARAM, "该分类下有子分类，移动后层级将超过三级")
    cat.parent_id = req.parent_id
    cat.name = req.name
    cat.sort = req.sort
    _rebuild_path(db, cat)
    db.commit()
    cache_delete("dict:categories")  # 分类树缓存失效
    return ok()


@router.delete("/categories/{cat_id}", dependencies=[Depends(require_permission("base:category"))])
def delete_category(cat_id: int, db: Session = Depends(get_db)) -> dict:
    cat = db.get(BaseCategory, cat_id)
    if cat is None:
        raise BizError(E_NOT_FOUND, "分类不存在", http_status=404)
    child_cnt = db.scalar(select(func.count()).select_from(BaseCategory).where(BaseCategory.parent_id == cat_id)) or 0
    product_cnt = db.scalar(select(func.count()).select_from(BaseProduct).where(BaseProduct.category_id == cat_id)) or 0
    if child_cnt or product_cnt:
        raise BizError(E_PARAM, "分类下存在子分类或商品，禁止删除", http_status=409)
    db.delete(cat)
    db.commit()
    cache_delete("dict:categories")  # 分类树缓存失效
    return ok()


def _product_cnt(db: Session, category_id: int) -> int:
    """分类下启用材料数（与 /categories 的 product_count 同口径）。"""
    return db.scalar(
        select(func.count()).select_from(BaseProduct).where(
            BaseProduct.category_id == category_id, BaseProduct.status == 1
        )
    ) or 0


def _cat_depth(cat: BaseCategory) -> int:
    """分类层级：path 每段一个祖先（"/" 顶级=1、"/a/" 二级=2、"/a/b/" 三级=3）。"""
    return len([p for p in cat.path.split("/") if p]) + 1


def _is_descendant(db: Session, node_id: int, ancestor_id: int) -> bool:
    """node_id 是否为 ancestor_id 的子孙。"""
    node = db.get(BaseCategory, node_id)
    cur = node.parent_id if node else 0
    while cur:
        if cur == ancestor_id:
            return True
        cur_node = db.get(BaseCategory, cur)
        cur = cur_node.parent_id if cur_node else 0
    return False


def _subtree_height(db: Session, cat: BaseCategory) -> int:
    """以 cat 为根的子分类树最大深度（根记 1）。"""
    children = db.scalars(select(BaseCategory).where(BaseCategory.parent_id == cat.id)).all()
    if not children:
        return 1
    return 1 + max(_subtree_height(db, c) for c in children)


def _require_leaf_category(db: Session, category_id: int) -> None:
    """材料分类规则（三级体系）：材料只能挂二级/三级分类（parent_id != 0），顶级分类仅作分组。"""
    cat = db.get(BaseCategory, category_id)
    if cat is None:
        raise BizError(E_NOT_FOUND, "分类不存在")
    if cat.parent_id == 0:
        raise BizError(E_PARAM, "顶级分类仅作分组，材料请挂到其二级或三级子分类")


# ============================ 单位 ============================


@router.get("/units")
def list_units(db: Session = Depends(get_db)) -> dict:
    def _load() -> list[dict]:
        units = db.scalars(select(BaseUnit).order_by(BaseUnit.id)).all()
        return [UnitOut(id=u.id, name=u.name, remark=u.remark).model_dump() for u in units]

    return ok(cache_aside("dict:units", _DICT_TTL, _load))



# 垃圾单位名：短名（≤6 字）+ 6 位 hex 随机尾（测试/误操作产生），禁止入库
_UNIT_NAME_GARBAGE = re.compile(r"^.{1,6}[0-9a-f]{6}$")


def _check_unit_name(name: str) -> str:
    name = name.strip()
    if not name or len(name) > 20:
        raise BizError(E_PARAM, "单位名称长度须为 1-20 字")
    if _UNIT_NAME_GARBAGE.match(name):
        raise BizError(E_PARAM, "单位名称不规范（疑似随机后缀），请使用规范单位名")
    return name


@router.post("/units", dependencies=[Depends(require_permission("base:product"))])
def create_unit(req: UnitReq, db: Session = Depends(get_db)) -> dict:
    name = _check_unit_name(req.name)
    if db.scalar(select(BaseUnit.id).where(BaseUnit.name == name)):
        raise BizError(E_PARAM, "单位已存在")
    unit = BaseUnit(name=name, remark=req.remark)
    db.add(unit)
    db.commit()
    db.refresh(unit)
    cache_delete("dict:units")
    return ok(UnitOut(id=unit.id, name=unit.name, remark=unit.remark).model_dump())


@router.put("/units/{unit_id}", dependencies=[Depends(require_permission("base:product"))])
def update_unit(unit_id: int, req: UnitReq, db: Session = Depends(get_db)) -> dict:
    unit = db.get(BaseUnit, unit_id)
    if unit is None:
        raise BizError(E_NOT_FOUND, "单位不存在")
    name = _check_unit_name(req.name)
    exists = db.scalar(select(BaseUnit.id).where(BaseUnit.name == name, BaseUnit.id != unit_id))
    if exists:
        raise BizError(E_PARAM, "单位已存在")
    unit.name = name
    unit.remark = req.remark
    db.commit()
    cache_delete("dict:units")
    return ok()


@router.delete("/units/{unit_id}", dependencies=[Depends(require_permission("base:product"))])
def delete_unit(unit_id: int, db: Session = Depends(get_db)) -> dict:
    unit = db.get(BaseUnit, unit_id)
    if unit is None:
        raise BizError(E_NOT_FOUND, "单位不存在")
    used = db.scalar(select(func.count()).select_from(BaseProductUnit).where(BaseProductUnit.unit_id == unit_id)) or 0
    # 产品主单位（base_product.unit_id）也计入引用，否则删除后产生悬空引用
    used += db.scalar(select(func.count()).select_from(BaseProduct).where(BaseProduct.unit_id == unit_id)) or 0
    if used:
        raise BizError(E_PARAM, "单位已被商品引用，禁止删除")
    db.delete(unit)
    db.commit()
    cache_delete("dict:units")
    return ok()


# ============================ 供应商 ============================


@router.get("/suppliers")
def list_suppliers(
    keyword: str = Query("", max_length=100),
    status: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    # 下拉场景（无关键词，取前 100 条）走缓存；带关键词的搜索不缓存（低频率、参数多变）
    if not keyword and page == 1:
        key = f"dict:suppliers:{status or 'all'}:{page_size}"

        def _load() -> dict:
            stmt = select(BaseSupplier)
            if status is not None:
                stmt = stmt.where(BaseSupplier.status == status)
            total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
            rows = db.scalars(stmt.order_by(BaseSupplier.id.desc()).limit(page_size)).all()
            return PageData(
                list=[SupplierOut.model_validate(s, from_attributes=True).model_dump() for s in rows],
                total=total, page=page, page_size=page_size,
            ).model_dump()

        return ok(cache_aside(key, _DICT_TTL, _load))

    stmt = select(BaseSupplier)
    if keyword:
        stmt = stmt.where(or_(BaseSupplier.name.like(f"%{keyword}%"), BaseSupplier.code.like(f"%{keyword}%")))
    if status is not None:
        stmt = stmt.where(BaseSupplier.status == status)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(BaseSupplier.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    return ok(PageData(
        list=[SupplierOut.model_validate(s, from_attributes=True).model_dump() for s in rows],
        total=total, page=page, page_size=page_size,
    ).model_dump())


@router.post("/suppliers", dependencies=[Depends(require_permission("base:supplier"))])
def create_supplier(req: SupplierReq, db: Session = Depends(get_db)) -> dict:
    if db.scalar(select(BaseSupplier.id).where(BaseSupplier.code == req.code)):
        raise BizError(E_PARAM, "供应商编码已存在")
    sup = BaseSupplier(**req.model_dump())
    db.add(sup)
    db.commit()
    db.refresh(sup)
    cache_delete_pattern("dict:suppliers*")
    return ok(SupplierOut.model_validate(sup, from_attributes=True).model_dump())


@router.put("/suppliers/{sup_id}", dependencies=[Depends(require_permission("base:supplier"))])
def update_supplier(sup_id: int, req: SupplierReq, db: Session = Depends(get_db)) -> dict:
    sup = db.get(BaseSupplier, sup_id)
    if sup is None:
        raise BizError(E_NOT_FOUND, "供应商不存在")
    exists = db.scalar(select(BaseSupplier.id).where(BaseSupplier.code == req.code, BaseSupplier.id != sup_id))
    if exists:
        raise BizError(E_PARAM, "供应商编码已存在")
    for k, v in req.model_dump().items():
        setattr(sup, k, v)
    db.commit()
    cache_delete_pattern("dict:suppliers*")
    return ok()


@router.delete("/suppliers/{sup_id}", dependencies=[Depends(require_permission("base:supplier"))])
def delete_supplier(sup_id: int, db: Session = Depends(get_db)) -> dict:
    sup = db.get(BaseSupplier, sup_id)
    if sup is None:
        raise BizError(E_NOT_FOUND, "供应商不存在")
    # 数据一致性：仍有启用材料关联时禁止停用，避免材料-供应商关系悬空
    linked = db.scalar(
        select(func.count())
        .select_from(BaseProductSupplier)
        .join(BaseProduct, BaseProduct.id == BaseProductSupplier.product_id)
        .where(BaseProductSupplier.supplier_id == sup_id, BaseProduct.status == 1)
    ) or 0
    if linked:
        raise BizError(E_PARAM, f"仍有 {linked} 个启用材料关联该供应商，请先在材料中解除关联")
    db.execute(BaseProductSupplier.__table__.delete().where(BaseProductSupplier.supplier_id == sup_id))  # 清理残留关联（停用材料的）
    sup.status = 0  # 软删除：停用
    db.commit()
    cache_delete_pattern("dict:suppliers*")
    return ok()


@router.post("/suppliers/merge", dependencies=[Depends(require_permission("base:supplier"))])
def merge_suppliers(body: dict, db: Session = Depends(get_db)) -> dict:
    """合并供应商（人工确认）：from 的关联材料/入库单转移至 to，from 停用（不物理删除）。"""
    from_id = int(body.get("from_id") or 0)
    to_id = int(body.get("to_id") or 0)
    if not from_id or not to_id or from_id == to_id:
        raise BizError(E_PARAM, "合并参数无效（from_id/to_id 必须不同）")
    a = db.get(BaseSupplier, from_id)
    b = db.get(BaseSupplier, to_id)
    if a is None or b is None:
        raise BizError(E_NOT_FOUND, "供应商不存在")
    # 材料关联转移（目标已有关联则删除重复）
    for link in db.scalars(select(BaseProductSupplier).where(BaseProductSupplier.supplier_id == a.id)).all():
        exists = db.scalar(
            select(BaseProductSupplier.id).where(
                BaseProductSupplier.supplier_id == b.id, BaseProductSupplier.product_id == link.product_id
            )
        )
        if exists:
            db.delete(link)
        else:
            link.supplier_id = b.id
    # 历史入库单归属转移
    db.execute(update(PchPurchaseIn).where(PchPurchaseIn.supplier_id == a.id).values(supplier_id=b.id))
    a.status = 0  # 停用被合并的供应商
    db.commit()
    cache_delete_pattern("dict:suppliers*")  # 供应商下拉缓存失效
    return ok({"merged_id": b.id})


@router.get("/suppliers/{sup_id}/products")
def supplier_products(sup_id: int, db: Session = Depends(get_db)) -> dict:
    """供应商信息中查看关联材料（含停用材料，按材料 id 倒序）。"""
    if db.get(BaseSupplier, sup_id) is None:
        raise BizError(E_NOT_FOUND, "供应商不存在")
    rows = db.scalars(
        select(BaseProduct)
        .join(BaseProductSupplier, BaseProductSupplier.product_id == BaseProduct.id)
        .where(BaseProductSupplier.supplier_id == sup_id)
        .order_by(BaseProduct.id.desc())
    ).all()
    return ok({"list": [_product_out(db, p) for p in rows], "total": len(rows)})


SUPPLIER_IMPORT_COLUMNS = ["编码", "名称", "联系人", "电话", "地址"]

_MAX_IMPORT_BYTES = 10 * 1024 * 1024  # 导入文件大小上限 10MB（防内存耗尽）
_MAX_IMPORT_ROWS = 5000  # 单次导入行数上限


@router.post("/suppliers/import", dependencies=[Depends(require_permission("base:supplier"))])
async def import_suppliers(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    data = await file.read()
    if len(data) > _MAX_IMPORT_BYTES:
        raise BizError(E_PARAM, "导入文件不能超过 10MB")
    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise BizError(E_PARAM, "文件为空")
    if len(rows) > _MAX_IMPORT_ROWS:
        raise BizError(E_PARAM, f"单次导入不能超过 {_MAX_IMPORT_ROWS} 行")
    headers = [str(h).strip() if h else "" for h in rows[0]]
    if headers[:5] != SUPPLIER_IMPORT_COLUMNS:
        raise BizError(E_PARAM, f"表头必须为：{'/'.join(SUPPLIER_IMPORT_COLUMNS)}")
    success, fail_rows = 0, []
    for idx, row in enumerate(rows[1:], start=2):
        vals = [str(v).strip() if v is not None else "" for v in row] + [""] * 5
        code, name = vals[0], vals[1]
        if not name:
            fail_rows.append({"row": idx, "reason": "名称为空"})
            continue
        if not code:
            code = "SUP" + uuid.uuid4().hex[:6].upper()
        if db.scalar(select(BaseSupplier.id).where(BaseSupplier.code == code)):
            fail_rows.append({"row": idx, "reason": f"编码 {code} 已存在"})
            continue
        db.add(BaseSupplier(code=code, name=name, contact=vals[2], phone=vals[3], address=vals[4]))
        success += 1
    db.commit()
    if success:
        cache_delete_pattern("dict:suppliers*")
    return ok({"success_count": success, "fail_rows": fail_rows})


# ============================ 商品 ============================


def _next_numeric_code(db: Session) -> str:
    """生成下一个纯数字商品编码（当前最大数字编码 + 1，空表从 1 开始）。"""
    v = db.execute(text("SELECT MAX(CAST(code AS UNSIGNED)) FROM base_product WHERE code REGEXP '^[0-9]+$'")).scalar()
    return str((v or 0) + 1)


def _product_out(db: Session, p: BaseProduct) -> dict:
    cat = db.get(BaseCategory, p.category_id)
    unit = db.get(BaseUnit, p.unit_id)
    units = db.scalars(
        select(BaseProductUnit).where(BaseProductUnit.product_id == p.id).order_by(BaseProductUnit.is_default.desc())
    ).all()
    # 关联供应商（按 id 升序，保证输出稳定）
    sups = db.scalars(
        select(BaseSupplier)
        .join(BaseProductSupplier, BaseProductSupplier.supplier_id == BaseSupplier.id)
        .where(BaseProductSupplier.product_id == p.id)
        .order_by(BaseSupplier.id)
    ).all()
    return ProductOut(
        id=p.id, code=p.code, material_code=p.material_code, barcode=p.barcode, sku=p.sku, name=p.name,
        category_id=p.category_id, category_name=cat.name if cat else "",
        spec=p.spec, unit_id=p.unit_id, unit_name=unit.name if unit else "",
        purchase_price=p.purchase_price, min_stock=p.min_stock, max_stock=p.max_stock,
        status=p.status, remark=p.remark, created_at=p.created_at,
        units=[{"id": u.id, "unit_id": u.unit_id, "unit_name": (db.get(BaseUnit, u.unit_id).name if db.get(BaseUnit, u.unit_id) else ""), "rate": format(u.rate, "f"), "is_default": u.is_default} for u in units],
        supplier_ids=[s.id for s in sups], supplier_names=[s.name for s in sups],
    ).model_dump()


def _apply_suppliers(db: Session, product_id: int, supplier_ids: list[int]) -> None:
    """写入材料-供应商关联（全量替换，与 _apply_units 同模式）。"""
    db.execute(BaseProductSupplier.__table__.delete().where(BaseProductSupplier.product_id == product_id))
    seen: set[int] = set()
    for sid in supplier_ids:
        if sid <= 0 or sid in seen:
            continue
        if db.get(BaseSupplier, sid) is None:
            raise BizError(E_PARAM, f"供应商 id={sid} 不存在")
        seen.add(sid)
        db.add(BaseProductSupplier(product_id=product_id, supplier_id=sid))


def _apply_units(db: Session, product_id: int, unit_id: int, units: list) -> None:
    """写入多单位换算：默认单位与 unit_id 一致（不存在则自动补）。

    units 元素为 dict（来自 schema model_dump）。
    """
    db.execute(BaseProductUnit.__table__.delete().where(BaseProductUnit.product_id == product_id))
    items = [u for u in units if (u.get("unit_id") or 0) > 0]
    if not any(u.get("is_default") for u in items):
        items.append({"unit_id": unit_id, "rate": "1", "is_default": 1})
    for u in items:
        uid = u["unit_id"]
        if db.get(BaseUnit, uid) is None:
            raise BizError(E_PARAM, f"单位 id={uid} 不存在")
        db.add(
            BaseProductUnit(
                product_id=product_id,
                unit_id=uid,
                rate=_parse_dec(u.get("rate", "1"), "换算率"),
                is_default=1 if u.get("is_default") else 0,
            )
        )


def _expand_keywords_local(keyword: str, max_kw: int = 25) -> list[str]:
    """本地语义扩展（P9-P2⑧，不调大模型）：去常见口语修饰词后，取 2-4 字连续子串做宽匹配。

    例：「测网络的工具」→ 去「的」→「测网络工具」→ 子串含「网络」「测试」「工具」→ 命中「网络测试仪」。
    """
    k = re.sub(r"[的了我你要买找查哪种给个是么]", "", keyword)
    if len(k) < 2:
        return []
    out: list[str] = []
    for w in range(min(4, len(k)), 1, -1):  # 优先长串（更可能是有意义词）
        for i in range(len(k) - w + 1):
            sub = k[i : i + w]
            if sub not in out:
                out.append(sub)
            if len(out) >= max_kw:
                break
        if len(out) >= max_kw:
            break
    return out


@router.get("/products")
def list_products(
    keyword: str = Query("", max_length=100),
    category_id: int = Query(0),
    descendants: int = Query(0, description="1 时 category_id 按子树过滤（含全部子孙分类，物料数据管理选中顶级分类用）"),
    uncategorized: int = Query(0, description="1 时过滤未分类材料（category_id=0，物料数据管理「未分类」入口）"),
    barcode: str = Query("", max_length=50),
    status: int = Query(1, description="1 启用（默认） / 0 停用；全部数据见导出接口"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    ai: int = Query(0, description="1 无结果时用大模型改写关键词重试（语义搜索）"),
    db: Session = Depends(get_db),
) -> dict:
    # 扫码场景：条码精确查询（无关键词）走缓存，命中直接返回分页结果
    if barcode and not keyword and status is not None:
        key = f"product:bc:{barcode.strip()}:{status}"

        def _load() -> dict:
            stmt = select(BaseProduct).where(BaseProduct.barcode == barcode.strip())
            if status is not None:
                stmt = stmt.where(BaseProduct.status == status)
            total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
            rows = db.scalars(stmt.order_by(BaseProduct.id.desc()).limit(page_size)).all()
            return PageData(
                list=[_product_out(db, p) for p in rows],
                total=total, page=page, page_size=page_size,
            ).model_dump()

        return ok(cache_aside(key, _PRODUCT_TTL, _load))

    stmt = select(BaseProduct)
    if keyword:
        like = f"%{keyword}%"
        stmt = stmt.where(or_(BaseProduct.name.like(like), BaseProduct.code.like(like), BaseProduct.material_code.like(like), BaseProduct.spec.like(like), BaseProduct.sku.like(like), BaseProduct.barcode.like(like)))
    if uncategorized:
        stmt = stmt.where(BaseProduct.category_id == 0)
    elif category_id:
        if descendants:
            # 子树聚合：该分类及全部子孙分类下的材料（path LIKE 一次查询）
            cat = db.get(BaseCategory, category_id)
            if cat is None:
                stmt = stmt.where(BaseProduct.category_id == -1)  # 分类不存在 → 空结果
            else:
                subtree = _subtree_category_ids(db, cat)
                stmt = stmt.where(BaseProduct.category_id.in_(subtree))
        else:
            stmt = stmt.where(BaseProduct.category_id == category_id)
    if barcode:
        stmt = stmt.where(BaseProduct.barcode == barcode)
    if status is not None:
        stmt = stmt.where(BaseProduct.status == status)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    ai_keywords: list[str] = []
    # 语义搜索（P9-P2⑧）：关键词无结果 → 本地子串扩展重查（不调大模型，毫秒级）
    # 注意：扩展时重建 stmt（去掉原 keyword 条件，否则 AND 叠加原条件仍无结果）
    if total == 0 and ai == 1 and keyword:
        ai_keywords = _expand_keywords_local(keyword)
        if ai_keywords:
            base = select(BaseProduct)
            if uncategorized:
                base = base.where(BaseProduct.category_id == 0)
            elif category_id:
                if descendants:
                    cat = db.get(BaseCategory, category_id)
                    if cat is not None:
                        base = base.where(BaseProduct.category_id.in_(_subtree_category_ids(db, cat)))
                else:
                    base = base.where(BaseProduct.category_id == category_id)
            if barcode:
                base = base.where(BaseProduct.barcode == barcode)
            if status is not None:
                base = base.where(BaseProduct.status == status)
            conds = []
            for k in ai_keywords:
                like = f"%{k}%"
                conds.append(or_(BaseProduct.name.like(like), BaseProduct.code.like(like), BaseProduct.material_code.like(like), BaseProduct.spec.like(like), BaseProduct.sku.like(like), BaseProduct.barcode.like(like)))
            if conds:
                stmt = base.where(or_(*conds))
                total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(BaseProduct.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    # 全仓库存合计（数量列）：单条按页内商品 id 分组查询，避免 N+1
    stock_qty_map: dict[int, Decimal] = {}
    if rows:
        stock_qty_map = dict(
            db.execute(
                select(StkStock.product_id, func.sum(StkStock.qty))
                .where(StkStock.product_id.in_([p.id for p in rows]))
                .group_by(StkStock.product_id)
            ).all()
        )
    data = PageData(
        list=[
            {**_product_out(db, p), "stock_qty": str(stock_qty_map.get(p.id, Decimal(0)))}
            for p in rows
        ],
        total=total, page=page, page_size=page_size,
    ).model_dump()
    if ai_keywords:
        data["ai_expanded"] = True
        data["ai_keywords"] = ai_keywords
    return ok(data)


@router.post("/products", dependencies=[Depends(require_permission("base:product"))])
def create_product(req: ProductReq, db: Session = Depends(get_db)) -> dict:
    code = req.code or _next_numeric_code(db)  # 商品编码纯数字，留空自动生成
    if db.scalar(select(BaseProduct.id).where(BaseProduct.code == code)):
        raise BizError(E_PARAM, "商品编码已存在")
    if db.get(BaseUnit, req.unit_id) is None:
        raise BizError(E_PARAM, "基本单位不存在")
    if req.category_id:
        _require_leaf_category(db, req.category_id)
    if req.barcode.strip() and db.scalar(select(BaseProduct.id).where(BaseProduct.barcode == req.barcode.strip())):
        raise BizError(E_PARAM, "条码已存在，请勿重复录入")
    p = BaseProduct(
        code=code, material_code=req.material_code, barcode=req.barcode.strip(), sku=req.sku, name=req.name,
        category_id=req.category_id, spec=req.spec, unit_id=req.unit_id,
        purchase_price=_parse_dec(req.purchase_price, "进价"),
        min_stock=_parse_dec(req.min_stock, "下限"), max_stock=_parse_dec(req.max_stock, "上限"),
        image_file_id=req.image_file_id, status=req.status, remark=req.remark,
    )
    db.add(p)
    db.flush()
    try:
        _apply_units(db, p.id, req.unit_id, [u.model_dump() for u in req.units])
        _apply_suppliers(db, p.id, req.supplier_ids or [])
        db.commit()
    except IntegrityError:
        db.rollback()
        raise BizError(E_PARAM, "商品编码已存在")
    cache_delete_pattern("product:*")  # 商品详情/条码缓存失效
    return ok(_product_out(db, db.get(BaseProduct, p.id)))


@router.get("/products/{product_id}")
def get_product(product_id: int, db: Session = Depends(get_db)) -> dict:
    def _load() -> dict:
        p = db.get(BaseProduct, product_id)
        if p is None:
            raise BizError(E_NOT_FOUND, "商品不存在")
        return _product_out(db, p)

    return ok(cache_aside(f"product:{product_id}", _PRODUCT_TTL, _load))


@router.put("/products/{product_id}", dependencies=[Depends(require_permission("base:product"))])
def update_product(product_id: int, req: ProductReq, db: Session = Depends(get_db)) -> dict:
    p = db.get(BaseProduct, product_id)
    if p is None:
        raise BizError(E_NOT_FOUND, "商品不存在")
    if req.code and db.scalar(select(BaseProduct.id).where(BaseProduct.code == req.code, BaseProduct.id != product_id)):
        raise BizError(E_PARAM, "商品编码已存在")
    if req.barcode.strip() and db.scalar(
        select(BaseProduct.id).where(BaseProduct.barcode == req.barcode.strip(), BaseProduct.id != product_id)
    ):
        raise BizError(E_PARAM, "条码已存在，请勿重复录入")
    # 编码留空 = 保持原编码（编辑时可不填）；supplier_ids 走关联表写入
    data = req.model_dump(exclude={"units", "code", "supplier_ids"}) if not req.code else req.model_dump(exclude={"units", "supplier_ids"})
    data["barcode"] = data["barcode"].strip()
    # 材料分类规则：仅二级分类可挂材料；历史数据保持原值（分类变更时才校验）
    if data.get("category_id") and data["category_id"] != p.category_id:
        _require_leaf_category(db, data["category_id"])
    for k, v in data.items():
        setattr(p, k, _parse_dec(v, k) if k in ("purchase_price", "min_stock", "max_stock") else v)
    try:
        _apply_units(db, p.id, req.unit_id, [u.model_dump() for u in req.units])
        if req.supplier_ids is not None:  # 缺省（None）保持原关联，避免状态切换等局部更新误清空
            _apply_suppliers(db, p.id, req.supplier_ids)
        db.commit()
    except IntegrityError:
        db.rollback()
        raise BizError(E_PARAM, "商品编码已存在")
    cache_delete_pattern("product:*")  # 商品详情/条码缓存失效
    return ok()


@router.put("/products/{product_id}/category", dependencies=[Depends(require_any_permission("base:category", "base:product"))])
def update_product_category(product_id: int, req: ProductCategoryReq, db: Session = Depends(get_db)) -> dict:
    """单独更新材料分类（分类管理页「取消挂载」）：category_id=0 取消挂载，>0 改挂二级/三级分类。"""
    p = db.get(BaseProduct, product_id)
    if p is None:
        raise BizError(E_NOT_FOUND, "商品不存在", http_status=404)
    if req.category_id:
        _require_leaf_category(db, req.category_id)
    p.category_id = req.category_id
    db.commit()
    cache_delete_pattern("product:*")  # 商品详情/条码缓存失效
    return ok()


@router.delete("/products/{product_id}", dependencies=[Depends(require_permission("base:product"))])
def delete_product(product_id: int, db: Session = Depends(get_db)) -> dict:
    p = db.get(BaseProduct, product_id)
    if p is None:
        raise BizError(E_NOT_FOUND, "商品不存在")
    p.status = 0  # 软删除：停用
    db.commit()
    cache_delete_pattern("product:*")  # 商品详情/条码缓存失效
    return ok()


def _product_export_rows(db: Session) -> list[list]:
    rows = db.scalars(select(BaseProduct).order_by(BaseProduct.id)).all()
    out = []
    for p in rows:
        cat = db.get(BaseCategory, p.category_id)
        unit = db.get(BaseUnit, p.unit_id)
        out.append([
            p.code, p.barcode, p.sku, p.name, cat.name if cat else "",
            p.spec, unit.name if unit else "", format(p.purchase_price, "f"),
            format(p.min_stock, "f"), format(p.max_stock, "f"),
            "启用" if p.status == 1 else "停用",
        ])
    return out


@static_router.get("/products/import-template", dependencies=[Depends(require_permission("base:product"))])
def product_import_template() -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "商品导入"
    ws.append(PRODUCT_IMPORT_COLUMNS)
    ws.append(["P001", "6901234567890", "SKU001", "示例商品", "五金件", "M8x30", "件", "1.50", "10", "1000"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=product_import_template.xlsx"})


@router.post("/products/import", dependencies=[Depends(require_permission("base:product"))])
async def import_products(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    data = await file.read()
    if len(data) > _MAX_IMPORT_BYTES:
        raise BizError(E_PARAM, "导入文件不能超过 10MB")
    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise BizError(E_PARAM, "文件为空")
    if len(rows) > _MAX_IMPORT_ROWS:
        raise BizError(E_PARAM, f"单次导入不能超过 {_MAX_IMPORT_ROWS} 行")
    headers = [str(h).strip() if h else "" for h in rows[0]]
    # 公司系统模板：按列名定位（序号|材料用途|材料大类|材料分类|物料编码|材料名称|型号规格|单位|数量|用途|用量(仅导入使用)|使用单位|备注）
    is_company = any("物料编码" in h for h in headers) and any("材料名称" in h for h in headers)
    if is_company:
        col = {h: i for i, h in enumerate(headers)}

        def v(row, key, default=""):
            i = col.get(key)
            return str(row[i]).strip() if i is not None and i < len(row) and row[i] is not None else default
        success, fail_rows = 0, []
        for idx, row in enumerate(rows[1:], start=2):
            material_code, name, spec, unit_name = v(row, "物料编码"), v(row, "材料名称"), v(row, "型号规格"), v(row, "单位")
            big_cat, sub_cat, remark = v(row, "材料大类"), v(row, "材料分类"), v(row, "备注")
            if not name:
                fail_rows.append({"row": idx, "reason": "材料名称为空"})
                continue
            # 物料编码=公司系统唯一编码（存 material_code），可为空（空则提示管理员补录）；商品编码=系统内部纯数字自动生成
            if material_code and db.scalar(select(BaseProduct.id).where(BaseProduct.material_code == material_code)):
                fail_rows.append({"row": idx, "reason": f"物料编码 {material_code} 已存在"})
                continue
            # 单位：不存在自动创建
            unit = db.scalar(select(BaseUnit).where(BaseUnit.name == unit_name)) if unit_name else None
            if unit is None:
                unit = BaseUnit(name=(unit_name if unit_name and not _UNIT_NAME_GARBAGE.match(unit_name) else "件"), remark="导入自动创建")
                db.add(unit)
                db.flush()
            # 分类：材料大类（一级）+ 材料分类（二级）自动创建
            cat_id = 0
            if big_cat:
                big = db.scalar(select(BaseCategory).where(BaseCategory.name == big_cat, BaseCategory.parent_id == 0))
                if big is None:
                    big = BaseCategory(parent_id=0, name=big_cat, path="/")
                    db.add(big)
                    db.flush()
                cat_id = big.id
                if sub_cat:
                    sub = db.scalar(select(BaseCategory).where(BaseCategory.name == sub_cat, BaseCategory.parent_id == big.id))
                    if sub is None:
                        sub = BaseCategory(parent_id=big.id, name=sub_cat, path=f"{big.path}{big.id}/")
                        db.add(sub)
                        db.flush()
                    cat_id = sub.id
            try:
                # 商品编码=系统内部纯数字（自动生成）；物料编码=公司系统编码（可能为空，空则提示管理员补录）
                code = _next_numeric_code(db)
                p = BaseProduct(
                    code=code, material_code=material_code, name=name, spec=spec,
                    category_id=cat_id, unit_id=unit.id,
                    purchase_price=Decimal(0), remark=remark,
                )
                db.add(p)
                db.flush()
                # 公司模板无条码列：条码留空，由管理员从公司系统抄写后补录（响应 notice 提示）
                db.add(BaseProductUnit(product_id=p.id, unit_id=unit.id, rate=Decimal(1), is_default=1))
                success += 1
            except IntegrityError:
                db.rollback()
                fail_rows.append({"row": idx, "reason": f"编码 {code} 已存在"})
        db.commit()
        notice = "公司模板不含条码列：已导入商品的条码留空，请管理员从公司系统抄写条码后补录（商品管理中编辑条码）"
        if any((v(row, "物料编码") or "").strip() == "" for row in rows[1:]):
            notice += "；部分商品未填物料编码（公司系统编码），请管理员补充"
        return ok({
            "success_count": success,
            "fail_rows": fail_rows,
            "notice": notice,
        })
    if headers[:10] != PRODUCT_IMPORT_COLUMNS:
        raise BizError(E_PARAM, f"表头必须为：{'/'.join(PRODUCT_IMPORT_COLUMNS)}（或公司模板：物料编码/材料名称/型号规格/单位/材料大类/材料分类）")
    success, fail_rows = 0, []
    for idx, row in enumerate(rows[1:], start=2):
        vals = [str(v).strip() if v is not None else "" for v in row] + [""] * 10
        code, name, unit_name = vals[0], vals[3], vals[6]
        if code and not code.isdigit():
            fail_rows.append({"row": idx, "reason": f"编码 {code} 不是纯数字"})
            continue
        if not name:
            fail_rows.append({"row": idx, "reason": "名称为空"})
            continue
        code = code or _next_numeric_code(db)  # 编码留空自动生成纯数字
        if db.scalar(select(BaseProduct.id).where(BaseProduct.code == code)):
            fail_rows.append({"row": idx, "reason": f"编码 {code} 已存在"})
            continue
        # 单位：存在则复用，不存在自动创建（默认单位）
        unit = db.scalar(select(BaseUnit).where(BaseUnit.name == unit_name)) if unit_name else None
        if unit is None:
            unit = BaseUnit(name=(unit_name if unit_name and not _UNIT_NAME_GARBAGE.match(unit_name) else "件"), remark="导入自动创建")
            db.add(unit)
            db.flush()
        # 分类：存在则复用，不存在自动创建（顶级）
        cat_name = vals[4]
        cat = db.scalar(select(BaseCategory).where(BaseCategory.name == cat_name, BaseCategory.parent_id == 0)) if cat_name else None
        if cat is None and cat_name:
            cat = BaseCategory(parent_id=0, name=cat_name, path="/")
            db.add(cat)
            db.flush()
        try:
            p = BaseProduct(
                code=code, barcode=vals[1], sku=vals[2], name=name,
                category_id=cat.id if cat else 0, spec=vals[5], unit_id=unit.id,
                purchase_price=Decimal(vals[7]) if _DECIMAL_RE.match(vals[7]) else Decimal(0),
                min_stock=Decimal(vals[8]) if _DECIMAL_RE.match(vals[8]) else Decimal(0),
                max_stock=Decimal(vals[9]) if _DECIMAL_RE.match(vals[9]) else Decimal(0),
            )
            db.add(p)
            db.flush()
            db.add(BaseProductUnit(product_id=p.id, unit_id=unit.id, rate=Decimal(1), is_default=1))
            success += 1
        except IntegrityError:
            db.rollback()
            fail_rows.append({"row": idx, "reason": f"编码 {code} 已存在"})
    db.commit()
    if success:
        cache_delete_pattern("product:*")  # 导入可能新增商品/单位/分类 → 全部失效
        cache_delete("dict:categories", "dict:units")
    return ok({"success_count": success, "fail_rows": fail_rows})


@static_router.get("/products/export", dependencies=[Depends(require_permission("base:product"))])
def export_products(db: Session = Depends(get_db)) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "商品"
    ws.append(PRODUCT_IMPORT_COLUMNS + ["状态"])
    for row in _product_export_rows(db):
        ws.append([safe_excel_value(x) for x in row])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=products.xlsx"})


# ============================ 仓库 / 货架 / 库位 ============================


@router.get("/warehouses")
def list_warehouses(db: Session = Depends(get_db)) -> dict:
    def _load() -> list[dict]:
        rows = db.scalars(select(BaseWarehouse).order_by(BaseWarehouse.id)).all()
        if not rows:
            return []
        wh_ids = [w.id for w in rows]
        shelf_cnt = dict(db.execute(
            select(BaseShelf.warehouse_id, func.count()).where(BaseShelf.warehouse_id.in_(wh_ids)).group_by(BaseShelf.warehouse_id)
        ).all())
        loc_cnt = dict(db.execute(
            select(BaseLocation.warehouse_id, func.count()).where(BaseLocation.warehouse_id.in_(wh_ids)).group_by(BaseLocation.warehouse_id)
        ).all())
        # 库内有货材料种数：该仓库有库存的不同材料数
        kind_cnt = dict(db.execute(
            select(StkStock.warehouse_id, func.count(func.distinct(StkStock.product_id)))
            .where(StkStock.warehouse_id.in_(wh_ids), StkStock.qty != 0)
            .group_by(StkStock.warehouse_id)
        ).all())
        out = []
        for w in rows:
            d = WarehouseOut.model_validate(w, from_attributes=True).model_dump()
            d["shelf_count"] = shelf_cnt.get(w.id, 0)
            d["location_count"] = loc_cnt.get(w.id, 0)
            d["product_kind_count"] = kind_cnt.get(w.id, 0)
            out.append(d)
        return out

    return ok(cache_aside("dict:warehouses", _DICT_TTL, _load))


@router.post("/warehouses", dependencies=[Depends(require_permission("base:warehouse"))])
def create_warehouse(req: WarehouseReq, db: Session = Depends(get_db)) -> dict:
    # 新建仓库编码=仓库名称（界面不展示编码，避免 WH 编码混淆）；重名即编码冲突
    if db.scalar(select(BaseWarehouse.id).where(BaseWarehouse.code == req.code)):
        raise BizError(E_PARAM, "仓库已存在（名称重复）")
    w = BaseWarehouse(**req.model_dump())
    db.add(w)
    db.commit()
    db.refresh(w)
    cache_delete("dict:warehouses")
    return ok(WarehouseOut.model_validate(w, from_attributes=True).model_dump())


@router.put("/warehouses/{wh_id}", dependencies=[Depends(require_permission("base:warehouse"))])
def update_warehouse(wh_id: int, req: WarehouseUpdateReq, db: Session = Depends(get_db)) -> dict:
    w = db.get(BaseWarehouse, wh_id)
    if w is None:
        raise BizError(E_NOT_FOUND, "仓库不存在")
    # 名称改动后同步编码（编码=名称，保持唯一与一致）
    if db.scalar(select(BaseWarehouse.id).where(BaseWarehouse.code == req.name.strip(), BaseWarehouse.id != wh_id)):
        raise BizError(E_PARAM, "仓库已存在（名称重复）")
    w.code = req.name.strip()
    for k, v in req.model_dump().items():
        setattr(w, k, v)
    db.commit()
    cache_delete("dict:warehouses")
    return ok()


@router.delete("/warehouses/{wh_id}", dependencies=[Depends(require_permission("base:warehouse"))])
def delete_warehouse(wh_id: int, db: Session = Depends(get_db)) -> dict:
    """停用仓库：有启用货架或库存时禁止（确认规则：仓库有启用货架/库存禁停用）。"""
    w = db.get(BaseWarehouse, wh_id)
    if w is None:
        raise BizError(E_NOT_FOUND, "仓库不存在")
    enabled_shelf = db.scalar(select(BaseShelf.id).where(BaseShelf.warehouse_id == wh_id).limit(1))
    if enabled_shelf:
        raise BizError(E_PARAM, "仓库下存在货架，禁止停用（请先删除货架）")
    stock_cnt = db.scalar(select(func.count()).select_from(StkStock).where(StkStock.warehouse_id == wh_id, StkStock.qty != 0)) or 0
    if stock_cnt:
        raise BizError(E_PARAM, "仓库存在库存，禁止停用（请先移走库存）")
    w.status = 0  # 软删除
    db.commit()
    cache_delete("dict:warehouses")
    return ok()


@router.get("/warehouses/{wh_id}/shelves")
def list_shelves(
    wh_id: int,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """仓库货架列表；wh_id=0 返回全部货架（单位管理页一次拉取，避免逐仓库请求）；非超管按所属单位过滤。
    每个货架附带实际维度（layers/rows/cols，由库位推导），供 2.5D 货架视图布局。"""
    stmt = select(BaseShelf)
    if wh_id:
        stmt = stmt.where(BaseShelf.warehouse_id == wh_id)
    visible = _visible_shelf_ids(db, user)
    if visible is not None:
        stmt = stmt.where(BaseShelf.id.in_(visible) if visible else False)
    # 无单位过滤（超管/管理者）时全量货架可缓存；受可见性限制的用户不缓存（正确性优先）
    if visible is None:
        def _load() -> list[dict]:
            rows = db.scalars(stmt.order_by(BaseShelf.code)).all()
            return _shelves_out(db, rows)

        return ok(cache_aside(f"dict:shelves:{wh_id}", _DICT_TTL, _load))
    rows = db.scalars(stmt.order_by(BaseShelf.code)).all()
    return ok(_shelves_out(db, rows))


def _shelves_out(db: Session, rows: list[BaseShelf]) -> list[dict]:
    """货架输出：附带实际维度（由库位最大层/行/列推导）。"""
    if not rows:
        return []
    shelf_ids = [s.id for s in rows]
    loc_rows = db.execute(
        select(BaseLocation.shelf_id, func.max(BaseLocation.layer_no), func.max(BaseLocation.row_no), func.max(BaseLocation.col_no))
        .where(BaseLocation.shelf_id.in_(shelf_ids))
        .group_by(BaseLocation.shelf_id)
    ).all()
    dims = {sid: (layers, rows_, cols) for sid, layers, rows_, cols in loc_rows}
    out = []
    for s in rows:
        layers, rows_, cols = dims.get(s.id, (1, 1, 1))
        d = ShelfOut.model_validate(s, from_attributes=True).model_dump()
        d["layers"] = layers
        d["rows"] = rows_
        d["cols"] = cols
        out.append(d)
    return out


@router.post("/shelves", dependencies=[Depends(require_permission("base:warehouse"))])
def create_shelf(req: ShelfReq, db: Session = Depends(get_db)) -> dict:
    """新建货架；提供 layers/rows/cols 时按 层×行×列 批量生成库位（隔，2.5D 视图用）。"""
    if db.get(BaseWarehouse, req.warehouse_id) is None:
        raise BizError(E_PARAM, "仓库不存在")
    dup = db.scalar(select(BaseShelf.id).where(BaseShelf.warehouse_id == req.warehouse_id, BaseShelf.code == req.code))
    if dup:
        raise BizError(E_PARAM, "同仓库下货架编码已存在")
    s = BaseShelf(warehouse_id=req.warehouse_id, code=req.code, name=req.name, remark=req.remark)
    db.add(s)
    db.flush()
    if req.layers is not None and req.rows is not None and req.cols is not None:
        _generate_shelf_cells(db, s, req.layers, req.rows, req.cols)
    db.commit()
    db.refresh(s)
    cache_delete_pattern("dict:shelves*", "dict:locations*", "stock:locsum:*")  # 货架图缓存同失效
    dims = _shelves_out(db, [s])[0]
    return ok(dims)


def _generate_shelf_cells(db: Session, shelf: BaseShelf, layers: int, rows: int, cols: int) -> None:
    """按 层×行×列 批量生成货架库位（隔），编码：仓库编码-货架编码-L{层}R{行}C{列}。"""
    wh = db.get(BaseWarehouse, shelf.warehouse_id)
    wh_code = wh.code if wh else "WH"
    for layer in range(1, layers + 1):
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                code = f"{wh_code}-{shelf.code}-L{layer}R{row}C{col}"
                if not db.scalar(select(BaseLocation.id).where(BaseLocation.code == code)):
                    db.add(BaseLocation(
                        warehouse_id=shelf.warehouse_id, shelf_id=shelf.id,
                        layer_no=layer, row_no=row, col_no=col, code=code,
                    ))


@router.put("/shelves/{shelf_id}", dependencies=[Depends(require_permission("base:warehouse"))])
def update_shelf(shelf_id: int, req: ShelfReq, db: Session = Depends(get_db)) -> dict:
    s = db.get(BaseShelf, shelf_id)
    if s is None:
        raise BizError(E_NOT_FOUND, "货架不存在")
    # 只更新可编辑字段（名称/备注/编码）；层行列维度由库位推导，不在编辑范围
    for k in ("code", "name", "remark"):
        v = getattr(req, k)
        if v is not None:
            setattr(s, k, v)
    db.commit()
    cache_delete_pattern("dict:shelves*", "stock:locsum:*")  # 货架图缓存同失效
    return ok()


@router.delete("/shelves/{shelf_id}", dependencies=[Depends(require_permission("base:warehouse"))])
def delete_shelf(shelf_id: int, db: Session = Depends(get_db)) -> dict:
    s = db.get(BaseShelf, shelf_id)
    if s is None:
        raise BizError(E_NOT_FOUND, "货架不存在")
    loc_cnt = db.scalar(select(func.count()).select_from(BaseLocation).where(BaseLocation.shelf_id == shelf_id)) or 0
    if loc_cnt:
        raise BizError(E_PARAM, "货架下存在库位，禁止删除")
    db.delete(s)
    db.commit()
    cache_delete_pattern("dict:shelves*", "stock:locsum:*")  # 货架图缓存同失效
    return ok()


def _loc_out(db: Session, rows: list[BaseLocation]) -> list[dict]:
    """库位输出（含 display 友好名：仓库名-货架编码-L{层}R{行}C{列}，界面不显示 WH 仓库编码）。"""
    if not rows:
        return []
    whs = {w.id: w for w in db.scalars(select(BaseWarehouse).where(BaseWarehouse.id.in_({r.warehouse_id for r in rows}))).all()}
    shelves = {s.id: s for s in db.scalars(select(BaseShelf).where(BaseShelf.id.in_({r.shelf_id for r in rows}))).all()}
    out = []
    for l in rows:
        d = LocationOut.model_validate(l, from_attributes=True).model_dump()
        wh = whs.get(l.warehouse_id)
        shelf = shelves.get(l.shelf_id)
        d["display"] = f"{wh.name if wh else ''}-{shelf.code if shelf else ''}-L{l.layer_no}R{l.row_no}C{l.col_no}"
        out.append(d)
    return out


@router.get("/locations")
def list_locations(
    warehouse_id: int = Query(0),
    shelf_id: int = Query(0),
    db: Session = Depends(get_db),
) -> dict:
    def _load() -> list[dict]:
        stmt = select(BaseLocation)
        if warehouse_id:
            stmt = stmt.where(BaseLocation.warehouse_id == warehouse_id)
        if shelf_id:
            stmt = stmt.where(BaseLocation.shelf_id == shelf_id)
        rows = db.scalars(stmt.order_by(BaseLocation.code)).all()
        return _loc_out(db, rows)

    return ok(cache_aside(f"dict:locations:{warehouse_id}:{shelf_id}", _DICT_TTL, _load))


@router.post("/locations", dependencies=[Depends(require_permission("base:stock-location"))])
def create_location(req: LocationReq, db: Session = Depends(get_db)) -> dict:
    """新建库位（隔）：货架内 层×行×列 定位；同货架同坐标重复 → 报错。"""
    shelf = db.get(BaseShelf, req.shelf_id)
    if shelf is None:
        raise BizError(E_PARAM, "货架不存在")
    if shelf.warehouse_id != req.warehouse_id:
        raise BizError(E_PARAM, "货架与仓库不匹配")
    wh = db.get(BaseWarehouse, req.warehouse_id)
    dup_coord = db.scalar(select(BaseLocation.id).where(
        BaseLocation.shelf_id == req.shelf_id,
        BaseLocation.layer_no == req.layer_no,
        BaseLocation.row_no == req.row_no,
        BaseLocation.col_no == req.col_no,
    ))
    if dup_coord:
        raise BizError(E_PARAM, f"货架 {shelf.code} 已存在 {req.layer_no}层{req.row_no}行{req.col_no}列的库位")
    code = req.code or f"{wh.code if wh else 'WH'}-{shelf.code}-L{req.layer_no}R{req.row_no}C{req.col_no}"
    if db.scalar(select(BaseLocation.id).where(BaseLocation.code == code)):
        raise BizError(E_PARAM, f"库位编码 {code} 已存在")
    loc = BaseLocation(
        warehouse_id=req.warehouse_id, shelf_id=req.shelf_id,
        layer_no=req.layer_no, row_no=req.row_no, col_no=req.col_no, code=code, remark=req.remark,
    )
    db.add(loc)
    db.commit()
    db.refresh(loc)
    cache_delete_pattern("dict:locations*", "stock:locsum:*")  # 货架图缓存同失效
    return ok(_loc_out(db, [loc])[0])


@router.put("/locations/{loc_id}", dependencies=[Depends(require_permission("base:stock-location"))])
def update_location(loc_id: int, req: LocationReq, db: Session = Depends(get_db)) -> dict:
    loc = db.get(BaseLocation, loc_id)
    if loc is None:
        raise BizError(E_NOT_FOUND, "库位不存在")
    for k, v in req.model_dump().items():
        setattr(loc, k, v)
    db.commit()
    cache_delete_pattern("dict:locations*", "stock:locsum:*")  # 货架图缓存同失效
    return ok()


@router.delete("/locations/{loc_id}", dependencies=[Depends(require_permission("base:stock-location"))])
def delete_location(loc_id: int, db: Session = Depends(get_db)) -> dict:
    loc = db.get(BaseLocation, loc_id)
    if loc is None:
        raise BizError(E_NOT_FOUND, "库位不存在")
    stock_cnt = db.execute(text("SELECT COUNT(*) FROM stk_stock WHERE location_id = :id"), {"id": loc_id}).scalar() or 0
    if stock_cnt:
        raise BizError(E_PARAM, "库位存在库存，禁止删除")
    db.delete(loc)
    db.commit()
    cache_delete_pattern("dict:locations*", "stock:locsum:*")  # 货架图缓存同失效
    return ok()


# ============================ 组织单位（部门） ============================


def _visible_shelf_ids(db: Session, user: SysUser) -> set[int] | None:
    """当前用户可见货架 id 集合；None 表示全部（超管/管理者/无单位角色不限）。"""
    role = db.get(SysRole, user.role_id)
    if role is None or not role.department_id:
        return None
    if role.code in (SUPER_ADMIN_ROLE_CODE, "manager"):
        return None
    rows = db.scalars(
        select(BaseDepartmentShelf.shelf_id).where(BaseDepartmentShelf.department_id == role.department_id)
    ).all()
    return set(rows)


def _dept_out(db: Session, d: BaseDepartment) -> dict:
    shelf_ids = db.scalars(
        select(BaseDepartmentShelf.shelf_id).where(BaseDepartmentShelf.department_id == d.id)
    ).all()
    return DeptOut(
        id=d.id, code=d.code, name=d.name, remark=d.remark, status=d.status, shelf_ids=list(shelf_ids),
    ).model_dump()


@router.get("/departments")
def list_departments(db: Session = Depends(get_db)) -> dict:
    def _load() -> list[dict]:
        rows = db.scalars(select(BaseDepartment).order_by(BaseDepartment.id)).all()
        return [_dept_out(db, d) for d in rows]

    return ok(cache_aside("dict:departments", _DICT_TTL, _load))


@router.post("/departments", dependencies=[Depends(require_permission("dept:manage"))])
def create_department(req: DeptReq, db: Session = Depends(get_db)) -> dict:
    # 编码由系统自动生成（数字编码，对用户隐藏），不接受前端传入
    d = BaseDepartment(code=_gen_dept_code(db), name=req.name, remark=req.remark, status=req.status)
    db.add(d)
    db.commit()
    db.refresh(d)
    cache_delete("dict:departments")
    return ok({"id": d.id, "code": d.code})


def _gen_dept_code(db: Session) -> str:
    """自动生成单位数字编码：{yyyyMMdd}{4位当日序号}（纯数字序列；删除后序号复用冲突时递增重试）。"""
    today = datetime.now().strftime("%Y%m%d")
    like = f"{today}%"
    cnt = db.scalar(select(func.count()).select_from(BaseDepartment).where(BaseDepartment.code.like(like))) or 0
    code = f"{today}{cnt + 1:04d}"
    while db.scalar(select(BaseDepartment.id).where(BaseDepartment.code == code)):
        cnt += 1
        code = f"{today}{cnt + 1:04d}"
    return code


@router.put("/departments/{dept_id}", dependencies=[Depends(require_permission("dept:manage"))])
def update_department(dept_id: int, req: DeptUpdateReq, db: Session = Depends(get_db)) -> dict:
    d = db.get(BaseDepartment, dept_id)
    if d is None:
        raise BizError(E_NOT_FOUND, "单位不存在")
    if req.name is not None:
        d.name = req.name
    if req.remark is not None:
        d.remark = req.remark
    if req.status is not None:
        d.status = req.status
    db.commit()
    cache_delete("dict:departments")
    return ok()


@router.delete("/departments/{dept_id}", dependencies=[Depends(require_permission("dept:manage"))])
def delete_department(dept_id: int, db: Session = Depends(get_db)) -> dict:
    d = db.get(BaseDepartment, dept_id)
    if d is None:
        raise BizError(E_NOT_FOUND, "单位不存在")
    if db.scalar(select(func.count()).select_from(SysRole).where(SysRole.department_id == dept_id)):
        raise BizError(E_PARAM, "该单位下还有角色，请先调整角色所属单位")
    db.execute(BaseDepartmentShelf.__table__.delete().where(BaseDepartmentShelf.department_id == dept_id))
    db.delete(d)
    db.commit()
    cache_delete("dict:departments")
    return ok()


@router.put("/departments/{dept_id}/shelves", dependencies=[Depends(require_permission("dept:manage"))])
def update_department_shelves(dept_id: int, req: DeptShelvesReq, db: Session = Depends(get_db)) -> dict:
    """设置单位可用显示的货架（超管/管理者不受限，其余角色仅见本单货架）。"""
    if db.get(BaseDepartment, dept_id) is None:
        raise BizError(E_NOT_FOUND, "单位不存在")
    valid = set(db.scalars(select(BaseShelf.id)).all())
    bad = [sid for sid in req.shelf_ids if sid not in valid]
    if bad:
        raise BizError(E_PARAM, f"货架不存在：{bad}")
    db.execute(BaseDepartmentShelf.__table__.delete().where(BaseDepartmentShelf.department_id == dept_id))
    for sid in req.shelf_ids:
        db.add(BaseDepartmentShelf(department_id=dept_id, shelf_id=sid))
    db.commit()
    cache_delete("dict:departments")  # 部门含 shelf_ids，变更需失效
    return ok()


@router.post("/products/dedupe-scan", dependencies=[Depends(require_permission("base:product"))])
def products_dedupe_scan(db: Session = Depends(get_db)) -> dict:
    """材料查重扫描：名称精确重复 + 本地相似规则分组 → 疑似重复分组（仅建议，不落库）。"""
    from app.services.ai.dedupe import dedupe_scan as _scan

    return ok({"groups": _scan(db)})


@router.post("/products/{product_id}/mark-duplicate", dependencies=[Depends(require_permission("base:product"))])
def mark_product_duplicate(product_id: int, db: Session = Depends(get_db)) -> dict:
    """人工标记材料为重复（写 remark 前缀，不物理删除——红线禁删）。"""
    p = db.get(BaseProduct, product_id)
    if p is None:
        raise BizError(E_NOT_FOUND, "材料不存在")
    mark = "【疑似重复】" if not (p.remark or "").startswith("【疑似重复】") else ""
    p.remark = f"{mark}{p.remark or ''}"
    db.commit()
    return ok()


# ============================ 删除审核（物料/分类删除审批流） ============================
# 物料数据管理 → 删除不直接执行：先提交删除申请，管理者及以上角色审核通过后才真正执行删除
# （材料删除 = 停用 status=0；分类删除 = 物理删除，仍受「有子分类/有材料禁删」保护）。

_MANAGER_ROLES = ("super_admin", "manager")


def _notify_managers(db: Session, title: str, content: str, exclude_user_id: int = 0, link: str = "") -> None:
    """站内通知管理者及以上角色（删除申请待审核）。link=业务联动目标，兼作自动已读唯一键。"""
    role_ids = db.scalars(select(SysRole.id).where(SysRole.code.in_(_MANAGER_ROLES))).all()
    uids = db.scalars(
        select(SysUser.id).where(SysUser.role_id.in_(role_ids), SysUser.status == 1)
    ).all()
    for uid in uids:
        if uid != exclude_user_id:
            db.add(SysNotification(user_id=uid, title=title, content=content, biz_type="待办", link=link))


def _clear_delete_review_todo(db: Session, review_id: int) -> None:
    """删除申请已处理（通过/驳回）后，自动把管理者们的「删除申请待审核」待办标记已读，
    避免已处理完的待办长期残留（业务联动的一部分）。"""
    db.execute(
        SysNotification.__table__.update()
        .where(
            SysNotification.link == f"/delete-reviews?review={review_id}",
            SysNotification.title == "删除申请待审核",
            SysNotification.is_read == 0,
        )
        .values(is_read=1)
    )


def _review_out(r: SysDeleteReview) -> dict:
    return DeleteReviewOut(
        id=r.id, biz_type=r.biz_type, target_id=r.target_id, target_name=r.target_name,
        target_desc=r.target_desc, reason=r.reason, status=r.status,
        applicant_id=r.applicant_id, applicant_name=r.applicant_name,
        handled_by=r.handled_by, handled_at=r.handled_at, review_remark=r.review_remark,
        created_at=r.created_at,
    ).model_dump()


@router.post("/delete-reviews", dependencies=[Depends(require_any_permission("base:product", "base:category", "fault:manage", "fault:report"))])
def create_delete_review(req: DeleteReviewReq, user: SysUser = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
    """提交删除申请（材料停用 / 分类删除 / 已关闭故障审核删除）：生成待审核记录并通知管理者；不直接执行删除。"""
    now = datetime.now()
    if req.biz_type == "product":
        p = db.get(BaseProduct, req.target_id)
        if p is None:
            raise BizError(E_NOT_FOUND, "材料不存在")
        if p.status == 0:
            raise BizError(E_PARAM, "该材料已停用，无需再申请删除")
        name, desc = p.name, f"编码 {p.code or '-'} · 规格 {p.spec or '-'}"
    elif req.biz_type == "category":
        c = db.get(BaseCategory, req.target_id)
        if c is None:
            raise BizError(E_NOT_FOUND, "分类不存在")
        name, desc = c.name, f"路径 {c.path or '/'}"
    elif req.biz_type == "fault":
        try:
            from app.modules.cable.models import CableFault  # 模块插件：已安装才可用
        except ImportError:
            raise BizError(E_PARAM, "cable 模块未启用") from None
        f = db.get(CableFault, req.target_id)
        if f is None or f.deleted:
            raise BizError(E_NOT_FOUND, "故障不存在")
        if f.status != 4:
            raise BizError(E_PARAM, "仅已关闭的故障需要审核删除")
        name = f"故障 #{f.id}"
        desc = f"{f.fault_type or '未分类'} · {float(f.lat):.6f},{float(f.lng):.6f}"
    else:
        raise BizError(E_PARAM, "未知删除类型")
    pending = db.scalar(
        select(SysDeleteReview.id).where(
            SysDeleteReview.biz_type == req.biz_type,
            SysDeleteReview.target_id == req.target_id,
            SysDeleteReview.status == 0,
        )
    )
    if pending:
        raise BizError(E_PARAM, "该对象已有待审核的删除申请，请等待审核结果")
    review = SysDeleteReview(
        biz_type=req.biz_type, target_id=req.target_id, target_name=name[:200], target_desc=desc[:500],
        reason=req.reason.strip()[:500], status=0,
        applicant_id=user.id, applicant_name=user.real_name or user.username,
    )
    db.add(review)
    db.commit()
    db.refresh(review)
    type_label = {"product": f"材料：{name}", "category": f"分类：{name}", "fault": f"故障：{name}"}.get(req.biz_type, name)
    _notify_managers(db, "删除申请待审核", f"{user.real_name or user.username} 申请删除{type_label}：{req.reason}", exclude_user_id=user.id, link=f"/delete-reviews?review={review.id}")
    db.commit()
    return ok(_review_out(review))


@router.get("/delete-reviews", dependencies=[Depends(require_any_permission("base:product", "base:category", "fault:manage"))])
def list_delete_reviews(
    status: int = Query(0, ge=0, le=2, description="0 待审核 / 1 已通过 / 2 已驳回"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    """删除审核列表（按状态筛选；申请人可查看自己提交的申请进度）。"""
    stmt = select(SysDeleteReview).order_by(SysDeleteReview.id.desc())
    if status is not None:
        stmt = stmt.where(SysDeleteReview.status == status)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    return ok({
        "list": [_review_out(r) for r in rows],
        "total": total, "page": page, "page_size": page_size,
    })


@router.post("/delete-reviews/{review_id}/approve", dependencies=[Depends(require_manager_role())])
def approve_delete_review(review_id: int, user: SysUser = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
    """审核通过：执行删除（材料 → 停用；分类 → 删除，仍受子分类/材料保护，不满足则自动驳回）。"""
    review = db.get(SysDeleteReview, review_id)
    if review is None:
        raise BizError(E_NOT_FOUND, "删除申请不存在")
    if review.status != 0:
        raise BizError(E_PARAM, "该申请已处理，请勿重复操作")
    now = datetime.now()
    reviewer = user.real_name or user.username
    if review.biz_type == "product":
        p = db.get(BaseProduct, review.target_id)
        if p is not None and p.status == 1:
            p.status = 0  # 软删除：停用
            cache_delete_pattern("product:*")
    elif review.biz_type == "category":
        c = db.get(BaseCategory, review.target_id)
        if c is not None:
            child_cnt = db.scalar(select(func.count()).select_from(BaseCategory).where(BaseCategory.parent_id == c.id)) or 0
            product_cnt = db.scalar(select(func.count()).select_from(BaseProduct).where(BaseProduct.category_id == c.id)) or 0
            if child_cnt or product_cnt:
                # 审核时已不满足删除条件 → 自动驳回（有子分类或材料）
                review.status = 2
                review.handled_by = user.id
                review.handled_at = now
                review.review_remark = "审核未通过：分类下仍有子分类或材料，无法删除"
                db.commit()
                _notify(db, review.applicant_id, "删除申请未通过", f"分类「{review.target_name}」删除申请被驳回：{review.review_remark}", "审批", link=f"/delete-reviews?review={review.id}")
                _clear_delete_review_todo(db, review.id)
                return ok(_review_out(review))
            db.delete(c)
            cache_delete("dict:categories")
    elif review.biz_type == "fault":
        try:
            from app.modules.cable.models import CableFault  # 模块插件：已安装才可用
        except ImportError:
            raise BizError(E_PARAM, "cable 模块未启用") from None
        f = db.get(CableFault, review.target_id)
        if f is not None and not f.deleted:
            f.deleted = 1  # 软删除：地图/列表不再显示，数据保留可追溯
    review.status = 1
    review.handled_by = user.id
    review.handled_at = now
    review.review_remark = review.review_remark or "审核通过"
    db.commit()
    _notify(db, review.applicant_id, "删除申请已通过", f"「{review.target_name}」已由 {reviewer} 审核通过并删除", "审批", link=f"/delete-reviews?review={review.id}")
    _clear_delete_review_todo(db, review.id)
    return ok(_review_out(review))


@router.post("/delete-reviews/{review_id}/reject", dependencies=[Depends(require_manager_role())])
def reject_delete_review(review_id: int, req: DeleteReviewRejectReq, user: SysUser = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
    """审核驳回：不执行删除。"""
    review = db.get(SysDeleteReview, review_id)
    if review is None:
        raise BizError(E_NOT_FOUND, "删除申请不存在")
    if review.status != 0:
        raise BizError(E_PARAM, "该申请已处理，请勿重复操作")
    review.status = 2
    review.handled_by = user.id
    review.handled_at = datetime.now()
    review.review_remark = req.remark.strip()[:500]
    db.commit()
    _notify(db, review.applicant_id, "删除申请被驳回", f"「{review.target_name}」删除申请被驳回：{review.review_remark}", "审批", link=f"/delete-reviews?review={review.id}")
    _clear_delete_review_todo(db, review.id)
    return ok(_review_out(review))


def _notify(db: Session, user_id: int, title: str, content: str, biz_type: str, link: str = "") -> None:
    """站内通知（调用方事务内）。"""
    db.add(SysNotification(user_id=user_id, title=title, content=content, biz_type=biz_type, link=link))
