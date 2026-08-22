"""库存接口：采购入库/期初库存/库存查询/流水（《后端API设计.md》§3、§6）。

库存一致性：所有入库/作废/期初过账必须走 services/stock.py 的 post_stock_change()，
与单据状态更新在同一事务提交（《开发规范.md》§4.5）。
"""
from __future__ import annotations

import io
import json
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.cache import cache_aside_json
from app.core.deps import get_current_user, require_any_permission, require_permission
from app.core.response import BizError, E_BILL_STATUS, E_NOT_FOUND, E_PARAM, ok
from app.db import get_db
from app.models.base import BaseCategory, BaseLocation, BaseProduct, BaseShelf, BaseSupplier, BaseUnit, BaseWarehouse
from app.models.stock import (
    PchPurchaseIn,
    PchPurchaseInItem,
    PchPurchasePlan,
    PchPurchasePlanItem,
    StkOpening,
    StkOpeningItem,
    StkStock,
    StkStockLog,
)
from app.models.ocr import OcrRecord
from app.models.sys import SysUser
from app.schemas.stock import (
    OpeningItemOut,
    OpeningOut,
    OpeningReq,
    PageData,
    PurchaseInItemOut,
    PurchaseInOut,
    PurchaseInReq,
    PurchasePlanItemOut,
    PurchasePlanItemReq,
    PurchasePlanOut,
    PurchasePlanReq,
    StockFlowRow,
    StockRow,
)
from app.services.stock import bill_no_conflict, generate_bill_no, post_stock_change

router = APIRouter(tags=["库存"], dependencies=[Depends(get_current_user)])

_DECIMAL_RE = re.compile(r"^\d+(\.\d+)?$")
_DEC2 = Decimal("0.01")
_DEC3 = Decimal("0.001")
_OPENING_IMPORT_COLUMNS = ["商品编码", "库位编码", "数量", "成本价"]
_MAX_IMPORT_BYTES = 10 * 1024 * 1024  # 导入文件大小上限 10MB（防内存耗尽）
_MAX_IMPORT_ROWS = 5000  # 单次导入行数上限


def _fmt_qty(v: Decimal | int | str | None) -> str:
    """数量统一格式化：保留至 3 位小数并去尾零（30.000 → 30，5.500 → 5.5）。"""
    d = Decimal(v or 0)
    return format(d.quantize(_DEC3), "f").rstrip("0").rstrip(".") or "0"


def _parse_dec(v: str, field: str) -> Decimal:
    if not _DECIMAL_RE.match(v):
        raise BizError(E_PARAM, f"{field} 必须是数字")
    return Decimal(v)


def _check_date_arg(v: str, field: str) -> None:
    """日期参数格式校验（YYYY-MM-DD）：非法值直接 4006，避免拼进 SQL 后 500。"""
    if v:
        try:
            datetime.fromisoformat(v)
        except ValueError:
            raise BizError(E_PARAM, f"{field} 格式错误，应为 YYYY-MM-DD")


def _user_name(db: Session, uid: int) -> str:
    u = db.get(SysUser, uid)
    return u.real_name if u else ""


def _loc_code(db: Session, loc_id: int) -> str:
    """库位显示名：仓库名-货架编码-L{层}R{行}C{列}（界面不显示 WH 仓库编码，避免混淆）。"""
    loc = db.get(BaseLocation, loc_id)
    if loc is None:
        return ""
    wh = db.get(BaseWarehouse, loc.warehouse_id)
    shelf = db.get(BaseShelf, loc.shelf_id)
    return f"{wh.name if wh else ''}-{shelf.code if shelf else ''}-L{loc.layer_no}R{loc.row_no}C{loc.col_no}"


def _purchase_out(db: Session, bill: PchPurchaseIn) -> dict:
    items = db.scalars(select(PchPurchaseInItem).where(PchPurchaseInItem.bill_id == bill.id).order_by(PchPurchaseInItem.sort)).all()
    wh = db.get(BaseWarehouse, bill.warehouse_id)
    sup = db.get(BaseSupplier, bill.supplier_id)
    plan = db.get(PchPurchasePlan, bill.plan_id) if bill.plan_id else None
    try:
        delivery_file_ids = json.loads(bill.delivery_file_ids) if bill.delivery_file_ids else []
    except (ValueError, TypeError):
        delivery_file_ids = []
    if not isinstance(delivery_file_ids, list):
        delivery_file_ids = []
    return PurchaseInOut(
        id=bill.id, bill_no=bill.bill_no, ocr_bill_no=bill.ocr_bill_no, supplier_id=bill.supplier_id,
        supplier_name=sup.name if sup else "",
        warehouse_id=bill.warehouse_id, warehouse_name=wh.name if wh else "",
        total_qty=bill.total_qty, total_amount=bill.total_amount, status=bill.status,
        bill_date=bill.bill_date, operator_name=_user_name(db, bill.operator_id), remark=bill.remark,
        ocr_record_id=bill.ocr_record_id,
        plan_id=bill.plan_id, plan_bill_no=plan.bill_no if plan else "",
        delivery_file_ids=delivery_file_ids,
        items=[
            PurchaseInItemOut(
                id=it.id, product_id=it.product_id,
                product_name=(p.name if (p := db.get(BaseProduct, it.product_id)) else ""),
                code=(p.code if (p := db.get(BaseProduct, it.product_id)) else ""),
                qty=it.qty, unit_name=it.unit_name, price=it.price, amount=it.amount,
                location_id=it.location_id, location_code=_loc_code(db, it.location_id),
            )
            for it in items
        ],
    ).model_dump()


# ============================ 采购入库 ============================


@router.post("/purchase-in", dependencies=[Depends(require_permission("pch:in"))])
def create_purchase_in(
    req: PurchaseInReq,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    if db.get(BaseWarehouse, req.warehouse_id) is None:
        raise BizError(E_PARAM, "仓库不存在")
    # 供应商可选，但传了就必须存在（避免悬空 supplier_id 入库单）
    if req.supplier_id and db.get(BaseSupplier, req.supplier_id) is None:
        raise BizError(E_PARAM, "供应商不存在")
    if req.ocr_record_id and db.get(OcrRecord, req.ocr_record_id) is None:
        raise BizError(E_PARAM, "送货单 OCR 记录不存在")
    # 采购计划单：存在且状态允许入库（已提交/部分入库）；已完成/作废不可再入库
    plan: PchPurchasePlan | None = None
    if req.plan_id:
        plan = db.get(PchPurchasePlan, req.plan_id)
        if plan is None:
            raise BizError(E_PARAM, "采购计划单不存在")
        if plan.status not in (1, 2):
            raise BizError(E_BILL_STATUS, "该采购计划单当前状态不可入库（仅已提交/部分入库可继续入库）")
        if plan.warehouse_id != req.warehouse_id:
            raise BizError(E_PARAM, "入库仓库与采购计划单仓库不一致")
    # 送货单图片存底：可选、最多 10 张、文件必须存在
    delivery_file_ids = list(dict.fromkeys(req.delivery_file_ids or []))  # 去重保序
    if len(delivery_file_ids) > 10:
        raise BizError(E_PARAM, "送货单图片最多 10 张")
    if delivery_file_ids:
        from app.models.sys import SysFile

        exist = set(db.scalars(select(SysFile.id).where(SysFile.id.in_(delivery_file_ids))).all())
        missing = [fid for fid in delivery_file_ids if fid not in exist]
        if missing:
            raise BizError(E_PARAM, f"送货单图片不存在：{missing}")
    for attempt in range(5):  # 单号并发冲突重试
        bill_no = generate_bill_no(db, "RK", PchPurchaseIn)
        bill = PchPurchaseIn(
            bill_no=bill_no, supplier_id=req.supplier_id, warehouse_id=req.warehouse_id,
            status=1, bill_date=req.bill_date or datetime.now(), operator_id=user.id,
            ocr_record_id=req.ocr_record_id, ocr_bill_no=req.ocr_bill_no.strip(), remark=req.remark,
            plan_id=req.plan_id, delivery_file_ids=json.dumps(delivery_file_ids),
        )
        db.add(bill)
        db.flush()
        total_qty, total_amount = Decimal(0), Decimal(0)
        try:
            for idx, item in enumerate(req.items):
                product = db.get(BaseProduct, item.product_id)
                if product is None:
                    raise BizError(E_NOT_FOUND, f"商品 id={item.product_id} 不存在")
                if db.get(BaseLocation, item.location_id) is None:
                    raise BizError(E_NOT_FOUND, f"库位 id={item.location_id} 不存在")
                qty = _parse_dec(item.qty, "数量")
                price = _parse_dec(item.price, "进价")
                if qty <= 0:
                    raise BizError(E_PARAM, "数量必须大于 0")
                if price < 0:
                    raise BizError(E_PARAM, "进价不能为负数")
                # 明细带分类（大模型识别/人工确认）：校验分类存在并更新材料分类；
                # 规则（三级体系）：顶级分类仅作分组，材料挂二级/三级；历史值不变时放行
                if item.category_id:
                    cat = db.get(BaseCategory, item.category_id)
                    if cat is None:
                        raise BizError(E_PARAM, f"分类 id={item.category_id} 不存在")
                    if cat.parent_id == 0 and item.category_id != product.category_id:
                        raise BizError(E_PARAM, "顶级分类仅作分组，材料请挂到其二级或三级子分类")
                    product.category_id = item.category_id
                amount = (qty * price).quantize(_DEC2)
                unit_name = item.unit_name or (db.get(BaseUnit, product.unit_id).name if product.unit_id else "")
                item_row = PchPurchaseInItem(
                    bill_id=bill.id, product_id=product.id, qty=qty, unit_name=unit_name,
                    price=price, amount=amount, location_id=item.location_id,
                    photo_file_id=item.photo_file_id, sort=idx,
                )
                db.add(item_row)
                db.flush()
                post_stock_change(
                    db,
                    product_id=product.id, warehouse_id=req.warehouse_id, location_id=item.location_id,
                    change_type="采购入库", bill_type="pch_purchase_in", bill_no=bill_no,
                    bill_item_id=item_row.id, qty_delta=qty, cost_price=price,
                    photo_file_id=item.photo_file_id, operator_id=user.id,
                )
                total_qty += qty
                total_amount += amount
            bill.total_qty = total_qty.quantize(Decimal("0.001"))
            bill.total_amount = total_amount.quantize(_DEC2)
            db.commit()
            # 关联采购计划单 → 按累计实际入库量推进状态（部分入库 / 已完成）
            if plan:
                _advance_plan_status(db, plan.id)
                db.commit()
            return ok({"id": bill.id, "bill_no": bill_no})
        except IntegrityError as exc:
            db.rollback()
            if not bill_no_conflict(exc):
                # 非单号冲突（外键/其他唯一约束）：如实报错，不掩盖真实原因
                raise BizError(E_PARAM, "入库单保存失败，请重试（详情见系统日志）") from exc
            continue  # 单号冲突，换号重试
    raise BizError(E_PARAM, "单据编号生成失败，请重试")


@router.get(
    "/purchase-in/history-price",
    dependencies=[Depends(require_any_permission("stk:query", "pch:in", "pch:ocr"))],
)
def purchase_in_history_price(
    product_id: int = Query(0, description="材料 id（0=全部材料，历史价格管理页用）"),
    keyword: str = Query("", max_length=100, description="材料名称/编码/物料编码模糊查询"),
    supplier_id: int = Query(0, description="供应商 id（0=全部供应商）"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    """历史采购价格：按入库日期倒序分页。

    - 入库/OCR 录入参考：传 product_id 查某材料最近价格（可带 supplier_id 过滤）
    - 历史价格管理页：product_id=0 + keyword/supplier_id 全量查询，展示「谁供的货」
    """
    if product_id and db.get(BaseProduct, product_id) is None:
        raise BizError(E_NOT_FOUND, "材料不存在")
    stmt = (
        select(PchPurchaseInItem, PchPurchaseIn)
        .join(PchPurchaseIn, PchPurchaseIn.id == PchPurchaseInItem.bill_id)
        .join(BaseProduct, BaseProduct.id == PchPurchaseInItem.product_id)
        .where(PchPurchaseIn.status == 1)
    )
    if product_id:
        stmt = stmt.where(PchPurchaseInItem.product_id == product_id)
    if supplier_id:
        stmt = stmt.where(PchPurchaseIn.supplier_id == supplier_id)
    if keyword:
        like = f"%{keyword.strip()}%"
        stmt = stmt.where(
            or_(BaseProduct.name.like(like), BaseProduct.code.like(like), BaseProduct.material_code.like(like), BaseProduct.spec.like(like))
        )
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.execute(
        stmt.order_by(PchPurchaseIn.bill_date.desc(), PchPurchaseIn.id.desc())
        .offset((page - 1) * page_size).limit(page_size)
    ).all()
    return ok({
        "list": [
            {
                "bill_no": b.bill_no,
                "bill_date": b.bill_date,
                "price": format(it.price, "f"),
                "qty": format(it.qty, "f"),
                "amount": format(it.amount, "f"),
                "supplier_id": b.supplier_id,
                "supplier_name": (s.name if (s := db.get(BaseSupplier, b.supplier_id)) else ""),
                "unit_name": ((u.name if (u := db.get(BaseUnit, p.unit_id)) else it.unit_name) if (p := db.get(BaseProduct, it.product_id)) else it.unit_name),
                "product_id": it.product_id,
                "product_name": (p.name if (p := db.get(BaseProduct, it.product_id)) else ""),
                "material_code": (p.material_code if (p := db.get(BaseProduct, it.product_id)) else ""),
                "spec": (p.spec if (p := db.get(BaseProduct, it.product_id)) else ""),
            }
            for it, b in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    })


@router.get("/purchase-in")
def list_purchase_in(
    bill_no: str = Query("", max_length=30),
    supplier_id: int = Query(0),
    warehouse_id: int = Query(0),
    start: str = Query(""),
    end: str = Query(""),
    status: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    stmt = select(PchPurchaseIn)
    if bill_no:
        stmt = stmt.where(PchPurchaseIn.bill_no.like(f"%{bill_no}%"))
    if supplier_id:
        stmt = stmt.where(PchPurchaseIn.supplier_id == supplier_id)
    if warehouse_id:
        stmt = stmt.where(PchPurchaseIn.warehouse_id == warehouse_id)
    if start:
        _check_date_arg(start, "start")
        stmt = stmt.where(PchPurchaseIn.bill_date >= f"{start} 00:00:00")
    if end:
        _check_date_arg(end, "end")
        stmt = stmt.where(PchPurchaseIn.bill_date <= f"{end} 23:59:59")
    if status is not None:
        stmt = stmt.where(PchPurchaseIn.status == status)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(PchPurchaseIn.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    return ok(PageData(
        list=[_purchase_out(db, b) for b in rows],
        total=total, page=page, page_size=page_size,
    ).model_dump())


@router.get("/purchase-in/{bill_id}")
def get_purchase_in(bill_id: int, db: Session = Depends(get_db)) -> dict:
    bill = db.get(PchPurchaseIn, bill_id)
    if bill is None:
        raise BizError(E_NOT_FOUND, "入库单不存在")
    return ok(_purchase_out(db, bill))


@router.post("/purchase-in/{bill_id}/void", dependencies=[Depends(require_permission("pch:in"))])
def void_purchase_in(
    bill_id: int,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    bill = db.get(PchPurchaseIn, bill_id)
    if bill is None:
        raise BizError(E_NOT_FOUND, "入库单不存在")
    if bill.status != 1:
        raise BizError(E_BILL_STATUS, "仅已入库单据可作废")
    if bill.bill_date.date() != datetime.now().date():
        raise BizError(E_BILL_STATUS, "仅当日单据可作废")
    bill.status = -1
    items = db.scalars(select(PchPurchaseInItem).where(PchPurchaseInItem.bill_id == bill.id)).all()
    for it in items:
        post_stock_change(
            db,
            product_id=it.product_id, warehouse_id=bill.warehouse_id, location_id=it.location_id,
            change_type="采购入库作废", bill_type="pch_purchase_in", bill_no=bill.bill_no,
            bill_item_id=it.id, qty_delta=-it.qty, cost_price=it.price,
            operator_id=user.id, remark="作废冲销",
        )
    if bill.plan_id:
        _advance_plan_status(db, bill.plan_id)  # 作废后实收减少，计划状态相应回退
    db.commit()
    return ok()


# ============================ 期初库存 ============================


def _opening_out(db: Session, bill: StkOpening) -> dict:
    items = db.scalars(select(StkOpeningItem).where(StkOpeningItem.bill_id == bill.id)).all()
    wh = db.get(BaseWarehouse, bill.warehouse_id)
    return OpeningOut(
        id=bill.id, bill_no=bill.bill_no, warehouse_id=bill.warehouse_id,
        warehouse_name=wh.name if wh else "", status=bill.status, remark=bill.remark,
        creator_name=_user_name(db, bill.creator_id),
        items=[
            OpeningItemOut(
                id=it.id, product_id=it.product_id,
                product_name=(p.name if (p := db.get(BaseProduct, it.product_id)) else ""),
                code=(p.code if (p := db.get(BaseProduct, it.product_id)) else ""),
                location_id=it.location_id, location_code=_loc_code(db, it.location_id),
                qty=it.qty, cost_price=it.cost_price,
            ).model_dump()
            for it in items
        ],
    ).model_dump()


@router.post("/opening", dependencies=[Depends(require_permission("pch:in"))])
def create_opening(
    req: OpeningReq,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    if db.get(BaseWarehouse, req.warehouse_id) is None:
        raise BizError(E_PARAM, "仓库不存在")
    for attempt in range(5):  # 单号并发冲突重试
        bill_no = generate_bill_no(db, "QCK", StkOpening)
        bill = StkOpening(bill_no=bill_no, warehouse_id=req.warehouse_id, status=0, remark=req.remark, creator_id=user.id)
        db.add(bill)
        db.flush()
        try:
            for idx, item in enumerate(req.items):
                if db.get(BaseProduct, item.product_id) is None:
                    raise BizError(E_NOT_FOUND, f"商品 id={item.product_id} 不存在")
                if db.get(BaseLocation, item.location_id) is None:
                    raise BizError(E_NOT_FOUND, f"库位 id={item.location_id} 不存在")
                db.add(StkOpeningItem(
                    bill_id=bill.id, product_id=item.product_id, location_id=item.location_id,
                    qty=_parse_dec(item.qty, "数量"), cost_price=_parse_dec(item.cost_price, "成本价"),
                ))
            db.commit()
            return ok({"id": bill.id, "bill_no": bill_no})
        except IntegrityError as exc:
            db.rollback()
            if not bill_no_conflict(exc):
                raise BizError(E_PARAM, "期初单保存失败，请重试（详情见系统日志）") from exc
    raise BizError(E_PARAM, "单据编号生成失败，请重试")


@router.get("/opening")
def list_opening(
    status: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    stmt = select(StkOpening)
    if status is not None:
        stmt = stmt.where(StkOpening.status == status)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(StkOpening.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    return ok(PageData(list=[_opening_out(db, b) for b in rows], total=total, page=page, page_size=page_size).model_dump())


@router.get("/opening/{bill_id}")
def get_opening(bill_id: int, db: Session = Depends(get_db)) -> dict:
    bill = db.get(StkOpening, bill_id)
    if bill is None:
        raise BizError(E_NOT_FOUND, "期初单不存在")
    return ok(_opening_out(db, bill))


@router.put("/opening/{bill_id}", dependencies=[Depends(require_permission("pch:in"))])
def update_opening(bill_id: int, req: OpeningReq, db: Session = Depends(get_db)) -> dict:
    bill = db.get(StkOpening, bill_id)
    if bill is None:
        raise BizError(E_NOT_FOUND, "期初单不存在")
    if bill.status != 0:
        raise BizError(E_BILL_STATUS, "仅草稿可修改")
    bill.warehouse_id = req.warehouse_id
    bill.remark = req.remark
    db.execute(StkOpeningItem.__table__.delete().where(StkOpeningItem.bill_id == bill.id))
    for item in req.items:
        if db.get(BaseProduct, item.product_id) is None:
            raise BizError(E_NOT_FOUND, f"商品 id={item.product_id} 不存在")
        db.add(StkOpeningItem(
            bill_id=bill.id, product_id=item.product_id, location_id=item.location_id,
            qty=_parse_dec(item.qty, "数量"), cost_price=_parse_dec(item.cost_price, "成本价"),
        ))
    db.commit()
    return ok()


@router.post("/opening/{bill_id}/post", dependencies=[Depends(require_permission("pch:in"))])
def post_opening(
    bill_id: int,
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    bill = db.get(StkOpening, bill_id)
    if bill is None:
        raise BizError(E_NOT_FOUND, "期初单不存在")
    if bill.status != 0:
        raise BizError(E_BILL_STATUS, "仅草稿可过账")
    items = db.scalars(select(StkOpeningItem).where(StkOpeningItem.bill_id == bill.id)).all()
    if not items:
        raise BizError(E_PARAM, "期初单明细为空")
    bill.status = 1
    for it in items:
        post_stock_change(
            db,
            product_id=it.product_id, warehouse_id=bill.warehouse_id, location_id=it.location_id,
            change_type="期初", bill_type="stk_opening", bill_no=bill.bill_no,
            bill_item_id=it.id, qty_delta=it.qty, cost_price=it.cost_price,
            operator_id=user.id,
        )
    db.commit()
    return ok()


@router.post("/opening/import", dependencies=[Depends(require_permission("pch:in"))])
async def import_opening(
    warehouse_id: int = Query(..., gt=0),
    file: UploadFile = File(...),
    user: SysUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Excel 导入期初明细（表头：商品编码/库位编码/数量/成本价），生成草稿。"""
    if db.get(BaseWarehouse, warehouse_id) is None:
        raise BizError(E_PARAM, "仓库不存在")
    data = await file.read()
    if len(data) > _MAX_IMPORT_BYTES:
        raise BizError(E_PARAM, "导入文件不能超过 10MB")
    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise BizError(E_PARAM, "文件为空")
    if len(rows) > _MAX_IMPORT_ROWS:
        raise BizError(E_PARAM, f"单次导入不能超过 {_MAX_IMPORT_ROWS} 行")
    headers = [str(h).strip() if h else "" for h in rows[0]]
    if headers[:4] != _OPENING_IMPORT_COLUMNS:
        raise BizError(E_PARAM, f"表头必须为：{'/'.join(_OPENING_IMPORT_COLUMNS)}")

    for attempt in range(5):  # 单号并发冲突重试
        bill = StkOpening(bill_no=generate_bill_no(db, "QCK", StkOpening), warehouse_id=warehouse_id, status=0, creator_id=user.id)
        db.add(bill)
        db.flush()
        success, fail_rows = 0, []
        for idx, row in enumerate(rows[1:], start=2):
            vals = [str(v).strip() if v is not None else "" for v in row] + [""] * 4
            product = db.scalar(select(BaseProduct).where(BaseProduct.code == vals[0])) if vals[0] else None
            location = db.scalar(select(BaseLocation).where(BaseLocation.code == vals[1])) if vals[1] else None
            if product is None:
                fail_rows.append({"row": idx, "reason": f"商品编码 {vals[0]} 不存在"})
                continue
            if location is None:
                fail_rows.append({"row": idx, "reason": f"库位编码 {vals[1]} 不存在"})
                continue
            if not _DECIMAL_RE.match(vals[2]) or Decimal(vals[2]) <= 0:
                fail_rows.append({"row": idx, "reason": "数量必须为正数"})
                continue
            cost = Decimal(vals[3]) if _DECIMAL_RE.match(vals[3]) else Decimal(0)
            db.add(StkOpeningItem(bill_id=bill.id, product_id=product.id, location_id=location.id, qty=Decimal(vals[2]), cost_price=cost))
            success += 1
        if success == 0:
            db.rollback()
            raise BizError(E_PARAM, "导入全部失败，未生成草稿")
        try:
            db.commit()
            return ok({"draft_id": bill.id, "bill_no": bill.bill_no, "success_count": success, "fail_rows": fail_rows})
        except IntegrityError as exc:
            db.rollback()
            if not bill_no_conflict(exc):
                raise BizError(E_PARAM, "期初导入保存失败，请重试（详情见系统日志）") from exc
    raise BizError(E_PARAM, "单据编号生成失败，请重试")


# ============================ 库存查询 / 流水 ============================


@router.get("/stock")
def list_stock(
    product_id: int = Query(0),
    warehouse_id: int = Query(0),
    location_id: int = Query(0),
    keyword: str = Query("", max_length=100),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
) -> dict:
    stmt = select(StkStock)
    if product_id:
        stmt = stmt.where(StkStock.product_id == product_id)
    if warehouse_id:
        stmt = stmt.where(StkStock.warehouse_id == warehouse_id)
    if location_id:
        stmt = stmt.where(StkStock.location_id == location_id)
    if keyword:
        like = f"%{keyword}%"
        # 子查询代替 IN(prod_ids)：避免无上限的关键字命中全量加载进内存
        prod_q = select(BaseProduct.id).where(or_(
            BaseProduct.name.like(like), BaseProduct.code.like(like),
            BaseProduct.barcode.like(like), BaseProduct.sku.like(like),
        ))
        stmt = stmt.where(StkStock.product_id.in_(prod_q))
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(StkStock.product_id, StkStock.location_id).offset((page - 1) * page_size).limit(page_size)).all()
    # 批量预取商品/仓库/库位，避免每行 3 次回表（N+1）
    pids = {s.product_id for s in rows}
    wids = {s.warehouse_id for s in rows}
    lids = {s.location_id for s in rows}
    prod_map = {p.id: p for p in db.scalars(select(BaseProduct).where(BaseProduct.id.in_(pids))).all()} if pids else {}
    wh_map = {w.id: w for w in db.scalars(select(BaseWarehouse).where(BaseWarehouse.id.in_(wids))).all()} if wids else {}
    loc_map = {l.id: l for l in db.scalars(select(BaseLocation).where(BaseLocation.id.in_(lids))).all()} if lids else {}
    out = []
    for s in rows:
        p = prod_map.get(s.product_id)
        wh = wh_map.get(s.warehouse_id)
        loc = loc_map.get(s.location_id)
        out.append(StockRow(
            product_id=s.product_id, product_name=p.name if p else "", code=p.code if p else "",
            material_code=p.material_code if p else "", barcode=p.barcode if p else "", spec=p.spec if p else "",
            warehouse_id=s.warehouse_id, warehouse_name=wh.name if wh else "",
            location_id=s.location_id, location_code=_loc_code(db, s.location_id),
            qty=s.qty, cost_price=s.cost_price, amount=(s.qty * s.cost_price).quantize(_DEC2),
        ).model_dump())
    return ok(PageData(list=out, total=total, page=page, page_size=page_size).model_dump())


@router.get("/stock/flow")
def list_stock_flow(
    product_id: int = Query(0),
    bill_no: str = Query("", max_length=30),
    change_type: str = Query("", max_length=20),
    start: str = Query(""),
    end: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
) -> dict:
    stmt = select(StkStockLog)
    if product_id:
        stmt = stmt.where(StkStockLog.product_id == product_id)
    if bill_no:
        stmt = stmt.where(StkStockLog.bill_no.like(f"%{bill_no}%"))
    if change_type:
        stmt = stmt.where(StkStockLog.change_type == change_type)
    if start:
        _check_date_arg(start, "start")
        stmt = stmt.where(StkStockLog.created_at >= f"{start} 00:00:00")
    if end:
        _check_date_arg(end, "end")
        stmt = stmt.where(StkStockLog.created_at <= f"{end} 23:59:59")
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(StkStockLog.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    # 批量预取商品/仓库/库位/操作人，避免每行 3+1 次回表（N+1）
    pids = {log.product_id for log in rows}
    wids = {log.warehouse_id for log in rows}
    lids = {log.location_id for log in rows}
    uids = {log.operator_id for log in rows if log.operator_id}
    prod_map = {p.id: p for p in db.scalars(select(BaseProduct).where(BaseProduct.id.in_(pids))).all()} if pids else {}
    wh_map = {w.id: w for w in db.scalars(select(BaseWarehouse).where(BaseWarehouse.id.in_(wids))).all()} if wids else {}
    loc_map = {l.id: l for l in db.scalars(select(BaseLocation).where(BaseLocation.id.in_(lids))).all()} if lids else {}
    user_map = {u.id: u for u in db.scalars(select(SysUser).where(SysUser.id.in_(uids))).all()} if uids else {}
    out = []
    for log in rows:
        p = prod_map.get(log.product_id)
        wh = wh_map.get(log.warehouse_id)
        op = user_map.get(log.operator_id)
        out.append(StockFlowRow(
            id=log.id, product_id=log.product_id, product_name=p.name if p else "", code=p.code if p else "",
            warehouse_name=wh.name if wh else "", location_code=_loc_code(db, log.location_id),
            change_type=log.change_type, bill_no=log.bill_no,
            before_qty=log.before_qty, change_qty=log.change_qty, after_qty=log.after_qty,
            cost_price=log.cost_price, operator_name=op.real_name if op else "",
            remark=log.remark, created_at=log.created_at,
        ).model_dump())
    return ok(PageData(list=out, total=total, page=page, page_size=page_size).model_dump())


@router.get("/stock/location-summary")
def location_summary(
    warehouse_id: int = Query(0),
    shelf_id: int = Query(0),
    db: Session = Depends(get_db),
) -> dict:
    """2D 货架图数据源：仓库/货架下每个库位的商品库存 + 预警状态（绿=正常/红=低/黄=高）。

    与 /locations 的区别：本接口一次返回库位上的库存明细（商品+数量+预警），
    避免货架图 N+1 请求。60 秒 TTL 缓存，库存变动（post_stock_change）即时失效。
    """
    def _load() -> list[dict]:
        stmt = select(BaseLocation)
        if warehouse_id:
            stmt = stmt.where(BaseLocation.warehouse_id == warehouse_id)
        if shelf_id:
            stmt = stmt.where(BaseLocation.shelf_id == shelf_id)
        locations = db.scalars(stmt.order_by(BaseLocation.layer_no, BaseLocation.row_no, BaseLocation.col_no)).all()
        if not locations:
            return []
        loc_ids = [loc.id for loc in locations]

        stocks = db.scalars(
            select(StkStock).where(StkStock.location_id.in_(loc_ids), StkStock.qty != 0)
        ).all()
        prod_ids = {s.product_id for s in stocks}
        products = {p.id: p for p in db.scalars(select(BaseProduct).where(BaseProduct.id.in_(prod_ids))).all()}

        by_loc: dict[int, list[dict]] = {}
        for s in stocks:
            p = products.get(s.product_id)
            if p is None:
                continue
            if p.min_stock and s.qty < p.min_stock:
                alert = "low"
            elif p.max_stock and s.qty > p.max_stock:
                alert = "high"
            else:
                alert = "normal"
            by_loc.setdefault(s.location_id, []).append({
                "product_id": s.product_id,
                "code": p.code,
                "name": p.name,
                "spec": p.spec,
                "qty": _fmt_qty(s.qty),
                "min_stock": _fmt_qty(p.min_stock),
                "max_stock": _fmt_qty(p.max_stock),
                "alert": alert,
            })
        return [
            {
                "location_id": loc.id,
                "location_code": _loc_code(db, loc.id),
                "layer_no": loc.layer_no,
                "row_no": loc.row_no,
                "col_no": loc.col_no,
                "shelf_id": loc.shelf_id,
                "items": by_loc.get(loc.id, []),
            }
            for loc in locations
        ]

    return ok(cache_aside_json(f"stock:locsum:{warehouse_id}:{shelf_id}", 60, _load))


# ============================ 采购计划单 ============================
# 事物流：采购计划单 → 材料入库（关联 plan_id，送货单图片可选存底）→ 库存落账。
# 计划状态自动推进：0 草稿 → 1 已提交 → 2 部分入库 / 3 已完成（按计划明细累计实收）→ -1 作废。


def _plan_received_map(db: Session, plan_id: int) -> dict[int, Decimal]:
    """计划明细 id → 已累计入库数量（按 product_id 匹配该计划下已入库单明细求和）。"""
    plan_items = db.scalars(select(PchPurchasePlanItem).where(PchPurchasePlanItem.plan_id == plan_id)).all()
    result: dict[int, Decimal] = {pi.id: Decimal(0) for pi in plan_items}
    bill_ids = list(db.scalars(
        select(PchPurchaseIn.id).where(PchPurchaseIn.plan_id == plan_id, PchPurchaseIn.status == 1)
    ).all())
    if bill_ids:
        rows = db.execute(
            select(PchPurchaseInItem.product_id, func.sum(PchPurchaseInItem.qty))
            .where(PchPurchaseInItem.bill_id.in_(bill_ids))
            .group_by(PchPurchaseInItem.product_id)
        ).all()
        by_product = {pid: qty for pid, qty in rows}
        for pi in plan_items:
            result[pi.id] = by_product.get(pi.product_id, Decimal(0))
    return result


def _advance_plan_status(db: Session, plan_id: int) -> None:
    """按计划明细累计实收推进状态：全部完成 → 已完成(3)；有部分实收 → 部分入库(2)。"""
    plan = db.get(PchPurchasePlan, plan_id)
    if plan is None or plan.status not in (1, 2):
        return
    items = db.scalars(select(PchPurchasePlanItem).where(PchPurchasePlanItem.plan_id == plan.id)).all()
    if not items:
        return
    received = _plan_received_map(db, plan.id)
    total_planned = sum((it.planned_qty for it in items), Decimal(0))
    total_received = sum((received.get(it.id, Decimal(0)) for it in items), Decimal(0))
    if total_planned > 0 and total_received >= total_planned:
        plan.status = 3
    elif total_received > 0:
        plan.status = 2


def _plan_out(db: Session, plan: PchPurchasePlan) -> dict:
    items = db.scalars(select(PchPurchasePlanItem).where(PchPurchasePlanItem.plan_id == plan.id).order_by(PchPurchasePlanItem.sort)).all()
    wh = db.get(BaseWarehouse, plan.warehouse_id)
    sup = db.get(BaseSupplier, plan.supplier_id)
    received = _plan_received_map(db, plan.id)
    return PurchasePlanOut(
        id=plan.id, bill_no=plan.bill_no, supplier_id=plan.supplier_id, supplier_name=sup.name if sup else "",
        warehouse_id=plan.warehouse_id, warehouse_name=wh.name if wh else "",
        status=plan.status, total_qty=plan.total_qty, total_amount=plan.total_amount,
        plan_date=plan.plan_date, remark=plan.remark, creator_name=_user_name(db, plan.creator_id),
        items=[
            PurchasePlanItemOut(
                id=it.id, product_id=it.product_id,
                product_name=(p.name if (p := db.get(BaseProduct, it.product_id)) else ""),
                code=(p.code if (p := db.get(BaseProduct, it.product_id)) else ""),
                planned_qty=it.planned_qty, unit_name=it.unit_name, est_price=it.est_price, amount=it.amount,
                remark=it.remark, received_qty=received.get(it.id, Decimal(0)),
            )
            for it in items
        ],
    ).model_dump()


def _apply_plan_items(db: Session, plan: PchPurchasePlan, items: list[PurchasePlanItemReq]) -> None:
    """重建计划明细（事务内调用，调用方负责 commit）。"""
    db.execute(PchPurchasePlanItem.__table__.delete().where(PchPurchasePlanItem.plan_id == plan.id))
    total_qty, total_amount = Decimal(0), Decimal(0)
    for idx, item in enumerate(items):
        product = db.get(BaseProduct, item.product_id)
        if product is None:
            raise BizError(E_NOT_FOUND, f"商品 id={item.product_id} 不存在")
        qty = _parse_dec(item.planned_qty, "计划数量")
        est = _parse_dec(item.est_price, "预计单价")
        if qty <= 0:
            raise BizError(E_PARAM, "计划数量必须大于 0")
        if est < 0:
            raise BizError(E_PARAM, "预计单价不能为负数")
        amount = (qty * est).quantize(_DEC2)
        unit_name = item.unit_name or (db.get(BaseUnit, product.unit_id).name if product.unit_id else "")
        db.add(PchPurchasePlanItem(
            plan_id=plan.id, product_id=product.id, planned_qty=qty, unit_name=unit_name,
            est_price=est, amount=amount, remark=item.remark.strip(), sort=idx,
        ))
        total_qty += qty
        total_amount += amount
    plan.total_qty = total_qty.quantize(Decimal("0.001"))
    plan.total_amount = total_amount.quantize(_DEC2)


@router.post("/purchase-plans", dependencies=[Depends(require_permission("pch:in"))])
def create_purchase_plan(req: PurchasePlanReq, user: SysUser = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
    """新建采购计划单（草稿）：不动库存，到货后由它生成材料入库。"""
    if db.get(BaseWarehouse, req.warehouse_id) is None:
        raise BizError(E_PARAM, "仓库不存在")
    if req.supplier_id and db.get(BaseSupplier, req.supplier_id) is None:
        raise BizError(E_PARAM, "供应商不存在")
    for attempt in range(5):
        bill_no = generate_bill_no(db, "JH", PchPurchasePlan)
        plan = PchPurchasePlan(
            bill_no=bill_no, supplier_id=req.supplier_id, warehouse_id=req.warehouse_id,
            status=0, plan_date=req.plan_date or datetime.now(), remark=req.remark.strip(),
            creator_id=user.id,
        )
        db.add(plan)
        db.flush()
        try:
            _apply_plan_items(db, plan, req.items)
            db.commit()
            return ok(_plan_out(db, db.get(PchPurchasePlan, plan.id)))
        except IntegrityError as exc:
            db.rollback()
            if not bill_no_conflict(exc):
                raise BizError(E_PARAM, "采购计划单保存失败，请重试（详情见系统日志）") from exc
            continue
    raise BizError(E_PARAM, "单据编号生成失败，请重试")


@router.get("/purchase-plans")
def list_purchase_plans(
    bill_no: str = Query("", max_length=30),
    supplier_id: int = Query(0),
    warehouse_id: int = Query(0),
    status: int | None = Query(None),
    start: str = Query(""),
    end: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    stmt = select(PchPurchasePlan)
    if bill_no:
        stmt = stmt.where(PchPurchasePlan.bill_no.like(f"%{bill_no}%"))
    if supplier_id:
        stmt = stmt.where(PchPurchasePlan.supplier_id == supplier_id)
    if warehouse_id:
        stmt = stmt.where(PchPurchasePlan.warehouse_id == warehouse_id)
    if start:
        _check_date_arg(start, "start")
        stmt = stmt.where(PchPurchasePlan.plan_date >= f"{start} 00:00:00")
    if end:
        _check_date_arg(end, "end")
        stmt = stmt.where(PchPurchasePlan.plan_date <= f"{end} 23:59:59")
    if status is not None:
        stmt = stmt.where(PchPurchasePlan.status == status)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(stmt.order_by(PchPurchasePlan.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    return ok(PageData(
        list=[_plan_out(db, p) for p in rows],
        total=total, page=page, page_size=page_size,
    ).model_dump())


@router.get("/purchase-plans/{plan_id}")
def get_purchase_plan(plan_id: int, db: Session = Depends(get_db)) -> dict:
    plan = db.get(PchPurchasePlan, plan_id)
    if plan is None:
        raise BizError(E_NOT_FOUND, "采购计划单不存在")
    return ok(_plan_out(db, plan))


@router.put("/purchase-plans/{plan_id}", dependencies=[Depends(require_permission("pch:in"))])
def update_purchase_plan(plan_id: int, req: PurchasePlanReq, db: Session = Depends(get_db)) -> dict:
    """编辑采购计划单（仅草稿可编辑；已提交后不允许改计划）。"""
    plan = db.get(PchPurchasePlan, plan_id)
    if plan is None:
        raise BizError(E_NOT_FOUND, "采购计划单不存在")
    if plan.status != 0:
        raise BizError(E_BILL_STATUS, "仅草稿状态的采购计划单可编辑")
    if db.get(BaseWarehouse, req.warehouse_id) is None:
        raise BizError(E_PARAM, "仓库不存在")
    if req.supplier_id and db.get(BaseSupplier, req.supplier_id) is None:
        raise BizError(E_PARAM, "供应商不存在")
    plan.supplier_id = req.supplier_id
    plan.warehouse_id = req.warehouse_id
    plan.plan_date = req.plan_date or plan.plan_date
    plan.remark = req.remark.strip()
    _apply_plan_items(db, plan, req.items)
    db.commit()
    return ok(_plan_out(db, db.get(PchPurchasePlan, plan.id)))


@router.post("/purchase-plans/{plan_id}/submit", dependencies=[Depends(require_permission("pch:in"))])
def submit_purchase_plan(plan_id: int, db: Session = Depends(get_db)) -> dict:
    """提交采购计划单（草稿 → 已提交，之后可据此入库）。"""
    plan = db.get(PchPurchasePlan, plan_id)
    if plan is None:
        raise BizError(E_NOT_FOUND, "采购计划单不存在")
    if plan.status != 0:
        raise BizError(E_BILL_STATUS, "仅草稿状态的采购计划单可提交")
    plan.status = 1
    db.commit()
    return ok(_plan_out(db, db.get(PchPurchasePlan, plan.id)))


@router.post("/purchase-plans/{plan_id}/void", dependencies=[Depends(require_permission("pch:in"))])
def void_purchase_plan(plan_id: int, db: Session = Depends(get_db)) -> dict:
    """作废采购计划单（未完成/未作废状态可作废；已完成的不可）。"""
    plan = db.get(PchPurchasePlan, plan_id)
    if plan is None:
        raise BizError(E_NOT_FOUND, "采购计划单不存在")
    if plan.status in (-1, 3):
        raise BizError(E_BILL_STATUS, "已作废或已完成的采购计划单不可作废")
    plan.status = -1
    db.commit()
    return ok(_plan_out(db, db.get(PchPurchasePlan, plan.id)))
